#!/usr/bin/env python3
"""Regression tests for existing features."""
import sys
import requests
from io import BytesIO
from openpyxl import load_workbook

BASE_URL = "https://github-auto-build.preview.emergentagent.com/api"

class RegressionTestRunner:
    def __init__(self):
        self.tests_run = 0
        self.tests_passed = 0
        self.admin_token = None
        self.siswa_token = None

    def log(self, msg, level="INFO"):
        prefix = {"INFO": "ℹ️", "PASS": "✅", "FAIL": "❌", "WARN": "⚠️"}.get(level, "•")
        print(f"{prefix} {msg}")

    def test(self, name, fn):
        """Run a test function."""
        self.tests_run += 1
        self.log(f"Testing: {name}")
        try:
            fn()
            self.tests_passed += 1
            self.log(f"PASSED: {name}", "PASS")
            return True
        except AssertionError as e:
            self.log(f"FAILED: {name} - {e}", "FAIL")
            return False
        except Exception as e:
            self.log(f"ERROR: {name} - {e}", "FAIL")
            return False

    def login(self, email, password):
        """Login and return token."""
        r = requests.post(f"{BASE_URL}/auth/login", json={"email": email, "password": password})
        assert r.status_code == 200, f"Login failed: {r.status_code} {r.text}"
        return r.json()["token"]

    def _test_login_admin(self):
        self.admin_token = self.login("hitoria532@gmail.com", "admin123")
        assert self.admin_token, "Admin token is empty"

    def _test_login_siswa(self):
        self.siswa_token = self.login("siswa@sekolah.id", "siswa123")
        assert self.siswa_token, "Siswa token is empty"

    def _test_session_excel_export(self):
        """Test session results Excel export"""
        # Get sessions
        r = requests.get(f"{BASE_URL}/sessions", headers={"Authorization": f"Bearer {self.admin_token}"})
        assert r.status_code == 200, f"Get sessions failed: {r.status_code}"
        sessions = r.json()
        
        if not sessions:
            self.log("No sessions found, skipping Excel export test", "WARN")
            return
        
        # Find a session with attempts
        session_with_attempts = None
        for session in sessions:
            r = requests.get(f"{BASE_URL}/results/session/{session['id']}", 
                           headers={"Authorization": f"Bearer {self.admin_token}"})
            if r.status_code == 200:
                data = r.json()
                if data.get('attempts') and len(data['attempts']) > 0:
                    session_with_attempts = session
                    break
        
        if not session_with_attempts:
            self.log("No sessions with attempts found, skipping Excel export test", "WARN")
            return
        
        # Test Excel export endpoint
        r = requests.get(f"{BASE_URL}/export/session/{session_with_attempts['id']}/xlsx",
                        headers={"Authorization": f"Bearer {self.admin_token}"})
        assert r.status_code == 200, f"Excel export failed: {r.status_code}"
        assert r.headers.get('Content-Type') == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', \
            "Wrong content type for Excel"
        
        # Verify it's a valid Excel file
        wb = load_workbook(BytesIO(r.content))
        assert len(wb.sheetnames) > 0, "Excel file has no sheets"
        self.log(f"Excel export successful with {len(wb.sheetnames)} sheet(s)")

    def _test_kartu_peserta_pdf(self):
        """Test student result card PDF generation"""
        # Get student's attempts
        r = requests.get(f"{BASE_URL}/results/me", headers={"Authorization": f"Bearer {self.siswa_token}"})
        assert r.status_code == 200, f"Get results failed: {r.status_code}"
        attempts = r.json()
        
        if not attempts:
            self.log("No attempts found for siswa, skipping PDF test", "WARN")
            return
        
        # Get PDF for first attempt
        attempt_id = attempts[0]['id']
        r = requests.get(f"{BASE_URL}/results/detail/{attempt_id}/pdf",
                        headers={"Authorization": f"Bearer {self.siswa_token}"})
        assert r.status_code == 200, f"PDF generation failed: {r.status_code}"
        assert r.headers.get('Content-Type') == 'application/pdf', "Wrong content type for PDF"
        assert len(r.content) > 1000, "PDF file seems too small"
        self.log(f"PDF generated successfully ({len(r.content)} bytes)")

    def _test_student_report_pdf(self):
        """Test student comprehensive report PDF"""
        r = requests.get(f"{BASE_URL}/report/student/me/pdf",
                        headers={"Authorization": f"Bearer {self.siswa_token}"})
        assert r.status_code == 200, f"Student report PDF failed: {r.status_code}"
        assert r.headers.get('Content-Type') == 'application/pdf', "Wrong content type for PDF"
        assert len(r.content) > 1000, "PDF file seems too small"
        self.log(f"Student report PDF generated successfully ({len(r.content)} bytes)")

    def _test_retake_flow_basic(self):
        """Test basic retake flow (max_attempts > 1)"""
        # Get sessions as siswa
        r = requests.get(f"{BASE_URL}/sessions", headers={"Authorization": f"Bearer {self.siswa_token}"})
        assert r.status_code == 200, f"Get sessions failed: {r.status_code}"
        sessions = r.json()
        
        # Find a session with max_attempts > 1
        retake_session = None
        for session in sessions:
            if session.get('max_attempts', 1) > 1:
                retake_session = session
                break
        
        if not retake_session:
            self.log("No retake sessions found, skipping retake test", "WARN")
            return
        
        # Verify session has retake info
        assert 'attempts_left' in retake_session, "attempts_left field missing"
        assert 'attempts_used' in retake_session, "attempts_used field missing"
        assert 'score_policy' in retake_session, "score_policy field missing"
        
        self.log(f"Retake session found: {retake_session['title']} " +
                f"(used: {retake_session['attempts_used']}, left: {retake_session['attempts_left']}, " +
                f"policy: {retake_session['score_policy']})")

    def _test_essay_grading_endpoint(self):
        """Test essay grading endpoint exists and works"""
        # Get sessions with attempts
        r = requests.get(f"{BASE_URL}/sessions", headers={"Authorization": f"Bearer {self.admin_token}"})
        assert r.status_code == 200
        sessions = r.json()
        
        # Find an attempt that needs grading
        attempt_needing_grading = None
        for session in sessions:
            r = requests.get(f"{BASE_URL}/results/session/{session['id']}", 
                           headers={"Authorization": f"Bearer {self.admin_token}"})
            if r.status_code == 200:
                data = r.json()
                for attempt in data.get('attempts', []):
                    if attempt.get('needs_grading'):
                        attempt_needing_grading = attempt
                        break
                if attempt_needing_grading:
                    break
        
        if not attempt_needing_grading:
            self.log("No attempts needing grading found, testing endpoint only", "WARN")
            # Just verify the endpoint exists by trying with a fake ID
            r = requests.post(f"{BASE_URL}/results/grade/fake-id", 
                            json={"scores": {}},
                            headers={"Authorization": f"Bearer {self.admin_token}"})
            # Should get 404 (not found) not 405 (method not allowed) or 404 (route not found)
            assert r.status_code in [404, 400], f"Essay grading endpoint issue: {r.status_code}"
            return
        
        # Test grading with empty scores (should work)
        r = requests.post(f"{BASE_URL}/results/grade/{attempt_needing_grading['id']}", 
                         json={"scores": {}},
                         headers={"Authorization": f"Bearer {self.admin_token}"})
        assert r.status_code == 200, f"Essay grading failed: {r.status_code}"
        self.log("Essay grading endpoint working")

    def _test_leaderboard_endpoint(self):
        """Test leaderboard endpoint"""
        # Get sessions
        r = requests.get(f"{BASE_URL}/sessions", headers={"Authorization": f"Bearer {self.admin_token}"})
        assert r.status_code == 200
        sessions = r.json()
        
        if not sessions:
            self.log("No sessions found, skipping leaderboard test", "WARN")
            return
        
        # Test leaderboard for first session
        session_id = sessions[0]['id']
        r = requests.get(f"{BASE_URL}/leaderboard/{session_id}",
                        headers={"Authorization": f"Bearer {self.admin_token}"})
        
        # Endpoint should exist (200 or 404 if no data, but not 405)
        assert r.status_code in [200, 404], f"Leaderboard endpoint issue: {r.status_code}"
        
        if r.status_code == 200:
            data = r.json()
            assert isinstance(data, list), "Leaderboard should return a list"
            self.log(f"Leaderboard working ({len(data)} entries)")
        else:
            self.log("Leaderboard endpoint exists but no data", "WARN")

    def run_all_tests(self):
        """Run all regression tests."""
        self.log("=" * 70)
        self.log("REGRESSION TESTS - Existing Features")
        self.log("=" * 70)

        # Authentication
        self.log("\n--- Authentication ---")
        self.test("Login as admin", self._test_login_admin)
        self.test("Login as siswa", self._test_login_siswa)

        # Regression Tests
        self.log("\n--- Regression Tests ---")
        self.test("Session Excel export", self._test_session_excel_export)
        self.test("Kartu peserta PDF generation", self._test_kartu_peserta_pdf)
        self.test("Student report PDF generation", self._test_student_report_pdf)
        self.test("Retake flow (max_attempts)", self._test_retake_flow_basic)
        self.test("Essay grading endpoint", self._test_essay_grading_endpoint)
        self.test("Leaderboard endpoint", self._test_leaderboard_endpoint)

        # Summary
        self.log("\n" + "=" * 70)
        self.log(f"RESULTS: {self.tests_passed}/{self.tests_run} tests passed")
        self.log("=" * 70)
        
        return 0 if self.tests_passed == self.tests_run else 1


if __name__ == "__main__":
    runner = RegressionTestRunner()
    sys.exit(runner.run_all_tests())
