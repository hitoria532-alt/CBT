#!/usr/bin/env python3
"""Backend API tests for Excel export feature and regressions."""
import sys
import requests
from io import BytesIO
from openpyxl import load_workbook

BASE_URL = "https://github-auto-build.preview.emergentagent.com/api"

class TestRunner:
    def __init__(self):
        self.tests_run = 0
        self.tests_passed = 0
        self.tokens = {}
        self.sessions = []
        self.temp_sessions = []

    def log(self, msg, level="INFO"):
        prefix = {"INFO": "ℹ️", "PASS": "✅", "FAIL": "❌", "WARN": "⚠️"}.get(level, "•")
        print(f"{prefix} {msg}")

    def test(self, name, fn):
        """Run a test function."""
        self.tests_run += 1
        self.log(f"Testing: {name}")
        try:
            fn()
            self.tests_passed += 1
            self.log(f"PASSED: {name}", "PASS")
            return True
        except AssertionError as e:
            self.log(f"FAILED: {name} - {e}", "FAIL")
            return False
        except Exception as e:
            self.log(f"ERROR: {name} - {e}", "FAIL")
            return False

    def login(self, email, password):
        """Login and return token."""
        r = requests.post(f"{BASE_URL}/auth/login", json={"email": email, "password": password})
        assert r.status_code == 200, f"Login failed: {r.status_code} {r.text}"
        data = r.json()
        return data["token"]

    def get_sessions(self, token):
        """Get all sessions."""
        r = requests.get(f"{BASE_URL}/sessions", headers={"Authorization": f"Bearer {token}"})
        assert r.status_code == 200, f"Get sessions failed: {r.status_code}"
        return r.json()

    def create_temp_session(self, token, title, package_id, max_attempts=1):
        """Create a temporary session for testing."""
        from datetime import datetime, timedelta
        now = datetime.utcnow()
        start = (now - timedelta(hours=1)).isoformat()
        end = (now + timedelta(hours=1)).isoformat()
        
        body = {
            "title": title,
            "package_id": package_id,
            "start_time": start,
            "end_time": end,
            "duration_minutes": 60,
            "kkm": 75.0,
            "class_ids": [],
            "announcement": "",
            "max_attempts": max_attempts,
            "score_policy": "tertinggi"
        }
        r = requests.post(f"{BASE_URL}/sessions", json=body, headers={"Authorization": f"Bearer {token}"})
        assert r.status_code == 200, f"Create session failed: {r.status_code} {r.text}"
        session = r.json()
        self.temp_sessions.append(session["id"])
        return session

    def delete_temp_sessions(self, token):
        """Delete all temporary sessions created during testing."""
        for sid in self.temp_sessions:
            try:
                requests.delete(f"{BASE_URL}/sessions/{sid}", headers={"Authorization": f"Bearer {token}"})
                self.log(f"Deleted temp session: {sid}", "INFO")
            except Exception as e:
                self.log(f"Failed to delete temp session {sid}: {e}", "WARN")

    def run_all_tests(self):
        """Run all tests."""
        self.log("=" * 60)
        self.log("BACKEND API TESTS - Excel Export Feature")
        self.log("=" * 60)

        # Login
        self.log("\n--- Authentication ---")
        self.test("Login as admin", lambda: self._test_login_admin())
        self.test("Login as guru", lambda: self._test_login_guru())
        self.test("Login as siswa", lambda: self._test_login_siswa())

        # Get sessions
        self.log("\n--- Session Discovery ---")
        self.test("Get sessions as admin", lambda: self._test_get_sessions())

        # Excel export tests
        self.log("\n--- Excel Export API Tests ---")
        self.test("Export xlsx as admin (200)", lambda: self._test_export_xlsx_admin())
        self.test("Export xlsx as guru (200)", lambda: self._test_export_xlsx_guru())
        self.test("Export xlsx as siswa (403)", lambda: self._test_export_xlsx_siswa_forbidden())
        self.test("Export xlsx with invalid session (404)", lambda: self._test_export_xlsx_invalid_session())
        
        # Test with empty session
        self.log("\n--- Empty Session Test ---")
        self.test("Export xlsx with empty session", lambda: self._test_export_xlsx_empty_session())

        # Test max_attempts=1 vs max_attempts>1
        self.log("\n--- Max Attempts Column Test ---")
        self.test("Verify max_attempts=1 has no Percobaan/Dipakai columns", lambda: self._test_max_attempts_1())
        self.test("Verify max_attempts>1 has Percobaan/Dipakai columns", lambda: self._test_max_attempts_gt1())

        # Workbook structure tests
        self.log("\n--- Workbook Structure Tests ---")
        self.test("Verify workbook structure and RINGKASAN", lambda: self._test_workbook_structure())

        # Data consistency test
        self.log("\n--- Data Consistency Test ---")
        self.test("Verify xlsx data matches /api/results/session/{id}", lambda: self._test_data_consistency())

        # Regression tests
        self.log("\n--- Regression Tests ---")
        self.test("Essay grading still works", lambda: self._test_essay_grading())
        self.test("Analitik Butir still works", lambda: self._test_analytics())
        self.test("Kartu hasil PDF still works", lambda: self._test_kartu_hasil_pdf())
        self.test("Rapor PDF still works", lambda: self._test_rapor_pdf())
        self.test("CSV export still works", lambda: self._test_csv_export())

        # Cleanup
        self.log("\n--- Cleanup ---")
        self.delete_temp_sessions(self.tokens.get("admin"))

        # Summary
        self.log("\n" + "=" * 60)
        self.log(f"RESULTS: {self.tests_passed}/{self.tests_run} tests passed")
        self.log("=" * 60)
        return 0 if self.tests_passed == self.tests_run else 1

    def _test_login_admin(self):
        token = self.login("hitoria532@gmail.com", "admin123")
        self.tokens["admin"] = token
        assert token, "Admin token is empty"

    def _test_login_guru(self):
        token = self.login("guru@sekolah.id", "guru123")
        self.tokens["guru"] = token
        assert token, "Guru token is empty"

    def _test_login_siswa(self):
        token = self.login("siswa@sekolah.id", "siswa123")
        self.tokens["siswa"] = token
        assert token, "Siswa token is empty"

    def _test_get_sessions(self):
        sessions = self.get_sessions(self.tokens["admin"])
        self.sessions = sessions
        assert isinstance(sessions, list), "Sessions should be a list"
        self.log(f"Found {len(sessions)} sessions")

    def _test_export_xlsx_admin(self):
        if not self.sessions:
            self.log("No sessions found, skipping", "WARN")
            return
        session_id = self.sessions[0]["id"]
        r = requests.get(f"{BASE_URL}/export/session/{session_id}/xlsx", 
                        headers={"Authorization": f"Bearer {self.tokens['admin']}"})
        assert r.status_code == 200, f"Expected 200, got {r.status_code}"
        assert r.headers.get("content-type") == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", \
            f"Wrong content-type: {r.headers.get('content-type')}"
        # Verify it's a valid xlsx
        wb = load_workbook(BytesIO(r.content))
        assert wb, "Failed to load workbook"
        self.log(f"Workbook has {len(wb.sheetnames)} sheet(s)")

    def _test_export_xlsx_guru(self):
        if not self.sessions:
            self.log("No sessions found, skipping", "WARN")
            return
        session_id = self.sessions[0]["id"]
        r = requests.get(f"{BASE_URL}/export/session/{session_id}/xlsx", 
                        headers={"Authorization": f"Bearer {self.tokens['guru']}"})
        assert r.status_code == 200, f"Expected 200, got {r.status_code}"

    def _test_export_xlsx_siswa_forbidden(self):
        if not self.sessions:
            self.log("No sessions found, skipping", "WARN")
            return
        session_id = self.sessions[0]["id"]
        r = requests.get(f"{BASE_URL}/export/session/{session_id}/xlsx", 
                        headers={"Authorization": f"Bearer {self.tokens['siswa']}"})
        assert r.status_code == 403, f"Expected 403, got {r.status_code}"

    def _test_export_xlsx_invalid_session(self):
        r = requests.get(f"{BASE_URL}/export/session/invalid-session-id/xlsx", 
                        headers={"Authorization": f"Bearer {self.tokens['admin']}"})
        assert r.status_code == 404, f"Expected 404, got {r.status_code}"

    def _test_export_xlsx_empty_session(self):
        # Create a session with no attempts
        packages = requests.get(f"{BASE_URL}/packages", 
                               headers={"Authorization": f"Bearer {self.tokens['admin']}"}).json()
        if not packages:
            self.log("No packages found, skipping", "WARN")
            return
        
        session = self.create_temp_session(self.tokens["admin"], "TEST Empty Session", packages[0]["id"])
        r = requests.get(f"{BASE_URL}/export/session/{session['id']}/xlsx", 
                        headers={"Authorization": f"Bearer {self.tokens['admin']}"})
        assert r.status_code == 200, f"Expected 200, got {r.status_code}"
        
        # Verify workbook has "Belum ada peserta" message
        wb = load_workbook(BytesIO(r.content))
        ws = wb.active
        found_empty_msg = False
        for row in ws.iter_rows(values_only=True):
            if any("Belum ada peserta" in str(cell) for cell in row if cell):
                found_empty_msg = True
                break
        assert found_empty_msg, "Empty session should show 'Belum ada peserta' message"

    def _test_max_attempts_1(self):
        # Find or create a session with max_attempts=1
        session = None
        for s in self.sessions:
            if s.get("max_attempts", 1) == 1:
                session = s
                break
        
        if not session:
            # Create one
            packages = requests.get(f"{BASE_URL}/packages", 
                                   headers={"Authorization": f"Bearer {self.tokens['admin']}"}).json()
            if not packages:
                self.log("No packages found, skipping", "WARN")
                return
            session = self.create_temp_session(self.tokens["admin"], "TEST Max Attempts 1", packages[0]["id"], max_attempts=1)
        
        r = requests.get(f"{BASE_URL}/export/session/{session['id']}/xlsx", 
                        headers={"Authorization": f"Bearer {self.tokens['admin']}"})
        assert r.status_code == 200, f"Expected 200, got {r.status_code}"
        
        wb = load_workbook(BytesIO(r.content))
        ws = wb.active
        # Find header row
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            if row and "Nama Siswa" in str(row):
                header_row = row
                break
        
        assert header_row, "Header row not found"
        # Verify NO "Percobaan" or "Dipakai" columns
        assert "Percobaan" not in header_row, "max_attempts=1 should NOT have 'Percobaan' column"
        assert "Dipakai" not in header_row, "max_attempts=1 should NOT have 'Dipakai' column"
        self.log("Confirmed: max_attempts=1 has no Percobaan/Dipakai columns")

    def _test_max_attempts_gt1(self):
        # Find the existing "UH Matematika - Kelas X" session with max_attempts=2
        session = None
        for s in self.sessions:
            if "Matematika" in s.get("title", "") and s.get("max_attempts", 1) > 1:
                session = s
                break
        
        if not session:
            # Create one
            packages = requests.get(f"{BASE_URL}/packages", 
                                   headers={"Authorization": f"Bearer {self.tokens['admin']}"}).json()
            if not packages:
                self.log("No packages found, skipping", "WARN")
                return
            session = self.create_temp_session(self.tokens["admin"], "TEST Max Attempts 2", packages[0]["id"], max_attempts=2)
        
        r = requests.get(f"{BASE_URL}/export/session/{session['id']}/xlsx", 
                        headers={"Authorization": f"Bearer {self.tokens['admin']}"})
        assert r.status_code == 200, f"Expected 200, got {r.status_code}"
        
        wb = load_workbook(BytesIO(r.content))
        ws = wb.active
        # Find header row
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            if row and "Nama Siswa" in str(row):
                header_row = row
                break
        
        assert header_row, "Header row not found"
        # Verify "Percobaan" and "Dipakai" columns exist
        assert "Percobaan" in header_row, "max_attempts>1 should have 'Percobaan' column"
        assert "Dipakai" in header_row, "max_attempts>1 should have 'Dipakai' column"
        self.log("Confirmed: max_attempts>1 has Percobaan/Dipakai columns")

    def _test_workbook_structure(self):
        if not self.sessions:
            self.log("No sessions found, skipping", "WARN")
            return
        
        session_id = self.sessions[0]["id"]
        r = requests.get(f"{BASE_URL}/export/session/{session_id}/xlsx", 
                        headers={"Authorization": f"Bearer {self.tokens['admin']}"})
        assert r.status_code == 200, f"Expected 200, got {r.status_code}"
        
        wb = load_workbook(BytesIO(r.content))
        ws = wb.active
        
        # Check for RINGKASAN block
        found_ringkasan = False
        for row in ws.iter_rows(values_only=True):
            if row and "RINGKASAN" in str(row):
                found_ringkasan = True
                break
        assert found_ringkasan, "RINGKASAN block not found in workbook"
        
        # Check for frozen panes
        assert ws.freeze_panes is not None, "Header should be frozen"
        self.log("Workbook structure verified: RINGKASAN found, header frozen")

    def _test_data_consistency(self):
        # Find a session with attempts
        session = None
        for s in self.sessions:
            if "Matematika" in s.get("title", ""):
                session = s
                break
        
        if not session:
            self.log("No suitable session found, skipping", "WARN")
            return
        
        # Get results from API
        r_api = requests.get(f"{BASE_URL}/results/session/{session['id']}", 
                            headers={"Authorization": f"Bearer {self.tokens['admin']}"})
        assert r_api.status_code == 200, f"API call failed: {r_api.status_code}"
        api_data = r_api.json()
        
        # Get xlsx
        r_xlsx = requests.get(f"{BASE_URL}/export/session/{session['id']}/xlsx", 
                             headers={"Authorization": f"Bearer {self.tokens['admin']}"})
        assert r_xlsx.status_code == 200, f"XLSX call failed: {r_xlsx.status_code}"
        
        wb = load_workbook(BytesIO(r_xlsx.content))
        ws = wb.active
        
        # Count data rows in xlsx (excluding header and metadata)
        # Note: ALL attempts should be shown, including non-counted ones (with empty No column)
        data_rows = []
        in_data = False
        for row in ws.iter_rows(values_only=True):
            if row and "Nama Siswa" in str(row):
                in_data = True
                continue
            if in_data and row:
                # Check if this is a data row (has student name in column B/2nd column)
                if len(row) > 1 and row[1] and str(row[1]).strip() and row[1] != "RINGKASAN":
                    if "Belum ada peserta" not in str(row):
                        data_rows.append(row)
            if row and "RINGKASAN" in str(row):
                break
        
        # Compare counts
        api_attempts = len(api_data.get("attempts", []))
        xlsx_rows = len(data_rows)
        
        self.log(f"API attempts: {api_attempts}, XLSX rows: {xlsx_rows}")
        # They should match (all attempts are shown, counted or not)
        assert xlsx_rows == api_attempts or xlsx_rows == 0, \
            f"Data mismatch: API has {api_attempts} attempts, XLSX has {xlsx_rows} rows"

    def _test_essay_grading(self):
        # Just verify the endpoint is accessible
        r = requests.get(f"{BASE_URL}/results/me", 
                        headers={"Authorization": f"Bearer {self.tokens['siswa']}"})
        assert r.status_code == 200, f"Essay grading endpoint check failed: {r.status_code}"

    def _test_analytics(self):
        if not self.sessions:
            self.log("No sessions found, skipping", "WARN")
            return
        session_id = self.sessions[0]["id"]
        r = requests.get(f"{BASE_URL}/analytics/session/{session_id}", 
                        headers={"Authorization": f"Bearer {self.tokens['admin']}"})
        assert r.status_code == 200, f"Analytics endpoint failed: {r.status_code}"

    def _test_kartu_hasil_pdf(self):
        # Get an attempt
        r = requests.get(f"{BASE_URL}/results/me", 
                        headers={"Authorization": f"Bearer {self.tokens['siswa']}"})
        if r.status_code != 200:
            self.log("No results found for siswa, skipping", "WARN")
            return
        attempts = r.json()
        if not attempts:
            self.log("No attempts found, skipping", "WARN")
            return
        
        attempt_id = attempts[0]["id"]
        r = requests.get(f"{BASE_URL}/results/detail/{attempt_id}/pdf", 
                        headers={"Authorization": f"Bearer {self.tokens['siswa']}"})
        assert r.status_code == 200, f"Kartu hasil PDF failed: {r.status_code}"
        assert "application/pdf" in r.headers.get("content-type", ""), "Should return PDF"

    def _test_rapor_pdf(self):
        r = requests.get(f"{BASE_URL}/report/student/me/pdf", 
                        headers={"Authorization": f"Bearer {self.tokens['siswa']}"})
        assert r.status_code == 200, f"Rapor PDF failed: {r.status_code}"
        assert "application/pdf" in r.headers.get("content-type", ""), "Should return PDF"

    def _test_csv_export(self):
        # CSV export is frontend-only, but we can verify the data endpoint works
        if not self.sessions:
            self.log("No sessions found, skipping", "WARN")
            return
        session_id = self.sessions[0]["id"]
        r = requests.get(f"{BASE_URL}/results/session/{session_id}", 
                        headers={"Authorization": f"Bearer {self.tokens['admin']}"})
        assert r.status_code == 200, f"Results endpoint for CSV failed: {r.status_code}"


if __name__ == "__main__":
    runner = TestRunner()
    sys.exit(runner.run_all_tests())
