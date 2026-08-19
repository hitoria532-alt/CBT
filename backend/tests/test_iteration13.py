"""Iteration 13: PG options A-E + polished Excel export of session results."""
import io
import os
import time
import uuid

import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "http://localhost:8001").rstrip("/")
API = f"{BASE_URL}/api"

ADMIN = {"email": "hitoria532@gmail.com", "password": "admin123"}
GURU = {"email": "guru@sekolah.id", "password": "guru123"}
SISWA = {"email": "ani@sekolah.id", "password": "siswa123"}
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _login(creds):
    r = requests.post(f"{API}/auth/login", json=creds, timeout=20)
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text}"
    return r.json()["token"]


def H(t):
    return {"Authorization": f"Bearer {t}"}


@pytest.fixture(scope="module")
def admin_token():
    return _login(ADMIN)


@pytest.fixture(scope="module")
def guru_token():
    return _login(GURU)


@pytest.fixture(scope="module")
def siswa_token():
    return _login(SISWA)


# ---------------------------------------------------------------- A-E options
class TestFiveOptions:
    def test_template_has_option_e(self, admin_token):
        r = requests.get(f"{API}/questions/import-template", headers=H(admin_token), timeout=20)
        assert r.status_code == 200
        header = r.text.splitlines()[0]
        assert "option_e" in header
        cols = [c.strip() for c in header.split(",")]
        # A..E must appear in order
        assert cols.index("option_a") < cols.index("option_b") < cols.index("option_c")
        assert cols.index("option_c") < cols.index("option_d") < cols.index("option_e")

    def test_create_question_with_five_options(self, admin_token):
        unique = uuid.uuid4().hex[:8]
        cat = requests.post(f"{API}/categories", headers=H(admin_token),
                            json={"name": f"TEST_I13_{unique}"}, timeout=20).json()
        try:
            r = requests.post(f"{API}/questions", headers=H(admin_token), json={
                "category_id": cat["id"], "type": "pg",
                "text": f"TEST_I13 lima opsi {unique}",
                "options": ["Bogor", "Bandung", "Bekasi", "Cimahi", "Depok"],
                "correct_answer": "4", "weight": 1.0}, timeout=20)
            assert r.status_code == 200, r.text
            q = r.json()
            assert len(q["options"]) == 5
            assert q["options"][4] == "Depok"
            assert q["correct_answer"] == "4"
            # round-trips through the list endpoint
            got = requests.get(f"{API}/questions", headers=H(admin_token),
                               params={"category_id": cat["id"]}, timeout=20).json()
            assert got[0]["options"][4] == "Depok"
            requests.delete(f"{API}/questions/{q['id']}", headers=H(admin_token), timeout=20)
        finally:
            requests.delete(f"{API}/categories/{cat['id']}", headers=H(admin_token), timeout=20)

    def test_import_option_e_and_key_e(self, admin_token):
        unique = uuid.uuid4().hex[:8]
        cat_name = f"TEST_I13_IMP_{unique}"
        csv = (
            "type,text,option_a,option_b,option_c,option_d,option_e,correct,weight,category\n"
            f"pg,TEST_I13_Q_E_{unique},Bogor,Bandung,Bekasi,Cimahi,Depok,E,1,{cat_name}\n"
            f"pg,TEST_I13_Q_lower_{unique},A1,B1,C1,D1,E1,e,2,{cat_name}\n"
            f"pg,TEST_I13_Q_ABC_{unique},x,y,z,,,C,1,{cat_name}\n"
        )
        files = {"file": (f"i13_{unique}.csv", csv.encode(), "text/csv")}
        r = requests.post(f"{API}/questions/import", headers=H(admin_token), files=files, timeout=60)
        assert r.status_code == 200, r.text
        assert r.json()["imported"] == 3, r.json()

        cats = requests.get(f"{API}/categories", headers=H(admin_token), timeout=20).json()
        cat = next(c for c in cats if c["name"] == cat_name)
        qs = requests.get(f"{API}/questions", headers=H(admin_token),
                          params={"category_id": cat["id"]}, timeout=20).json()
        by_text = {q["text"]: q for q in qs}

        qe = by_text[f"TEST_I13_Q_E_{unique}"]
        assert qe["options"] == ["Bogor", "Bandung", "Bekasi", "Cimahi", "Depok"]
        assert qe["correct_answer"] == "4"  # E

        ql = by_text[f"TEST_I13_Q_lower_{unique}"]
        assert ql["correct_answer"] == "4"  # lowercase 'e' also works
        assert ql["weight"] == 2.0

        # only A-C filled -> trailing blanks trimmed, key C still index 2
        qabc = by_text[f"TEST_I13_Q_ABC_{unique}"]
        assert qabc["options"] == ["x", "y", "z"]
        assert qabc["correct_answer"] == "2"

        for q in qs:
            requests.delete(f"{API}/questions/{q['id']}", headers=H(admin_token), timeout=20)
        requests.delete(f"{API}/categories/{cat['id']}", headers=H(admin_token), timeout=20)

    def test_import_rejects_key_pointing_at_empty_option(self, admin_token):
        unique = uuid.uuid4().hex[:8]
        cat_name = f"TEST_I13_BAD_{unique}"
        csv = (
            "type,text,option_a,option_b,option_c,option_d,option_e,correct,weight,category\n"
            f"pg,TEST_I13_BADQ_{unique},x,y,,,,D,1,{cat_name}\n"
        )
        files = {"file": (f"bad_{unique}.csv", csv.encode(), "text/csv")}
        r = requests.post(f"{API}/questions/import", headers=H(admin_token), files=files, timeout=60)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["imported"] == 0
        assert any("kosong" in e.lower() for e in body["errors"]), body

    def test_middle_blank_does_not_shift_key(self, admin_token):
        """option_c blank but option_d filled -> D must stay index 3."""
        unique = uuid.uuid4().hex[:8]
        cat_name = f"TEST_I13_MID_{unique}"
        csv = (
            "type,text,option_a,option_b,option_c,option_d,option_e,correct,weight,category\n"
            f"pg,TEST_I13_MIDQ_{unique},aa,bb,,dd,,D,1,{cat_name}\n"
        )
        files = {"file": (f"mid_{unique}.csv", csv.encode(), "text/csv")}
        r = requests.post(f"{API}/questions/import", headers=H(admin_token), files=files, timeout=60)
        assert r.status_code == 200, r.text
        assert r.json()["imported"] == 1, r.json()
        cats = requests.get(f"{API}/categories", headers=H(admin_token), timeout=20).json()
        cat = next(c for c in cats if c["name"] == cat_name)
        qs = requests.get(f"{API}/questions", headers=H(admin_token),
                          params={"category_id": cat["id"]}, timeout=20).json()
        q = qs[0]
        assert q["options"] == ["aa", "bb", "", "dd"]
        assert q["correct_answer"] == "3", "D must remain index 3"
        requests.delete(f"{API}/questions/{q['id']}", headers=H(admin_token), timeout=20)
        requests.delete(f"{API}/categories/{cat['id']}", headers=H(admin_token), timeout=20)


# ---------------------------------------------------- Session Excel export
class TestSessionExcelExport:
    @pytest.fixture(scope="class")
    def session_id(self, admin_token):
        sessions = requests.get(f"{API}/sessions", headers=H(admin_token), timeout=20).json()
        s = next(x for x in sessions if x["title"] == "UH Matematika - Kelas X")
        return s["id"]

    def test_export_returns_xlsx(self, admin_token, session_id):
        r = requests.get(f"{API}/export/session/{session_id}/xlsx",
                         headers=H(admin_token), timeout=60)
        assert r.status_code == 200, r.text
        assert r.headers["content-type"].startswith(XLSX_MIME)
        assert "attachment" in r.headers.get("content-disposition", "")
        assert r.content[:2] == b"PK", "must be a real xlsx (zip) payload"
        assert len(r.content) > 3000

    def test_workbook_structure_and_content(self, admin_token, session_id):
        from openpyxl import load_workbook
        r = requests.get(f"{API}/export/session/{session_id}/xlsx",
                         headers=H(admin_token), timeout=60)
        wb = load_workbook(io.BytesIO(r.content))
        assert wb.sheetnames[:3] == ["Rekap Nilai", "Rincian Jawaban", "Analisis Butir"]
        assert "Pelanggaran" in wb.sheetnames

        ws = wb["Rekap Nilai"]
        flat = [str(c) for row in ws.iter_rows(values_only=True) for c in row if c is not None]
        blob = " | ".join(flat)
        assert "REKAP NILAI HASIL UJIAN" in blob
        assert "UH Matematika - Kelas X" in blob
        for h in ["No", "Nama Siswa", "NISN/NIP", "Kelas", "Status", "Benar", "Salah",
                  "Kosong", "Poin", "Nilai", "Predikat", "Keterangan", "Waktu Kumpul"]:
            assert h in flat, f"missing column {h}"
        # session info + summary blocks
        for label in ["Paket Soal", "KKM", "Metode Penilaian", "RINGKASAN",
                      "Rata-rata Nilai", "Ketuntasan", "Jumlah Lulus"]:
            assert label in flat, f"missing label {label}"
        # seeded students and their scores
        assert "Ani Siswa" in flat and "Budi Siswa" in flat
        assert 100.0 in [c for row in ws.iter_rows(values_only=True) for c in row]
        # polish: header frozen, filter set, gridlines hidden
        assert ws.freeze_panes is not None
        assert ws.auto_filter.ref is not None
        assert ws.sheet_view.showGridLines is False

        ws2 = wb["Rincian Jawaban"]
        flat2 = [str(c) for row in ws2.iter_rows(values_only=True) for c in row if c is not None]
        assert "S1" in flat2 and "Ani Siswa" in flat2
        assert ws2.freeze_panes is not None

        ws3 = wb["Analisis Butir"]
        flat3 = [str(c) for row in ws3.iter_rows(values_only=True) for c in row if c is not None]
        for h in ["Tipe", "Soal", "Kunci", "% Benar", "Kesukaran"]:
            assert h in flat3, f"missing analysis column {h}"
        assert any(x in flat3 for x in ("Mudah", "Sedang", "Sulit"))

    def test_guru_allowed(self, guru_token, session_id):
        r = requests.get(f"{API}/export/session/{session_id}/xlsx",
                         headers=H(guru_token), timeout=60)
        assert r.status_code == 200

    def test_siswa_forbidden(self, siswa_token, session_id):
        r = requests.get(f"{API}/export/session/{session_id}/xlsx",
                         headers=H(siswa_token), timeout=60)
        assert r.status_code == 403

    def test_unauth_blocked(self, session_id):
        r = requests.get(f"{API}/export/session/{session_id}/xlsx", timeout=30)
        assert r.status_code == 401

    def test_unknown_session_404(self, admin_token):
        r = requests.get(f"{API}/export/session/does-not-exist/xlsx",
                         headers=H(admin_token), timeout=30)
        assert r.status_code == 404

    def test_export_session_without_attempts(self, admin_token):
        """A brand new session with no participants must still export cleanly."""
        from datetime import datetime, timedelta, timezone
        from openpyxl import load_workbook
        unique = uuid.uuid4().hex[:8]
        cat = requests.post(f"{API}/categories", headers=H(admin_token),
                            json={"name": f"TEST_I13_EMP_{unique}"}, timeout=20).json()
        q = requests.post(f"{API}/questions", headers=H(admin_token), json={
            "category_id": cat["id"], "type": "pg", "text": f"TEST_I13_EMPQ_{unique}",
            "options": ["a", "b", "c", "d", "e"], "correct_answer": "4", "weight": 1.0},
            timeout=20).json()
        pkg = requests.post(f"{API}/packages", headers=H(admin_token), json={
            "title": f"TEST_I13_EMP_PKG_{unique}", "category_id": cat["id"],
            "question_ids": [q["id"]], "scoring_method": "percentage"}, timeout=20).json()
        start = (datetime.now(timezone.utc) + timedelta(days=3)).isoformat()
        end = (datetime.now(timezone.utc) + timedelta(days=4)).isoformat()
        ses = requests.post(f"{API}/sessions", headers=H(admin_token), json={
            "title": f"TEST_I13_EMP_SES_{unique}", "package_id": pkg["id"],
            "start_time": start, "end_time": end, "duration_minutes": 30,
            "kkm": 75.0, "class_ids": []}, timeout=20).json()
        try:
            r = requests.get(f"{API}/export/session/{ses['id']}/xlsx",
                             headers=H(admin_token), timeout=60)
            assert r.status_code == 200, r.text
            wb = load_workbook(io.BytesIO(r.content))
            ws = wb["Rekap Nilai"]
            flat = [str(c) for row in ws.iter_rows(values_only=True) for c in row if c is not None]
            assert any("Belum ada peserta" in x for x in flat)
            # analysis sheet still lists the question
            flat3 = [str(c) for row in wb["Analisis Butir"].iter_rows(values_only=True)
                     for c in row if c is not None]
            assert any(f"TEST_I13_EMPQ_{unique}" in x for x in flat3)
        finally:
            requests.delete(f"{API}/sessions/{ses['id']}", headers=H(admin_token), timeout=20)
            requests.delete(f"{API}/packages/{pkg['id']}", headers=H(admin_token), timeout=20)
            requests.delete(f"{API}/questions/{q['id']}", headers=H(admin_token), timeout=20)
            requests.delete(f"{API}/categories/{cat['id']}", headers=H(admin_token), timeout=20)


# ------------------------------------------- E2E: answering option E is graded
class TestFiveOptionGrading:
    def test_student_can_pick_option_e_and_is_graded(self, admin_token):
        from datetime import datetime, timedelta, timezone
        unique = uuid.uuid4().hex[:8]
        cat = requests.post(f"{API}/categories", headers=H(admin_token),
                            json={"name": f"TEST_I13_G_{unique}"}, timeout=20).json()
        q = requests.post(f"{API}/questions", headers=H(admin_token), json={
            "category_id": cat["id"], "type": "pg", "text": f"TEST_I13_GQ_{unique}",
            "options": ["a", "b", "c", "d", "BENAR-E"], "correct_answer": "4",
            "weight": 1.0}, timeout=20).json()
        pkg = requests.post(f"{API}/packages", headers=H(admin_token), json={
            "title": f"TEST_I13_G_PKG_{unique}", "category_id": cat["id"],
            "question_ids": [q["id"]], "scoring_method": "percentage"}, timeout=20).json()
        start = (datetime.now(timezone.utc) - timedelta(minutes=5)).isoformat()
        end = (datetime.now(timezone.utc) + timedelta(hours=2)).isoformat()
        ses = requests.post(f"{API}/sessions", headers=H(admin_token), json={
            "title": f"TEST_I13_G_SES_{unique}", "package_id": pkg["id"],
            "start_time": start, "end_time": end, "duration_minutes": 60,
            "kkm": 75.0, "class_ids": []}, timeout=20).json()
        email = f"test_i13_{int(time.time())}@example.com"
        stu = requests.post(f"{API}/users", headers=H(admin_token), json={
            "email": email, "password": "pw12345", "name": "TEST I13 Stu",
            "role": "siswa"}, timeout=20).json()
        try:
            stok = _login({"email": email, "password": "pw12345"})
            st = requests.post(f"{API}/exam/start", headers=H(stok),
                               json={"session_id": ses["id"]}, timeout=30)
            assert st.status_code == 200, st.text
            shown = st.json()["questions"][0]
            assert len(shown["options"]) == 5, "student must see all 5 options"
            sub = requests.post(f"{API}/exam/submit", headers=H(stok), json={
                "session_id": ses["id"], "answers": {q["id"]: "4"}}, timeout=30)
            assert sub.status_code == 200, sub.text
            assert sub.json()["score"] == 100.0, sub.json()
        finally:
            requests.delete(f"{API}/users/{stu['id']}", headers=H(admin_token), timeout=20)
            requests.delete(f"{API}/sessions/{ses['id']}", headers=H(admin_token), timeout=20)
            requests.delete(f"{API}/packages/{pkg['id']}", headers=H(admin_token), timeout=20)
            requests.delete(f"{API}/questions/{q['id']}", headers=H(admin_token), timeout=20)
            requests.delete(f"{API}/categories/{cat['id']}", headers=H(admin_token), timeout=20)

    def test_shuffled_options_still_grade_option_e_correctly(self, admin_token):
        """With shuffle_options on, the client sends the displayed index; the server
        must decode it back to the real key (E)."""
        from datetime import datetime, timedelta, timezone
        unique = uuid.uuid4().hex[:8]
        cat = requests.post(f"{API}/categories", headers=H(admin_token),
                            json={"name": f"TEST_I13_S_{unique}"}, timeout=20).json()
        q = requests.post(f"{API}/questions", headers=H(admin_token), json={
            "category_id": cat["id"], "type": "pg", "text": f"TEST_I13_SQ_{unique}",
            "options": ["a", "b", "c", "d", "BENAR-E"], "correct_answer": "4",
            "weight": 1.0}, timeout=20).json()
        pkg = requests.post(f"{API}/packages", headers=H(admin_token), json={
            "title": f"TEST_I13_S_PKG_{unique}", "category_id": cat["id"],
            "question_ids": [q["id"]], "scoring_method": "percentage",
            "shuffle_options": True}, timeout=20).json()
        start = (datetime.now(timezone.utc) - timedelta(minutes=5)).isoformat()
        end = (datetime.now(timezone.utc) + timedelta(hours=2)).isoformat()
        ses = requests.post(f"{API}/sessions", headers=H(admin_token), json={
            "title": f"TEST_I13_S_SES_{unique}", "package_id": pkg["id"],
            "start_time": start, "end_time": end, "duration_minutes": 60,
            "kkm": 75.0, "class_ids": []}, timeout=20).json()
        email = f"test_i13s_{int(time.time())}@example.com"
        stu = requests.post(f"{API}/users", headers=H(admin_token), json={
            "email": email, "password": "pw12345", "name": "TEST I13 Shuf",
            "role": "siswa"}, timeout=20).json()
        try:
            stok = _login({"email": email, "password": "pw12345"})
            st = requests.post(f"{API}/exam/start", headers=H(stok),
                               json={"session_id": ses["id"]}, timeout=30)
            assert st.status_code == 200, st.text
            shown = st.json()["questions"][0]
            assert len(shown["options"]) == 5
            picked = shown["options"].index("BENAR-E")
            sub = requests.post(f"{API}/exam/submit", headers=H(stok), json={
                "session_id": ses["id"], "answers": {q["id"]: str(picked)}}, timeout=30)
            assert sub.status_code == 200, sub.text
            assert sub.json()["score"] == 100.0, (picked, sub.json())
        finally:
            requests.delete(f"{API}/users/{stu['id']}", headers=H(admin_token), timeout=20)
            requests.delete(f"{API}/sessions/{ses['id']}", headers=H(admin_token), timeout=20)
            requests.delete(f"{API}/packages/{pkg['id']}", headers=H(admin_token), timeout=20)
            requests.delete(f"{API}/questions/{q['id']}", headers=H(admin_token), timeout=20)
            requests.delete(f"{API}/categories/{cat['id']}", headers=H(admin_token), timeout=20)
