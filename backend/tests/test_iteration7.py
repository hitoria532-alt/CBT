"""Iteration 7: Leaderboard filters + Excel export."""
import os
import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL").rstrip("/")
API = f"{BASE_URL}/api"

ADMIN = ("hitoria532@gmail.com", "admin123")
GURU = ("guru@sekolah.id", "guru123")
SISWA = ("siswa@sekolah.id", "siswa123")


def _login(email, pw):
    r = requests.post(f"{API}/auth/login", json={"email": email, "password": pw}, timeout=30)
    assert r.status_code == 200, f"login failed for {email}: {r.status_code} {r.text}"
    return r.json()["token"]


@pytest.fixture(scope="module")
def admin_h():
    return {"Authorization": f"Bearer {_login(*ADMIN)}"}


@pytest.fixture(scope="module")
def guru_h():
    return {"Authorization": f"Bearer {_login(*GURU)}"}


@pytest.fixture(scope="module")
def siswa_h():
    return {"Authorization": f"Bearer {_login(*SISWA)}"}


# --- Filter Peringkat backend ---

class TestGlobalLeaderboardFilters:
    def test_no_filter_has_completed(self, admin_h):
        r = requests.get(f"{API}/leaderboard/global", headers=admin_h, timeout=30)
        assert r.status_code == 200
        rows = r.json()["rows"]
        assert isinstance(rows, list)
        assert any(row["completed"] > 0 for row in rows), "expected at least one student with completed>0"
        # ranking sanity
        avgs = [row["avg_score"] for row in rows]
        assert avgs == sorted(avgs, reverse=True)
        for i, row in enumerate(rows):
            assert row["rank"] == i + 1

    def test_far_past_date_range_zero(self, admin_h):
        r = requests.get(f"{API}/leaderboard/global",
                         params={"start": "2020-01-01", "end": "2020-12-31"},
                         headers=admin_h, timeout=30)
        assert r.status_code == 200
        rows = r.json()["rows"]
        assert all(row["completed"] == 0 for row in rows)
        assert all(row["avg_score"] == 0 for row in rows)

    def test_category_filter_applied(self, admin_h):
        # Get some category id
        cats = requests.get(f"{API}/categories", headers=admin_h, timeout=30).json()
        assert isinstance(cats, list) and len(cats) > 0
        cid = cats[0]["id"]
        r = requests.get(f"{API}/leaderboard/global",
                         params={"category_id": cid}, headers=admin_h, timeout=30)
        assert r.status_code == 200
        rows = r.json()["rows"]
        # completed values must be <= no-filter (orphaned attempts are excluded)
        base = requests.get(f"{API}/leaderboard/global", headers=admin_h,
                            timeout=30).json()["rows"]
        base_by_id = {row["student_id"]: row["completed"] for row in base}
        for row in rows:
            assert row["completed"] <= base_by_id.get(row["student_id"], 0)

    def test_guru_can_access(self, guru_h):
        r = requests.get(f"{API}/leaderboard/global", headers=guru_h, timeout=30)
        assert r.status_code == 200

    def test_siswa_can_access_global(self, siswa_h):
        # student page uses this endpoint too
        r = requests.get(f"{API}/leaderboard/global", headers=siswa_h, timeout=30)
        assert r.status_code in (200, 403)  # depends on policy; document actual


# --- Ekspor Peringkat ---

class TestExportLeaderboard:
    XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def test_admin_export_no_filter(self, admin_h):
        r = requests.get(f"{API}/export/leaderboard/xlsx", headers=admin_h, timeout=60)
        assert r.status_code == 200
        assert self.XLSX_CT in r.headers.get("content-type", "")
        assert r.content[:2] == b"PK"  # xlsx zip magic

    def test_admin_export_with_filters(self, admin_h):
        cats = requests.get(f"{API}/categories", headers=admin_h, timeout=30).json()
        cid = cats[0]["id"] if cats else None
        params = {"start": "2024-01-01", "end": "2030-12-31"}
        if cid:
            params["category_id"] = cid
        r = requests.get(f"{API}/export/leaderboard/xlsx", params=params,
                         headers=admin_h, timeout=60)
        assert r.status_code == 200
        assert r.content[:2] == b"PK"

    def test_guru_can_export(self, guru_h):
        r = requests.get(f"{API}/export/leaderboard/xlsx", headers=guru_h, timeout=60)
        assert r.status_code == 200
        assert r.content[:2] == b"PK"

    def test_siswa_forbidden(self, siswa_h):
        r = requests.get(f"{API}/export/leaderboard/xlsx", headers=siswa_h, timeout=30)
        assert r.status_code == 403

    def test_xlsx_content_structure(self, admin_h):
        import io
        from openpyxl import load_workbook
        r = requests.get(f"{API}/export/leaderboard/xlsx", headers=admin_h, timeout=60)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # Title cell
        assert "PERINGKAT" in str(ws["A1"].value).upper()
        # Filter summary row 2 contains Mapel and Periode
        row2 = " ".join(str(c.value or "") for c in ws[2])
        assert "Mapel" in row2 and "Periode" in row2
        # Header row (row 4)
        headers = [c.value for c in ws[4]]
        for h in ["Peringkat", "Nama Siswa", "NISN/NIP", "Kelas", "Ujian Selesai", "Rata-rata"]:
            assert h in headers


# --- Regression ---

class TestRegression:
    def test_class_leaderboard(self, admin_h):
        classes = requests.get(f"{API}/classes", headers=admin_h, timeout=30).json()
        if not classes:
            pytest.skip("no classes")
        r = requests.get(f"{API}/leaderboard/class/{classes[0]['id']}",
                         headers=admin_h, timeout=30)
        assert r.status_code == 200
        data = r.json()
        assert "rows" in data and "class_name" in data

    def test_student_leaderboard_me(self, siswa_h):
        r = requests.get(f"{API}/leaderboard/me", headers=siswa_h, timeout=30)
        assert r.status_code == 200
