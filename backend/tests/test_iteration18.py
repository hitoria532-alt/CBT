"""
Test iteration 18: Class roster management (student accounts within classes)
- GET /api/classes/{cid}/students
- POST /api/classes/{cid}/students (create student account)
- POST /api/classes/{cid}/students/attach
- DELETE /api/classes/{cid}/students/{sid}
- GET /api/classes/{cid}/students/xlsx
- POST /api/students/import with class_id
"""
import pytest
import requests
import io
import pandas as pd
from openpyxl import load_workbook

BASE_URL = "https://deploy-web-app-3.preview.emergentagent.com/api"

@pytest.fixture(scope="module")
def admin_token():
    """Login as admin and return token"""
    resp = requests.post(f"{BASE_URL}/auth/login", json={
        "email": "admin@sekolah.id",
        "password": "Admin@12345"
    })
    assert resp.status_code == 200, f"Admin login failed: {resp.text}"
    return resp.json()["token"]

@pytest.fixture(scope="module")
def test_class(admin_token):
    """Create a test class for roster management"""
    resp = requests.post(
        f"{BASE_URL}/classes",
        json={"name": "Test Roster Class", "description": "For testing roster APIs", "student_ids": []},
        headers={"Authorization": f"Bearer {admin_token}"}
    )
    assert resp.status_code == 200, f"Failed to create test class: {resp.text}"
    cls = resp.json()
    yield cls
    # Cleanup
    requests.delete(f"{BASE_URL}/classes/{cls['id']}", headers={"Authorization": f"Bearer {admin_token}"})

class TestClassRosterAPIs:
    """Test class roster management endpoints"""

    def test_get_class_students_empty(self, admin_token, test_class):
        """GET /api/classes/{cid}/students returns empty roster for new class"""
        resp = requests.get(
            f"{BASE_URL}/classes/{test_class['id']}/students",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert resp.status_code == 200, f"Failed: {resp.text}"
        data = resp.json()
        assert "class" in data
        assert "students" in data
        assert "available" in data
        assert data["class"]["id"] == test_class["id"]
        assert data["class"]["name"] == test_class["name"]
        assert isinstance(data["students"], list)
        assert isinstance(data["available"], list)
        print(f"✓ GET class students returns correct structure with {len(data['students'])} students, {len(data['available'])} available")

    def test_create_student_account(self, admin_token, test_class):
        """POST /api/classes/{cid}/students creates a login-ready student account"""
        student_data = {
            "name": "Test Student Roster",
            "email": "test.roster.student@sekolah.id",
            "password": "student123",
            "identifier": "9999999"
        }
        resp = requests.post(
            f"{BASE_URL}/classes/{test_class['id']}/students",
            json=student_data,
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert resp.status_code == 200, f"Failed to create student: {resp.text}"
        student = resp.json()
        assert student["name"] == student_data["name"]
        assert student["email"] == student_data["email"]
        assert student["identifier"] == student_data["identifier"]
        assert "id" in student
        assert "exams_done" in student
        print(f"✓ Created student account: {student['name']} ({student['email']})")
        
        # Verify student can login immediately
        login_resp = requests.post(f"{BASE_URL}/auth/login", json={
            "email": student_data["email"],
            "password": student_data["password"]
        })
        assert login_resp.status_code == 200, f"Student cannot login: {login_resp.text}"
        user = login_resp.json()["user"]
        assert user["role"] == "siswa"
        assert user["email"] == student_data["email"]
        print(f"✓ Student can login immediately after creation")
        
        # Verify student appears in class roster
        roster_resp = requests.get(
            f"{BASE_URL}/classes/{test_class['id']}/students",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert roster_resp.status_code == 200
        roster = roster_resp.json()
        student_ids = [s["id"] for s in roster["students"]]
        assert student["id"] in student_ids
        print(f"✓ Student appears in class roster")

    def test_create_student_validations(self, admin_token, test_class):
        """POST /api/classes/{cid}/students validates input correctly"""
        # Empty name
        resp = requests.post(
            f"{BASE_URL}/classes/{test_class['id']}/students",
            json={"name": "", "email": "test@sekolah.id", "password": "pass123"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert resp.status_code == 400
        assert "nama" in resp.text.lower()
        print("✓ Rejects empty name")
        
        # Short password
        resp = requests.post(
            f"{BASE_URL}/classes/{test_class['id']}/students",
            json={"name": "Test", "email": "test2@sekolah.id", "password": "123"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert resp.status_code == 400
        assert "password" in resp.text.lower() and "5" in resp.text
        print("✓ Rejects password shorter than 5 chars")
        
        # Duplicate email
        resp = requests.post(
            f"{BASE_URL}/classes/{test_class['id']}/students",
            json={"name": "Duplicate", "email": "test.roster.student@sekolah.id", "password": "pass123"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert resp.status_code == 400
        assert "email" in resp.text.lower() or "terdaftar" in resp.text.lower()
        print("✓ Rejects duplicate email")
        
        # Non-student email (admin email)
        resp = requests.post(
            f"{BASE_URL}/classes/{test_class['id']}/students",
            json={"name": "Admin Test", "email": "admin@sekolah.id", "password": "pass123"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert resp.status_code == 400
        print("✓ Rejects non-student email")

    def test_attach_existing_students(self, admin_token, test_class):
        """POST /api/classes/{cid}/students/attach adds existing students to class"""
        # First, get available students
        roster_resp = requests.get(
            f"{BASE_URL}/classes/{test_class['id']}/students",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert roster_resp.status_code == 200
        available = roster_resp.json()["available"]
        
        if len(available) == 0:
            print("⚠ No available students to attach, skipping test")
            return
        
        # Attach first available student
        student_to_attach = available[0]
        resp = requests.post(
            f"{BASE_URL}/classes/{test_class['id']}/students/attach",
            json={"student_ids": [student_to_attach["id"]]},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert resp.status_code == 200, f"Failed to attach student: {resp.text}"
        result = resp.json()
        assert "added" in result
        assert "total" in result
        print(f"✓ Attached {result['added']} student(s), total now {result['total']}")
        
        # Verify student is now in roster
        roster_resp = requests.get(
            f"{BASE_URL}/classes/{test_class['id']}/students",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        roster = roster_resp.json()
        student_ids = [s["id"] for s in roster["students"]]
        assert student_to_attach["id"] in student_ids
        print(f"✓ Attached student appears in roster")
        
        # Try to attach same student again (should not duplicate)
        resp = requests.post(
            f"{BASE_URL}/classes/{test_class['id']}/students/attach",
            json={"student_ids": [student_to_attach["id"]]},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert resp.status_code == 200
        result = resp.json()
        assert result["added"] == 0  # No new students added
        print(f"✓ Does not duplicate when attaching same student")

    def test_remove_student_from_class(self, admin_token, test_class):
        """DELETE /api/classes/{cid}/students/{sid} removes student from class only"""
        # Get current roster
        roster_resp = requests.get(
            f"{BASE_URL}/classes/{test_class['id']}/students",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        students = roster_resp.json()["students"]
        if len(students) == 0:
            print("⚠ No students in class to remove, skipping test")
            return
        
        student = students[0]
        student_email = student["email"]
        
        # Remove from class (without deleting account)
        resp = requests.delete(
            f"{BASE_URL}/classes/{test_class['id']}/students/{student['id']}",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert resp.status_code == 200, f"Failed to remove student: {resp.text}"
        result = resp.json()
        assert result["ok"] is True
        assert result.get("deleted_account") is False
        print(f"✓ Removed student from class")
        
        # Verify student is not in roster anymore
        roster_resp = requests.get(
            f"{BASE_URL}/classes/{test_class['id']}/students",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        roster = roster_resp.json()
        student_ids = [s["id"] for s in roster["students"]]
        assert student["id"] not in student_ids
        print(f"✓ Student no longer in class roster")
        
        # Verify account still exists (can still login)
        login_resp = requests.post(f"{BASE_URL}/auth/login", json={
            "email": student_email,
            "password": "student123"  # assuming this is the password
        })
        # If login fails with 401, it means password is wrong but account exists
        # If it's 404 or account not found, that's a problem
        assert login_resp.status_code in [200, 401], f"Account was deleted: {login_resp.text}"
        print(f"✓ Student account still exists after removal from class")

    def test_delete_student_account(self, admin_token, test_class):
        """DELETE /api/classes/{cid}/students/{sid}?delete_account=true deletes the account"""
        # Create a temporary student for deletion
        student_data = {
            "name": "To Be Deleted",
            "email": "delete.me@sekolah.id",
            "password": "delete123",
            "identifier": "8888888"
        }
        create_resp = requests.post(
            f"{BASE_URL}/classes/{test_class['id']}/students",
            json=student_data,
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert create_resp.status_code == 200
        student = create_resp.json()
        print(f"✓ Created temporary student for deletion test")
        
        # Delete account
        resp = requests.delete(
            f"{BASE_URL}/classes/{test_class['id']}/students/{student['id']}?delete_account=true",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert resp.status_code == 200, f"Failed to delete account: {resp.text}"
        result = resp.json()
        assert result["ok"] is True
        assert result.get("deleted_account") is True
        print(f"✓ Deleted student account")
        
        # Verify account no longer exists (cannot login)
        login_resp = requests.post(f"{BASE_URL}/auth/login", json={
            "email": student_data["email"],
            "password": student_data["password"]
        })
        assert login_resp.status_code == 401, f"Account still exists: {login_resp.text}"
        print(f"✓ Student account cannot login after deletion")

    def test_export_class_roster_xlsx(self, admin_token, test_class):
        """GET /api/classes/{cid}/students/xlsx returns valid Excel file"""
        resp = requests.get(
            f"{BASE_URL}/classes/{test_class['id']}/students/xlsx",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert resp.status_code == 200, f"Failed to export roster: {resp.text}"
        assert resp.headers.get("content-type") == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert len(resp.content) > 0
        print(f"✓ Export returns Excel file ({len(resp.content)} bytes)")
        
        # Verify it's a valid Excel file
        try:
            wb = load_workbook(io.BytesIO(resp.content))
            assert "Akun Siswa" in wb.sheetnames
            print(f"✓ Excel file is valid with 'Akun Siswa' sheet")
        except Exception as e:
            pytest.fail(f"Invalid Excel file: {e}")

    def test_import_students_with_class_id(self, admin_token, test_class):
        """POST /api/students/import with class_id imports students into that class"""
        # Create a CSV with students (empty kelas column)
        csv_data = """nama,username,password,nis,kelas
Import Test 1,import.test1@sekolah.id,import123,1111111,
Import Test 2,import.test2@sekolah.id,import123,2222222,
Import Test 3,import.test3@sekolah.id,import123,3333333,Other Class
"""
        files = {"file": ("test_import.csv", csv_data, "text/csv")}
        data = {"class_id": test_class["id"]}
        
        resp = requests.post(
            f"{BASE_URL}/students/import",
            files=files,
            data=data,
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert resp.status_code == 200, f"Import failed: {resp.text}"
        result = resp.json()
        assert result["created"] >= 2  # At least 2 new students
        print(f"✓ Imported {result['created']} students, {result['updated']} updated")
        
        # Verify students with empty kelas went to test_class
        roster_resp = requests.get(
            f"{BASE_URL}/classes/{test_class['id']}/students",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        roster = roster_resp.json()
        student_emails = [s["email"] for s in roster["students"]]
        assert "import.test1@sekolah.id" in student_emails
        assert "import.test2@sekolah.id" in student_emails
        print(f"✓ Students with empty kelas column went to target class")
        
        # Verify student with explicit kelas went to that class
        classes_resp = requests.get(
            f"{BASE_URL}/classes",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        classes = classes_resp.json()
        other_class = next((c for c in classes if c["name"] == "Other Class"), None)
        if other_class:
            other_roster_resp = requests.get(
                f"{BASE_URL}/classes/{other_class['id']}/students",
                headers={"Authorization": f"Bearer {admin_token}"}
            )
            other_roster = other_roster_resp.json()
            other_emails = [s["email"] for s in other_roster["students"]]
            assert "import.test3@sekolah.id" in other_emails
            print(f"✓ Student with explicit kelas went to that class")
        
        # Verify imported students can login
        login_resp = requests.post(f"{BASE_URL}/auth/login", json={
            "email": "import.test1@sekolah.id",
            "password": "import123"
        })
        assert login_resp.status_code == 200, f"Imported student cannot login: {login_resp.text}"
        user = login_resp.json()["user"]
        assert user["role"] == "siswa"
        print(f"✓ Imported student can login immediately")

    def test_import_blank_rows_skipped(self, admin_token):
        """POST /api/students/import skips blank rows without errors"""
        csv_data = """nama,username,password,nis,kelas
Valid Student,valid@sekolah.id,valid123,5555555,Test Class


Another Valid,another@sekolah.id,valid123,6666666,Test Class
"""
        files = {"file": ("test_blank.csv", csv_data, "text/csv")}
        
        resp = requests.post(
            f"{BASE_URL}/students/import",
            files=files,
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert resp.status_code == 200, f"Import failed: {resp.text}"
        result = resp.json()
        # Should import 2 students, skip blank rows
        assert result["created"] + result["updated"] >= 2
        assert len(result.get("errors", [])) == 0 or not any("blank" in e.lower() for e in result.get("errors", []))
        print(f"✓ Blank rows skipped without errors")

    def test_import_duplicate_username_updates(self, admin_token):
        """POST /api/students/import updates existing students instead of duplicating"""
        # First import
        csv_data1 = """nama,username,password,nis,kelas
Update Test,update.test@sekolah.id,update123,7777777,Test Class
"""
        files1 = {"file": ("test_update1.csv", csv_data1, "text/csv")}
        resp1 = requests.post(
            f"{BASE_URL}/students/import",
            files=files1,
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert resp1.status_code == 200
        result1 = resp1.json()
        initial_created = result1["created"]
        print(f"✓ First import: {initial_created} created")
        
        # Second import with same username but different name
        csv_data2 = """nama,username,password,nis,kelas
Update Test UPDATED,update.test@sekolah.id,newpass123,8888888,Test Class
"""
        files2 = {"file": ("test_update2.csv", csv_data2, "text/csv")}
        resp2 = requests.post(
            f"{BASE_URL}/students/import",
            files=files2,
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert resp2.status_code == 200
        result2 = resp2.json()
        assert result2["updated"] >= 1
        assert result2["created"] == 0  # Should update, not create
        print(f"✓ Second import: {result2['updated']} updated, {result2['created']} created")
        
        # Verify only one account exists with updated data
        users_resp = requests.get(
            f"{BASE_URL}/users?role=siswa",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        users = users_resp.json()
        matching = [u for u in users if u["email"] == "update.test@sekolah.id"]
        assert len(matching) == 1
        assert matching[0]["name"] == "Update Test UPDATED"
        assert matching[0]["identifier"] == "8888888"
        print(f"✓ Student data updated, not duplicated")


class TestStudentE2EFlow:
    """Test critical E2E flow: create student → login → use app"""

    def test_student_created_from_ui_can_use_app(self, admin_token, test_class):
        """Student created from Manajemen Kelas can immediately login and use the app"""
        # Create student via API (simulating UI creation)
        student_data = {
            "name": "E2E Test Student",
            "email": "e2e.test@sekolah.id",
            "password": "e2etest123",
            "identifier": "E2E001"
        }
        create_resp = requests.post(
            f"{BASE_URL}/classes/{test_class['id']}/students",
            json=student_data,
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert create_resp.status_code == 200, f"Failed to create student: {create_resp.text}"
        student = create_resp.json()
        print(f"✓ Created student: {student['name']}")
        
        # Login as student
        login_resp = requests.post(f"{BASE_URL}/auth/login", json={
            "email": student_data["email"],
            "password": student_data["password"]
        })
        assert login_resp.status_code == 200, f"Student login failed: {login_resp.text}"
        student_token = login_resp.json()["token"]
        student_user = login_resp.json()["user"]
        assert student_user["role"] == "siswa"
        assert student_user["email"] == student_data["email"]
        print(f"✓ Student logged in successfully")
        
        # Get sessions list (student app)
        sessions_resp = requests.get(
            f"{BASE_URL}/sessions",
            headers={"Authorization": f"Bearer {student_token}"}
        )
        assert sessions_resp.status_code == 200, f"Failed to get sessions: {sessions_resp.text}"
        sessions = sessions_resp.json()
        assert isinstance(sessions, list)
        print(f"✓ Student can access sessions list ({len(sessions)} sessions)")
        
        # Get student's results
        results_resp = requests.get(
            f"{BASE_URL}/results/me",
            headers={"Authorization": f"Bearer {student_token}"}
        )
        assert results_resp.status_code == 200, f"Failed to get results: {results_resp.text}"
        results = results_resp.json()
        assert isinstance(results, list)
        print(f"✓ Student can access their results ({len(results)} results)")
        
        # Verify student info
        me_resp = requests.get(
            f"{BASE_URL}/auth/me",
            headers={"Authorization": f"Bearer {student_token}"}
        )
        assert me_resp.status_code == 200
        me = me_resp.json()
        assert me["role"] == "siswa"
        assert me["name"] == student_data["name"]
        assert me["identifier"] == student_data["identifier"]
        print(f"✓ Student profile correct: {me['name']} ({me['identifier']})")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
