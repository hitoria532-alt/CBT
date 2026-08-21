"""Iteration 14: bulk student import from Excel (Manajemen Kelas)."""
import io
import os
import uuid

import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "http://localhost:8001").rstrip("/")
API = f"{BASE_URL}/api"

ADMIN = {"email": "hitoria532@gmail.com", "password": "admin123"}
GURU = {"email": "guru@sekolah.id", "password": "guru123"}
SISWA = {"email": "ani@sekolah.id", "password": "siswa123"}
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _login(creds):
    r = requests.post(f"{API}/auth/login", json=creds, timeout=20)
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text}"
    return r.json()["token"]


def H(t):
    return {"Authorization": f"Bearer {t}"}


@pytest.fixture(scope="module")
def admin_token():
    return _login(ADMIN)


@pytest.fixture(scope="module")
def guru_token():
    return _login(GURU)


@pytest.fixture(scope="module")
def siswa_token():
    return _login(SISWA)


def build_workbook(rows):
    """Create an xlsx matching the template layout (headers on row 3)."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.active.title = "Petunjuk"
    wb["Petunjuk"]["A1"] = "petunjuk"
    ws = wb.create_sheet("Data Siswa")
    ws["A1"] = "DATA SISWA"
    ws["A2"] = "catatan"
    for i, h in enumerate(["nama", "kelas", "nis", "username", "password"], start=1):
        ws.cell(row=3, column=i, value=h)
    for ri, row in enumerate(rows, start=4):
        for ci, v in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def cleanup(admin_token, emails, class_names):
    users = requests.get(f"{API}/users", headers=H(admin_token), timeout=20).json()
    for u in users:
        if u["email"] in emails:
            requests.delete(f"{API}/users/{u['id']}", headers=H(admin_token), timeout=20)
    classes = requests.get(f"{API}/classes", headers=H(admin_token), timeout=20).json()
    for c in classes:
        if c["name"] in class_names:
            requests.delete(f"{API}/classes/{c['id']}", headers=H(admin_token), timeout=20)


# ------------------------------------------------------------------ template
class TestStudentTemplate:
    def test_template_download(self, admin_token):
        r = requests.get(f"{API}/students/import-template", headers=H(admin_token), timeout=30)
        assert r.status_code == 200, r.text
        assert r.headers["content-type"].startswith(XLSX_MIME)
        assert "template_data_siswa.xlsx" in r.headers.get("content-disposition", "")
        assert r.content[:2] == b"PK"

    def test_template_structure(self, admin_token):
        from openpyxl import load_workbook
        r = requests.get(f"{API}/students/import-template", headers=H(admin_token), timeout=30)
        wb = load_workbook(io.BytesIO(r.content))
        assert "Petunjuk" in wb.sheetnames and "Data Siswa" in wb.sheetnames

        ws = wb["Data Siswa"]
        header = [ws.cell(row=3, column=i).value for i in range(1, 6)]
        assert header == ["nama", "kelas", "nis", "username", "password"]
        # sample rows present and usable as a guide
        assert ws.cell(row=4, column=1).value
        assert "@" in str(ws.cell(row=4, column=4).value)
        assert ws.freeze_panes == "A4"
        # each header carries an explanatory comment
        for i in range(1, 6):
            assert ws.cell(row=3, column=i).comment is not None

        guide = wb["Petunjuk"]
        flat = " ".join(str(c) for row in guide.iter_rows(values_only=True)
                        for c in row if c is not None)
        assert "TEMPLATE IMPOR DATA SISWA" in flat
        assert "LANGKAH PENGISIAN" in flat and "KETERANGAN KOLOM" in flat
        for col in ["nama", "kelas", "nis", "username", "password"]:
            assert col in flat

    def test_guru_can_download_template(self, guru_token):
        r = requests.get(f"{API}/students/import-template", headers=H(guru_token), timeout=30)
        assert r.status_code == 200

    def test_siswa_forbidden(self, siswa_token):
        r = requests.get(f"{API}/students/import-template", headers=H(siswa_token), timeout=30)
        assert r.status_code == 403


# ------------------------------------------------------------------ import
class TestStudentImport:
    def test_creates_students_classes_and_login_works(self, admin_token):
        u = uuid.uuid4().hex[:6]
        cls_a, cls_b = f"TEST_I14_Kelas_A_{u}", f"TEST_I14_Kelas_B_{u}"
        e1, e2, e3 = (f"test_i14_a_{u}@sekolah.id", f"test_i14_b_{u}@sekolah.id",
                      f"test_i14_c_{u}@sekolah.id")
        data = build_workbook([
            (f"TEST_I14 Andi {u}", cls_a, "0091", e1, "rahasia123"),
            (f"TEST_I14 Bella {u}", cls_a, "0092", e2, "rahasia123"),
            (f"TEST_I14 Cahyo {u}", cls_b, "0093", e3, "rahasia123"),
        ])
        try:
            r = requests.post(f"{API}/students/import", headers=H(admin_token),
                              files={"file": (f"i14_{u}.xlsx", data, XLSX_MIME)}, timeout=90)
            assert r.status_code == 200, r.text
            body = r.json()
            assert body["created"] == 3, body
            assert body["updated"] == 0
            assert body["errors"] == []
            assert set(body["classes_created"]) == {cls_a, cls_b}
            assert body["added_to_class"] == 3

            # accounts exist with role siswa + NIS
            users = requests.get(f"{API}/users", headers=H(admin_token),
                                 params={"role": "siswa"}, timeout=20).json()
            by_email = {x["email"]: x for x in users}
            assert e1 in by_email and by_email[e1]["role"] == "siswa"
            assert by_email[e1]["identifier"] == "0091"

            # classes were created with the right members
            classes = requests.get(f"{API}/classes", headers=H(admin_token), timeout=20).json()
            ca = next(c for c in classes if c["name"] == cls_a)
            cb = next(c for c in classes if c["name"] == cls_b)
            assert ca["student_count"] == 2
            assert cb["student_count"] == 1
            assert by_email[e1]["id"] in ca["student_ids"]

            # the imported student can actually log in with the given password
            tok = _login({"email": e1, "password": "rahasia123"})
            me = requests.get(f"{API}/auth/me", headers=H(tok), timeout=20).json()
            assert me["role"] == "siswa"
            assert me["name"] == f"TEST_I14 Andi {u}"
        finally:
            cleanup(admin_token, {e1, e2, e3}, {cls_a, cls_b})

    def test_reimport_updates_instead_of_duplicating(self, admin_token):
        u = uuid.uuid4().hex[:6]
        cls = f"TEST_I14_Kelas_U_{u}"
        email = f"test_i14_upd_{u}@sekolah.id"
        try:
            first = build_workbook([(f"TEST_I14 Lama {u}", cls, "0101", email, "rahasia123")])
            r = requests.post(f"{API}/students/import", headers=H(admin_token),
                              files={"file": ("a.xlsx", first, XLSX_MIME)}, timeout=90)
            assert r.json()["created"] == 1, r.json()

            second = build_workbook([(f"TEST_I14 Baru {u}", cls, "0102", email, "sandibaru123")])
            r2 = requests.post(f"{API}/students/import", headers=H(admin_token),
                               files={"file": ("b.xlsx", second, XLSX_MIME)}, timeout=90)
            body = r2.json()
            assert body["created"] == 0, body
            assert body["updated"] == 1, body
            assert body["classes_created"] == []

            users = requests.get(f"{API}/users", headers=H(admin_token),
                                 params={"role": "siswa"}, timeout=20).json()
            matches = [x for x in users if x["email"] == email]
            assert len(matches) == 1, "must not duplicate the account"
            assert matches[0]["name"] == f"TEST_I14 Baru {u}"
            assert matches[0]["identifier"] == "0102"
            # new password works
            _login({"email": email, "password": "sandibaru123"})
        finally:
            cleanup(admin_token, {email}, {cls})

    def test_row_validation_errors(self, admin_token):
        u = uuid.uuid4().hex[:6]
        good = f"test_i14_ok_{u}@sekolah.id"
        dup = f"test_i14_dup_{u}@sekolah.id"
        cls = f"TEST_I14_Kelas_V_{u}"
        data = build_workbook([
            (f"TEST_I14 Ok {u}", cls, "0111", good, "rahasia123"),
            (f"TEST_I14 BadEmail {u}", cls, "0112", "bukanemail", "rahasia123"),
            ("", cls, "0113", f"test_i14_noname_{u}@sekolah.id", "rahasia123"),
            (f"TEST_I14 NoUser {u}", cls, "0114", "", "rahasia123"),
            (f"TEST_I14 ShortPw {u}", cls, "0115", f"test_i14_pw_{u}@sekolah.id", "123"),
            (f"TEST_I14 Dup1 {u}", cls, "0116", dup, "rahasia123"),
            (f"TEST_I14 Dup2 {u}", cls, "0117", dup, "rahasia123"),
        ])
        try:
            r = requests.post(f"{API}/students/import", headers=H(admin_token),
                              files={"file": ("v.xlsx", data, XLSX_MIME)}, timeout=90)
            assert r.status_code == 200, r.text
            body = r.json()
            assert body["created"] == 2, body          # good + first dup only
            assert len(body["errors"]) == 5, body
            blob = " | ".join(body["errors"]).lower()
            assert "harus berupa email" in blob
            assert "nama wajib diisi" in blob
            assert "username (email) wajib diisi" in blob
            assert "password minimal 5 karakter" in blob
            assert "dobel" in blob
        finally:
            cleanup(admin_token, {good, dup}, {cls})

    def test_student_without_class_is_created(self, admin_token):
        u = uuid.uuid4().hex[:6]
        email = f"test_i14_nocls_{u}@sekolah.id"
        try:
            data = build_workbook([(f"TEST_I14 Tanpa Kelas {u}", "", "0121", email, "rahasia123")])
            r = requests.post(f"{API}/students/import", headers=H(admin_token),
                              files={"file": ("n.xlsx", data, XLSX_MIME)}, timeout=90)
            body = r.json()
            assert body["created"] == 1, body
            assert body["added_to_class"] == 0
            assert body["classes_created"] == []
        finally:
            cleanup(admin_token, {email}, set())

    def test_csv_is_accepted(self, admin_token):
        u = uuid.uuid4().hex[:6]
        email = f"test_i14_csv_{u}@sekolah.id"
        cls = f"TEST_I14_Kelas_C_{u}"
        csv = ("nama,kelas,nis,username,password\n"
               f"TEST_I14 Csv {u},{cls},0131,{email},rahasia123\n")
        try:
            r = requests.post(f"{API}/students/import", headers=H(admin_token),
                              files={"file": (f"i14_{u}.csv", csv.encode(), "text/csv")},
                              timeout=90)
            assert r.status_code == 200, r.text
            assert r.json()["created"] == 1, r.json()
            assert r.json()["classes_created"] == [cls]
        finally:
            cleanup(admin_token, {email}, {cls})

    def test_column_aliases_accepted(self, admin_token):
        """'nama siswa'/'email'/'nisn'/'rombel' should map onto the real columns."""
        u = uuid.uuid4().hex[:6]
        email = f"test_i14_alias_{u}@sekolah.id"
        cls = f"TEST_I14_Kelas_AL_{u}"
        csv = ("nama siswa,rombel,nisn,email,password\n"
               f"TEST_I14 Alias {u},{cls},0141,{email},rahasia123\n")
        try:
            r = requests.post(f"{API}/students/import", headers=H(admin_token),
                              files={"file": (f"al_{u}.csv", csv.encode(), "text/csv")},
                              timeout=90)
            assert r.status_code == 200, r.text
            assert r.json()["created"] == 1, r.json()
        finally:
            cleanup(admin_token, {email}, {cls})

    def test_missing_required_columns_rejected(self, admin_token):
        csv = "foo,bar\n1,2\n"
        r = requests.post(f"{API}/students/import", headers=H(admin_token),
                          files={"file": ("bad.csv", csv.encode(), "text/csv")}, timeout=60)
        assert r.status_code == 400
        assert "nama" in r.json()["detail"].lower()

    def test_cannot_hijack_a_non_student_account(self, admin_token):
        """Importing an email that belongs to admin/guru must be refused."""
        csv = ("nama,kelas,nis,username,password\n"
               "TEST_I14 Hijack,,0151,guru@sekolah.id,rahasia123\n")
        r = requests.post(f"{API}/students/import", headers=H(admin_token),
                          files={"file": ("h.csv", csv.encode(), "text/csv")}, timeout=60)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["created"] == 0 and body["updated"] == 0, body
        assert any("guru" in e for e in body["errors"]), body
        # guru login still intact
        _login(GURU)

    def test_guru_cannot_import(self, guru_token):
        csv = "nama,kelas,nis,username,password\nX,,1,x@y.id,rahasia123\n"
        r = requests.post(f"{API}/students/import", headers=H(guru_token),
                          files={"file": ("g.csv", csv.encode(), "text/csv")}, timeout=60)
        assert r.status_code == 403

    def test_siswa_cannot_import(self, siswa_token):
        csv = "nama,kelas,nis,username,password\nX,,1,x@y.id,rahasia123\n"
        r = requests.post(f"{API}/students/import", headers=H(siswa_token),
                          files={"file": ("s.csv", csv.encode(), "text/csv")}, timeout=60)
        assert r.status_code == 403

    def test_unauth_blocked(self):
        csv = "nama,kelas,nis,username,password\nX,,1,x@y.id,rahasia123\n"
        r = requests.post(f"{API}/students/import",
                          files={"file": ("u.csv", csv.encode(), "text/csv")}, timeout=60)
        assert r.status_code == 401

    def test_imported_student_sees_only_their_class_sessions(self, admin_token):
        """End-to-end: an imported student joins a class and the session filter works."""
        u = uuid.uuid4().hex[:6]
        email = f"test_i14_e2e_{u}@sekolah.id"
        cls = f"TEST_I14_Kelas_E2E_{u}"
        try:
            data = build_workbook([(f"TEST_I14 E2E {u}", cls, "0161", email, "rahasia123")])
            r = requests.post(f"{API}/students/import", headers=H(admin_token),
                              files={"file": ("e.xlsx", data, XLSX_MIME)}, timeout=90)
            assert r.json()["created"] == 1, r.json()
            stok = _login({"email": email, "password": "rahasia123"})
            # seeded sessions target every class (class_ids empty) -> visible
            sessions = requests.get(f"{API}/sessions", headers=H(stok), timeout=20).json()
            assert any(s["title"] == "UH Matematika - Kelas X" for s in sessions), sessions
            # leaderboard for their own class resolves
            lb = requests.get(f"{API}/leaderboard/me", headers=H(stok), timeout=20)
            assert lb.status_code == 200
            assert any(x["class_name"] == cls for x in lb.json()), lb.json()
        finally:
            cleanup(admin_token, {email}, {cls})
