"""Iteration 17 — Ekspor akun login siswa (Excel + kartu login PDF).

Akun yang tercantum pada berkas ekspor dijamin aktif dan bisa langsung dipakai login.
"""
import io
import os

import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "http://localhost:8001").rstrip("/")
API = f"{BASE_URL}/api"

ADMIN = {"email": "hitoria532@gmail.com", "password": "admin123"}
GURU = {"email": "guru@sekolah.id", "password": "guru123"}
STUDENT_PASSWORD = "siswa123"
XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _login(creds):
    r = requests.post(f"{API}/auth/login", json=creds, timeout=20)
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text}"
    return r.json()["token"]


def H(tok):
    return {"Authorization": f"Bearer {tok}"}


def _rows(content):
    """Baris data (setelah baris header 'No') dari workbook akun."""
    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(content)).active
    grid = [[c.value for c in row] for row in ws.iter_rows()]
    head = next(i for i, r in enumerate(grid) if r and r[0] == "No")
    return grid[head], [r for r in grid[head + 1:] if r and r[0] is not None]


@pytest.fixture(scope="module")
def admin():
    return _login(ADMIN)


@pytest.fixture(scope="module")
def klass(admin, worker_id):
    """Kelas + siswa milik test ini sendiri.

    Wajib terisolasi: ekspor dapat me-reset password, jadi jangan pernah menyentuh
    akun demo yang dipakai suite lain. Nama disuffix `worker_id` karena pytest-xdist
    menjalankan fixture module-scope sekali per worker.
    """
    tag = worker_id or "gw0"
    students = []
    for i in (1, 2, 3):
        email = f"pytest.akun{tag}{i}@sekolah.id"
        for u in requests.get(f"{API}/users", params={"role": "siswa"},
                              headers=H(admin), timeout=20).json():
            if u["email"] == email:
                requests.delete(f"{API}/users/{u['id']}", headers=H(admin), timeout=20)
        r = requests.post(f"{API}/users", headers=H(admin), json={
            "email": email, "password": STUDENT_PASSWORD, "name": f"Pytest Akun {tag}{i}",
            "role": "siswa", "identifier": f"P8{tag}{i}"}, timeout=20)
        assert r.status_code == 200, r.text
        students.append(r.json()["id"])
    cls = requests.post(f"{API}/classes", headers=H(admin), json={
        "name": f"[PYTEST] Kelas Akun {tag}", "description": "kelas uji ekspor akun",
        "student_ids": students}, timeout=20).json()
    yield cls
    requests.delete(f"{API}/classes/{cls['id']}", headers=H(admin), timeout=20)
    for sid in students:
        requests.delete(f"{API}/users/{sid}", headers=H(admin), timeout=20)


class TestAccountExportXlsx:
    def test_headers_and_all_rows_can_login(self, admin, klass):
        r = requests.get(f"{API}/export/class/{klass['id']}/accounts/xlsx",
                         headers=H(admin), timeout=40)
        assert r.status_code == 200, r.text[:200]
        assert r.headers["content-type"].startswith(XLSX_MEDIA)
        header, rows = _rows(r.content)
        assert header[:7] == ["No", "Nama", "NIS", "Kelas", "Username (Email)",
                              "Password", "Status Akun"]
        assert len(rows) == len(klass["student_ids"])
        for row in rows:
            assert row[3] == klass["name"]
            assert row[5], "kolom Password tidak boleh kosong"
            assert "Aktif" in row[6]
            # inti permintaan: kredensial di berkas ekspor benar-benar bisa login
            login = requests.post(f"{API}/auth/login",
                                  json={"email": row[4], "password": row[5]}, timeout=20)
            assert login.status_code == 200, f"{row[4]} tidak bisa login: {login.text[:120]}"

    def test_second_export_keeps_same_password(self, admin, klass):
        a = _rows(requests.get(f"{API}/export/class/{klass['id']}/accounts/xlsx",
                               headers=H(admin), timeout=40).content)[1]
        b = _rows(requests.get(f"{API}/export/class/{klass['id']}/accounts/xlsx",
                               headers=H(admin), timeout=40).content)[1]
        assert [x[5] for x in a] == [x[5] for x in b]
        assert all("Tersimpan" in x[6] for x in b)

    def test_reset_changes_password_and_invalidates_old(self, admin, klass):
        before = _rows(requests.get(f"{API}/export/class/{klass['id']}/accounts/xlsx",
                                    headers=H(admin), timeout=40).content)[1]
        after = _rows(requests.get(f"{API}/export/class/{klass['id']}/accounts/xlsx",
                                   params={"reset": "true"}, headers=H(admin), timeout=40).content)[1]
        assert [x[5] for x in after] != [x[5] for x in before]
        assert all("Baru" in x[6] for x in after)
        for row in after:
            assert requests.post(f"{API}/auth/login",
                                 json={"email": row[4], "password": row[5]},
                                 timeout=20).status_code == 200
        old = before[0]
        assert requests.post(f"{API}/auth/login",
                             json={"email": old[4], "password": old[5]},
                             timeout=20).status_code == 401

    def test_all_classes_workbook_has_sheet_per_class(self, admin, klass):
        from openpyxl import load_workbook
        r = requests.get(f"{API}/export/accounts/xlsx", headers=H(admin), timeout=60)
        assert r.status_code == 200, r.text[:200]
        wb = load_workbook(io.BytesIO(r.content))
        # Jangan bandingkan dengan jumlah kelas saat itu: modul lain / worker xdist lain
        # bisa menambah-hapus kelas secara bersamaan. Cukup pastikan tiap kelas punya
        # sheet-nya sendiri dan kelas milik test ini ikut terekspor.
        assert len(wb.sheetnames) >= 1
        assert len(wb.sheetnames) == len(set(wb.sheetnames)), "nama sheet harus unik"
        assert any(klass["name"].strip("[]").split("]")[-1].strip() in s or s in klass["name"]
                   for s in wb.sheetnames), wb.sheetnames


class TestLoginCardPdf:
    def test_class_cards_pdf(self, admin, klass):
        r = requests.get(f"{API}/export/class/{klass['id']}/accounts/pdf",
                         headers=H(admin), timeout=60)
        assert r.status_code == 200, r.text[:200]
        assert r.headers["content-type"] == "application/pdf"
        assert r.content[:5] == b"%PDF-"
        assert len(r.content) > 1000

    def test_all_classes_cards_pdf(self, admin, klass):
        r = requests.get(f"{API}/export/accounts/pdf", headers=H(admin), timeout=60)
        assert r.status_code == 200
        assert r.content[:5] == b"%PDF-"


class TestAccountExportGuards:
    def test_unknown_class_404(self, admin):
        assert requests.get(f"{API}/export/class/tidak-ada/accounts/xlsx",
                            headers=H(admin), timeout=20).status_code == 404
        assert requests.get(f"{API}/export/class/tidak-ada/accounts/pdf",
                            headers=H(admin), timeout=20).status_code == 404

    def test_guru_and_student_forbidden(self, klass):
        gtok = _login(GURU)
        for url in (f"{API}/export/class/{klass['id']}/accounts/xlsx",
                    f"{API}/export/class/{klass['id']}/accounts/pdf",
                    f"{API}/export/accounts/xlsx",
                    f"{API}/export/accounts/pdf"):
            assert requests.get(url, headers=H(gtok), timeout=20).status_code == 403, url

    def test_unauthenticated_forbidden(self, klass):
        r = requests.get(f"{API}/export/class/{klass['id']}/accounts/xlsx", timeout=20)
        assert r.status_code in (401, 403)


class TestInitialPasswordPrivacy:
    def test_initial_password_never_exposed(self, admin):
        users = requests.get(f"{API}/users", params={"role": "siswa"},
                             headers=H(admin), timeout=20).json()
        assert users, "butuh minimal satu siswa"
        for u in users:
            assert not [k for k in u if "password" in k.lower()], u.keys()

    def test_new_account_password_is_reused_on_export(self, admin, klass, worker_id):
        """Akun yang baru dibuat tidak perlu di-reset saat diekspor."""
        email = f"pytest.kartu{worker_id or 'gw0'}@sekolah.id"
        existing = [u for u in requests.get(f"{API}/users", params={"role": "siswa"},
                                            headers=H(admin), timeout=20).json()
                    if u["email"] == email]
        for u in existing:
            requests.delete(f"{API}/users/{u['id']}", headers=H(admin), timeout=20)
        created = requests.post(f"{API}/users", headers=H(admin), json={
            "email": email, "password": "rahasia123", "name": "Pytest Kartu",
            "role": "siswa", "identifier": "P9001"}, timeout=20)
        assert created.status_code == 200, created.text
        uid = created.json()["id"]
        cls_students = list(klass["student_ids"]) + [uid]
        requests.put(f"{API}/classes/{klass['id']}", headers=H(admin), json={
            "name": klass["name"], "description": klass.get("description", ""),
            "student_ids": cls_students}, timeout=20)
        try:
            _, rows = _rows(requests.get(f"{API}/export/class/{klass['id']}/accounts/xlsx",
                                         headers=H(admin), timeout=40).content)
            row = next(r for r in rows if r[4] == email)
            assert row[5] == "rahasia123"           # password asli dipakai apa adanya
            assert "Tersimpan" in row[6]
            assert requests.post(f"{API}/auth/login",
                                 json={"email": email, "password": "rahasia123"},
                                 timeout=20).status_code == 200
        finally:
            requests.put(f"{API}/classes/{klass['id']}", headers=H(admin), json={
                "name": klass["name"], "description": klass.get("description", ""),
                "student_ids": klass["student_ids"]}, timeout=20)
            requests.delete(f"{API}/users/{uid}", headers=H(admin), timeout=20)
