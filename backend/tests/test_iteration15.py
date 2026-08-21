"""Iteration 15: Mode Ujian Ketat (exam lockdown) — violations + auto submit."""
import io
import os
import time
import uuid
from datetime import datetime, timedelta, timezone

import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "http://localhost:8001").rstrip("/")
API = f"{BASE_URL}/api"
ADMIN = {"email": "hitoria532@gmail.com", "password": "admin123"}
SISWA = {"email": "ani@sekolah.id", "password": "siswa123"}


def _login(c):
    r = requests.post(f"{API}/auth/login", json=c, timeout=20)
    assert r.status_code == 200, r.text
    return r.json()["token"]


def H(t):
    return {"Authorization": f"Bearer {t}"}


@pytest.fixture(scope="module")
def admin_token():
    return _login(ADMIN)


@pytest.fixture(scope="module")
def lock_cfg(admin_token):
    """Pin the limit to 3 so the tests are order independent."""
    r = requests.put(f"{API}/settings/exam-lock", headers=H(admin_token),
                     json={"enabled": True, "max_violations": 3}, timeout=20)
    assert r.status_code == 200, r.text
    yield r.json()
    requests.put(f"{API}/settings/exam-lock", headers=H(admin_token),
                 json={"enabled": True, "max_violations": 3}, timeout=20)


def make_live_session(admin_token, tag):
    cat = requests.post(f"{API}/categories", headers=H(admin_token),
                        json={"name": f"TEST_I15_{tag}"}, timeout=20).json()
    q = requests.post(f"{API}/questions", headers=H(admin_token), json={
        "category_id": cat["id"], "type": "pg", "text": f"TEST_I15_Q_{tag}",
        "options": ["a", "b", "c", "d", "e"], "correct_answer": "0", "weight": 1.0},
        timeout=20).json()
    pkg = requests.post(f"{API}/packages", headers=H(admin_token), json={
        "title": f"TEST_I15_PKG_{tag}", "category_id": cat["id"],
        "question_ids": [q["id"]], "scoring_method": "percentage"}, timeout=20).json()
    ses = requests.post(f"{API}/sessions", headers=H(admin_token), json={
        "title": f"TEST_I15_SES_{tag}", "package_id": pkg["id"],
        "start_time": (datetime.now(timezone.utc) - timedelta(minutes=5)).isoformat(),
        "end_time": (datetime.now(timezone.utc) + timedelta(hours=2)).isoformat(),
        "duration_minutes": 60, "kkm": 75.0, "class_ids": []}, timeout=20).json()
    email = f"test_i15_{tag}_{int(time.time())}@example.com"
    stu = requests.post(f"{API}/users", headers=H(admin_token), json={
        "email": email, "password": "pw12345", "name": f"TEST I15 {tag}",
        "role": "siswa"}, timeout=20).json()
    return {"cat": cat, "q": q, "pkg": pkg, "ses": ses, "stu": stu, "email": email}


def teardown(admin_token, ctx):
    for path, key in [("users", "stu"), ("sessions", "ses"), ("packages", "pkg"),
                      ("questions", "q"), ("categories", "cat")]:
        requests.delete(f"{API}/{path}/{ctx[key]['id']}", headers=H(admin_token), timeout=20)


class TestExamLockSettings:
    def test_default_and_update(self, admin_token):
        r = requests.get(f"{API}/settings/exam-lock", headers=H(admin_token), timeout=20)
        assert r.status_code == 200
        assert "enabled" in r.json() and "max_violations" in r.json()
        r2 = requests.put(f"{API}/settings/exam-lock", headers=H(admin_token),
                          json={"enabled": True, "max_violations": 5}, timeout=20)
        assert r2.status_code == 200 and r2.json()["max_violations"] == 5
        # clamped to 1..20
        r3 = requests.put(f"{API}/settings/exam-lock", headers=H(admin_token),
                          json={"enabled": True, "max_violations": 999}, timeout=20)
        assert r3.json()["max_violations"] == 20
        requests.put(f"{API}/settings/exam-lock", headers=H(admin_token),
                     json={"enabled": True, "max_violations": 3}, timeout=20)

    def test_student_can_read_config(self):
        tok = _login(SISWA)
        r = requests.get(f"{API}/settings/exam-lock", headers=H(tok), timeout=20)
        assert r.status_code == 200
        assert r.json()["max_violations"] >= 1

    def test_student_cannot_change_config(self):
        tok = _login(SISWA)
        r = requests.put(f"{API}/settings/exam-lock", headers=H(tok),
                         json={"enabled": False, "max_violations": 99}, timeout=20)
        assert r.status_code == 403

    def test_unauth_blocked(self):
        assert requests.get(f"{API}/settings/exam-lock", timeout=20).status_code == 401


class TestViolations:
    def test_exam_start_exposes_lock_config(self, admin_token, lock_cfg):
        ctx = make_live_session(admin_token, uuid.uuid4().hex[:5])
        try:
            stok = _login({"email": ctx["email"], "password": "pw12345"})
            r = requests.post(f"{API}/exam/start", headers=H(stok),
                              json={"session_id": ctx["ses"]["id"]}, timeout=30)
            assert r.status_code == 200, r.text
            body = r.json()
            assert body["lock"]["enabled"] is True
            assert body["lock"]["max_violations"] == 3
            assert body["violations"] == 0
        finally:
            teardown(admin_token, ctx)

    def test_violations_accumulate_then_auto_submit(self, admin_token, lock_cfg):
        ctx = make_live_session(admin_token, uuid.uuid4().hex[:5])
        sid = ctx["ses"]["id"]
        try:
            stok = _login({"email": ctx["email"], "password": "pw12345"})
            requests.post(f"{API}/exam/start", headers=H(stok),
                          json={"session_id": sid}, timeout=30)
            for expected in (1, 2):
                r = requests.post(f"{API}/exam/violation", headers=H(stok),
                                  json={"session_id": sid, "type": "tab_hidden"}, timeout=30)
                assert r.status_code == 200, r.text
                assert r.json()["count"] == expected
                assert r.json()["auto_submitted"] is False
            r3 = requests.post(f"{API}/exam/violation", headers=H(stok),
                               json={"session_id": sid, "type": "fullscreen_exit"}, timeout=30)
            assert r3.json()["count"] == 3
            assert r3.json()["auto_submitted"] is True

            mine = requests.get(f"{API}/results/me", headers=H(stok), timeout=20).json()
            att = next(a for a in mine if a["session_id"] == sid)
            assert att["status"] in ("selesai", "menunggu_koreksi")
            assert len(att["violations"]) == 3
            assert att["auto_submitted_reason"] == "pelanggaran"
            labels = [v["label"] for v in att["violations"]]
            assert "Pindah tab / minimize" in labels
            assert "Keluar dari layar penuh" in labels
            # further reports are a no-op, not an error
            r4 = requests.post(f"{API}/exam/violation", headers=H(stok),
                               json={"session_id": sid, "type": "tab_hidden"}, timeout=30)
            assert r4.status_code == 200
            assert r4.json().get("already_submitted") is True
        finally:
            teardown(admin_token, ctx)

    def test_teacher_sees_violations_in_results(self, admin_token, lock_cfg):
        ctx = make_live_session(admin_token, uuid.uuid4().hex[:5])
        sid = ctx["ses"]["id"]
        try:
            stok = _login({"email": ctx["email"], "password": "pw12345"})
            requests.post(f"{API}/exam/start", headers=H(stok),
                          json={"session_id": sid}, timeout=30)
            requests.post(f"{API}/exam/violation", headers=H(stok),
                          json={"session_id": sid, "type": "tab_hidden"}, timeout=30)
            res = requests.get(f"{API}/results/session/{sid}", headers=H(admin_token),
                               timeout=20).json()
            att = res["attempts"][0]
            assert len(att["violations"]) == 1
            assert att["violations"][0]["label"] == "Pindah tab / minimize"
            assert att["violations"][0]["at"]
        finally:
            teardown(admin_token, ctx)

    def test_violations_appear_in_excel_export(self, admin_token, lock_cfg):
        from openpyxl import load_workbook
        ctx = make_live_session(admin_token, uuid.uuid4().hex[:5])
        sid = ctx["ses"]["id"]
        try:
            stok = _login({"email": ctx["email"], "password": "pw12345"})
            requests.post(f"{API}/exam/start", headers=H(stok),
                          json={"session_id": sid}, timeout=30)
            requests.post(f"{API}/exam/violation", headers=H(stok),
                          json={"session_id": sid, "type": "copy_attempt"}, timeout=30)
            r = requests.get(f"{API}/export/session/{sid}/xlsx", headers=H(admin_token), timeout=60)
            assert r.status_code == 200, r.text
            wb = load_workbook(io.BytesIO(r.content))
            assert "Pelanggaran" in wb.sheetnames
            rekap = [str(c) for row in wb["Rekap Nilai"].iter_rows(values_only=True)
                     for c in row if c is not None]
            assert "Pelanggaran" in rekap
            log = [str(c) for row in wb["Pelanggaran"].iter_rows(values_only=True)
                   for c in row if c is not None]
            for h in ["Nama Siswa", "Pelanggaran ke-", "Jenis Pelanggaran", "Waktu",
                      "Dikumpulkan Otomatis"]:
                assert h in log, f"missing {h}"
            assert any("Mencoba menyalin teks" in x for x in log)
        finally:
            teardown(admin_token, ctx)

    def test_violation_without_attempt_is_404(self):
        tok = _login(SISWA)
        r = requests.post(f"{API}/exam/violation", headers=H(tok),
                          json={"session_id": "tidak-ada", "type": "tab_hidden"}, timeout=20)
        assert r.status_code == 404

    def test_admin_cannot_post_violation(self, admin_token):
        r = requests.post(f"{API}/exam/violation", headers=H(admin_token),
                          json={"session_id": "x", "type": "tab_hidden"}, timeout=20)
        assert r.status_code == 403
