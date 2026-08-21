from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import io
import re
import hmac
import base64
import gzip
import json
import asyncio
import logging
import uuid
import random
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Any

import jwt
import bcrypt
import requests
import pandas as pd
from bson import ObjectId
from fastapi import (FastAPI, APIRouter, HTTPException, Depends, Request, Response,
                     UploadFile, File, Form, BackgroundTasks, Header, Query)
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr

# ------------------------------------------------------------------ DB setup
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

JWT_ALGORITHM = "HS256"

# ------------------------------------------------------------------ object storage
# Two backends so the app runs both on Emergent and on self-hosted platforms
# (Render/Railway/VPS) where the Emergent object-storage proxy is unavailable:
#   STORAGE_MODE=auto (default) -> Emergent proxy when EMERGENT_LLM_KEY is set,
#                                  otherwise files are kept in MongoDB.
#   STORAGE_MODE=emergent|mongo -> force one backend.
STORAGE_BASE = (os.environ.get("INTEGRATION_PROXY_URL") or "").strip() or "https://integrations.emergentagent.com"
STORAGE_URL = STORAGE_BASE.rstrip("/") + "/objstore/api/v1/storage"
EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY")
STORAGE_MODE = (os.environ.get("STORAGE_MODE") or "auto").strip().lower()
APP_NAME = "cbt-ujian"
_storage_key = None


def use_emergent_storage() -> bool:
    if STORAGE_MODE == "mongo":
        return False
    if STORAGE_MODE == "emergent":
        return True
    return bool(EMERGENT_KEY)


def init_storage(force: bool = False):
    global _storage_key
    if _storage_key and not force:
        return _storage_key
    resp = requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_KEY}, timeout=30)
    resp.raise_for_status()
    _storage_key = resp.json()["storage_key"]
    return _storage_key


def put_object(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    resp = requests.put(f"{STORAGE_URL}/objects/{path}",
                        headers={"X-Storage-Key": key, "Content-Type": content_type},
                        data=data, timeout=120)
    if resp.status_code == 404:
        key = init_storage(force=True)
        resp = requests.put(f"{STORAGE_URL}/objects/{path}",
                            headers={"X-Storage-Key": key, "Content-Type": content_type},
                            data=data, timeout=120)
    resp.raise_for_status()
    return resp.json()


def get_object(path: str):
    key = init_storage()
    resp = requests.get(f"{STORAGE_URL}/objects/{path}", headers={"X-Storage-Key": key}, timeout=60)
    if resp.status_code == 404:
        key = init_storage(force=True)
        resp = requests.get(f"{STORAGE_URL}/objects/{path}", headers={"X-Storage-Key": key}, timeout=60)
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")


async def store_file(path: str, data: bytes, content_type: str) -> dict:
    """Save an uploaded file with whichever storage backend is available."""
    if use_emergent_storage():
        try:
            return put_object(path, data, content_type)
        except Exception as e:
            if STORAGE_MODE == "emergent":
                raise
            logging.getLogger(__name__).warning(
                "Object storage unavailable (%s) — falling back to MongoDB", e)
    from bson.binary import Binary
    await db.file_blobs.replace_one(
        {"path": path},
        {"path": path, "data": Binary(data), "content_type": content_type,
         "size": len(data), "created_at": now_iso()},
        upsert=True)
    return {"path": path, "size": len(data)}


async def load_file(path: str):
    """Read a stored file: MongoDB copy first, then the Emergent proxy."""
    doc = await db.file_blobs.find_one({"path": path})
    if doc is not None:
        return bytes(doc["data"]), doc.get("content_type", "application/octet-stream")
    if use_emergent_storage():
        return get_object(path)
    raise FileNotFoundError(path)



def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]


# ------------------------------------------------------------------ helpers
def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode("utf-8"), salt).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))


def create_access_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id, "email": email, "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "access",
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def new_id() -> str:
    return str(uuid.uuid4())


def clean_user(u: dict) -> dict:
    u = dict(u)
    u["id"] = str(u.pop("_id"))
    u.pop("password_hash", None)
    return u


# ------------------------------------------------------------------ auth dep
async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return clean_user(user)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


def require_roles(*roles):
    async def checker(user: dict = Depends(get_current_user)) -> dict:
        if user["role"] not in roles:
            raise HTTPException(status_code=403, detail="Akses ditolak")
        return user
    return checker


# ------------------------------------------------------------------ models
class LoginBody(BaseModel):
    email: EmailStr
    password: str


class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str  # siswa | guru | admin
    identifier: Optional[str] = None  # NISN / NIP


class UserUpdate(BaseModel):
    name: Optional[str] = None
    password: Optional[str] = None
    role: Optional[str] = None
    identifier: Optional[str] = None


class Category(BaseModel):
    id: str = Field(default_factory=new_id)
    name: str
    description: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class CategoryBody(BaseModel):
    name: str
    description: Optional[str] = ""


class Question(BaseModel):
    id: str = Field(default_factory=new_id)
    category_id: Optional[str] = None
    type: str  # pg | truefalse | essay
    text: str
    options: List[str] = []          # for pg
    correct_answer: Optional[str] = None  # index string for pg, "true"/"false" for tf
    weight: float = 1.0
    image_path: Optional[str] = None
    created_at: str = Field(default_factory=now_iso)


class QuestionBody(BaseModel):
    category_id: Optional[str] = None
    type: str
    text: str
    options: List[str] = []
    correct_answer: Optional[str] = None
    weight: float = 1.0
    image_path: Optional[str] = None


class Package(BaseModel):
    id: str = Field(default_factory=new_id)
    title: str
    description: Optional[str] = ""
    category_id: Optional[str] = None
    question_ids: List[str] = []
    scoring_method: str = "percentage"  # percentage | weighted
    shuffle_questions: bool = False
    shuffle_options: bool = False
    min_score: float = 0.0
    rounding: str = "2desimal"  # 2desimal | 1desimal | bulat
    easy_min: Optional[float] = None
    medium_min: Optional[float] = None
    created_by: Optional[str] = None
    is_public: bool = False
    created_at: str = Field(default_factory=now_iso)


class PackageBody(BaseModel):
    title: str
    description: Optional[str] = ""
    category_id: Optional[str] = None
    question_ids: List[str] = []
    scoring_method: str = "percentage"
    shuffle_questions: bool = False
    shuffle_options: bool = False
    min_score: float = 0.0
    rounding: str = "2desimal"
    easy_min: Optional[float] = None
    medium_min: Optional[float] = None
    is_public: bool = False


class Session(BaseModel):
    id: str = Field(default_factory=new_id)
    title: str
    package_id: str
    start_time: str
    end_time: str
    duration_minutes: int = 60
    kkm: float = 75.0
    class_ids: List[str] = []
    announcement: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class SessionBody(BaseModel):
    title: str
    package_id: str
    start_time: str
    end_time: str
    duration_minutes: int = 60
    kkm: float = 75.0
    class_ids: List[str] = []
    announcement: Optional[str] = ""


class MakeupBody(BaseModel):
    """Jadwalkan ujian susulan untuk satu atau beberapa siswa pada sebuah sesi."""
    session_id: str
    student_ids: List[str]
    start_time: str
    end_time: str
    duration_minutes: Optional[int] = None  # None = ikut durasi sesi induk
    reason: Optional[str] = ""


class MakeupUpdateBody(BaseModel):
    start_time: str
    end_time: str
    duration_minutes: Optional[int] = None
    reason: Optional[str] = ""


class SchoolClass(BaseModel):
    id: str = Field(default_factory=new_id)
    name: str
    description: Optional[str] = ""
    student_ids: List[str] = []
    created_at: str = Field(default_factory=now_iso)


class ClassBody(BaseModel):
    name: str
    description: Optional[str] = ""
    student_ids: List[str] = []


class DifficultyBody(BaseModel):
    easy_min: float = 70.0
    medium_min: float = 40.0


DEFAULT_THEME_COLOR = "157 35% 18%"
THEME_HSL_RE = re.compile(r"^-?\d+(\.\d+)?\s+\d+(\.\d+)?%\s+\d+(\.\d+)?%$")
LEGACY_THEME_MAP = {
    "green": "157 35% 18%", "#1e3a30": "157 35% 18%",
    "blue": "215 60% 30%", "#1f4e8c": "215 60% 30%",
    "purple": "265 45% 40%", "#5b3a94": "265 45% 40%",
    "red": "0 60% 40%", "#a32626": "0 60% 40%",
    "teal": "200 70% 30%", "#106688": "200 70% 30%",
    "brown": "25 60% 35%", "#8a4b1e": "25 60% 35%",
}


def sanitize_theme_color(value: Optional[str]) -> Optional[str]:
    """Only allow a Tailwind HSL triplet ('157 35% 18%'); map known legacy names.

    An invalid value (e.g. 'green') would produce `hsl(green)` in the browser and
    make every element that uses --primary render as blank/white text.
    """
    if not isinstance(value, str):
        return None
    v = value.strip()
    if not v:
        return None
    if THEME_HSL_RE.match(v):
        return v
    return LEGACY_THEME_MAP.get(v.lower())


class SchoolBody(BaseModel):
    name: str = ""
    address: str = ""
    logo_path: Optional[str] = None
    theme_color: Optional[str] = None


class StartAttemptBody(BaseModel):
    session_id: str


class SubmitBody(BaseModel):
    session_id: str
    answers: dict  # {question_id: answer}


class GradeEssayBody(BaseModel):
    scores: dict  # {question_id: points_earned}


# ------------------------------------------------------------------ AUTH routes
def set_auth_cookie(response: Response, token: str):
    response.set_cookie(key="access_token", value=token, httponly=True,
                        secure=True, samesite="none", max_age=604800, path="/")


@api_router.post("/auth/login")
async def login(body: LoginBody, response: Response):
    email = body.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Email atau password salah")
    token = create_access_token(str(user["_id"]), email, user["role"])
    set_auth_cookie(response, token)
    return {"token": token, "user": clean_user(user)}


@api_router.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user


@api_router.post("/auth/logout")
async def logout(response: Response, user: dict = Depends(get_current_user)):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}


# ------------------------------------------------------------------ ACCOUNTS (admin)
@api_router.get("/users")
async def list_users(role: Optional[str] = None, user: dict = Depends(require_roles("admin", "guru"))):
    q = {}
    if role:
        q["role"] = role
    users = await db.users.find(q).sort("created_at", -1).to_list(1000)
    return [clean_user(u) for u in users]


@api_router.post("/users")
async def create_user(body: UserCreate, user: dict = Depends(require_roles("admin"))):
    if body.role not in ("siswa", "guru", "admin"):
        raise HTTPException(status_code=400, detail="Role tidak valid")
    email = body.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email sudah terdaftar")
    doc = {
        "email": email, "password_hash": hash_password(body.password),
        "name": body.name, "role": body.role, "identifier": body.identifier or "",
        "created_at": now_iso(),
    }
    res = await db.users.insert_one(doc)
    doc["_id"] = res.inserted_id
    return clean_user(doc)


@api_router.put("/users/{user_id}")
async def update_user(user_id: str, body: UserUpdate, user: dict = Depends(require_roles("admin"))):
    update = {}
    if body.name is not None:
        update["name"] = body.name
    if body.role is not None:
        update["role"] = body.role
    if body.identifier is not None:
        update["identifier"] = body.identifier
    if body.password:
        update["password_hash"] = hash_password(body.password)
    if update:
        await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": update})
    u = await db.users.find_one({"_id": ObjectId(user_id)})
    return clean_user(u)


@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, user: dict = Depends(require_roles("admin"))):
    target = await db.users.find_one({"_id": ObjectId(user_id)}, {"_id": 0, "id": 1})
    await db.users.delete_one({"_id": ObjectId(user_id)})
    if target and target.get("id"):
        await db.makeups.delete_many({"student_id": target["id"]})
    return {"ok": True}


# ------------------------------------------------------------------ CATEGORIES
@api_router.get("/categories")
async def list_categories(user: dict = Depends(get_current_user)):
    return await db.categories.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)


@api_router.post("/categories")
async def create_category(body: CategoryBody, user: dict = Depends(require_roles("admin", "guru"))):
    cat = Category(**body.model_dump())
    await db.categories.insert_one(cat.model_dump())
    return cat.model_dump()


@api_router.put("/categories/{cid}")
async def update_category(cid: str, body: CategoryBody, user: dict = Depends(require_roles("admin", "guru"))):
    await db.categories.update_one({"id": cid}, {"$set": body.model_dump()})
    return await db.categories.find_one({"id": cid}, {"_id": 0})


@api_router.delete("/categories/{cid}")
async def delete_category(cid: str, user: dict = Depends(require_roles("admin", "guru"))):
    await db.categories.delete_one({"id": cid})
    return {"ok": True}


# ------------------------------------------------------------------ QUESTIONS
@api_router.get("/questions")
async def list_questions(category_id: Optional[str] = None, user: dict = Depends(require_roles("admin", "guru"))):
    q = {}
    if category_id:
        q["category_id"] = category_id
    return await db.questions.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)


@api_router.post("/questions")
async def create_question(body: QuestionBody, user: dict = Depends(require_roles("admin", "guru"))):
    ques = Question(**body.model_dump())
    await db.questions.insert_one(ques.model_dump())
    return ques.model_dump()


@api_router.put("/questions/{qid}")
async def update_question(qid: str, body: QuestionBody, user: dict = Depends(require_roles("admin", "guru"))):
    await db.questions.update_one({"id": qid}, {"$set": body.model_dump()})
    return await db.questions.find_one({"id": qid}, {"_id": 0})


@api_router.delete("/questions/{qid}")
async def delete_question(qid: str, user: dict = Depends(require_roles("admin", "guru"))):
    await db.questions.delete_one({"id": qid})
    return {"ok": True}


# ------------------------------------------------------------------ PACKAGES
@api_router.get("/packages")
async def list_packages(user: dict = Depends(require_roles("admin", "guru"))):
    if user["role"] == "admin":
        q = {}
    else:
        q = {"$or": [{"created_by": user["id"]}, {"is_public": True},
                     {"created_by": None}, {"created_by": {"$exists": False}}]}
    pkgs = await db.packages.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)
    owner_ids = {p.get("created_by") for p in pkgs if p.get("created_by")}
    owners = {}
    if owner_ids:
        docs = await db.users.find({"_id": {"$in": [ObjectId(o) for o in owner_ids]}}).to_list(2000)
        owners = {str(d["_id"]): d["name"] for d in docs}
    for p in pkgs:
        p["question_count"] = len(p.get("question_ids", []))
        p["owner_name"] = owners.get(p.get("created_by"), "—")
        p["is_owner"] = user["role"] == "admin" or not p.get("created_by") or p.get("created_by") == user["id"]
    return pkgs


@api_router.get("/packages/{pid}")
async def get_package(pid: str, user: dict = Depends(require_roles("admin", "guru"))):
    p = await db.packages.find_one({"id": pid}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Paket tidak ditemukan")
    return p


def _check_pkg_thresholds(body: "PackageBody"):
    if body.easy_min is not None or body.medium_min is not None:
        if body.easy_min is None or body.medium_min is None:
            raise HTTPException(status_code=400, detail="Ambang Mudah & Sedang harus diisi keduanya")
        if not (0 <= body.medium_min < body.easy_min <= 100):
            raise HTTPException(status_code=400, detail="Harus 0 ≤ Sedang < Mudah ≤ 100")


@api_router.post("/packages")
async def create_package(body: PackageBody, user: dict = Depends(require_roles("admin", "guru"))):
    _check_pkg_thresholds(body)
    data = body.model_dump()
    data["created_by"] = user["id"]
    pkg = Package(**data)
    await db.packages.insert_one(pkg.model_dump())
    return pkg.model_dump()


@api_router.put("/packages/{pid}")
async def update_package(pid: str, body: PackageBody, user: dict = Depends(require_roles("admin", "guru"))):
    _check_pkg_thresholds(body)
    existing = await db.packages.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Paket tidak ditemukan")
    if user["role"] != "admin" and existing.get("created_by") and existing["created_by"] != user["id"]:
        raise HTTPException(status_code=403, detail="Hanya pemilik yang dapat mengubah paket ini")
    await db.packages.update_one({"id": pid}, {"$set": body.model_dump()})
    return await db.packages.find_one({"id": pid}, {"_id": 0})


@api_router.delete("/packages/{pid}")
async def delete_package(pid: str, user: dict = Depends(require_roles("admin", "guru"))):
    existing = await db.packages.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Paket tidak ditemukan")
    if user["role"] != "admin" and existing.get("created_by") and existing["created_by"] != user["id"]:
        raise HTTPException(status_code=403, detail="Hanya pemilik yang dapat menghapus paket ini")
    await db.packages.delete_one({"id": pid})
    return {"ok": True}


@api_router.post("/packages/{pid}/duplicate")
async def duplicate_package(pid: str, user: dict = Depends(require_roles("admin", "guru"))):
    src = await db.packages.find_one({"id": pid}, {"_id": 0})
    if not src:
        raise HTTPException(status_code=404, detail="Paket tidak ditemukan")
    if (user["role"] != "admin" and src.get("created_by")
            and src["created_by"] != user["id"] and not src.get("is_public")):
        raise HTTPException(status_code=403, detail="Paket ini tidak dapat disalin")
    new = Package(
        title=f"{src.get('title', 'Paket')} (Salinan)", description=src.get("description", ""),
        category_id=src.get("category_id"), question_ids=src.get("question_ids", []),
        scoring_method=src.get("scoring_method", "percentage"),
        shuffle_questions=src.get("shuffle_questions", False),
        shuffle_options=src.get("shuffle_options", False),
        min_score=src.get("min_score", 0), rounding=src.get("rounding", "2desimal"),
        easy_min=src.get("easy_min"), medium_min=src.get("medium_min"),
        created_by=user["id"], is_public=False,
    )
    await db.packages.insert_one(new.model_dump())
    return new.model_dump()


# ------------------------------------------------------------------ SESSIONS
async def enrich_session(s: dict) -> dict:
    pkg = await db.packages.find_one({"id": s["package_id"]}, {"_id": 0})
    s["package_title"] = pkg["title"] if pkg else "-"
    s["question_count"] = len(pkg.get("question_ids", [])) if pkg else 0
    now = datetime.now(timezone.utc)
    start = datetime.fromisoformat(s["start_time"])
    end = datetime.fromisoformat(s["end_time"])
    if now < start:
        s["status"] = "akan_datang"
    elif now > end:
        s["status"] = "selesai"
    else:
        s["status"] = "berlangsung"
    return s


@api_router.get("/sessions")
async def list_sessions(user: dict = Depends(get_current_user)):
    sessions = await db.sessions.find({}, {"_id": 0}).sort("start_time", -1).to_list(1000)
    for s in sessions:
        await enrich_session(s)
    if user["role"] == "siswa":
        my_classes = await db.classes.find({"student_ids": user["id"]}, {"_id": 0, "id": 1}).to_list(1000)
        my_class_ids = {c["id"] for c in my_classes}
        my_makeups = await db.makeups.find({"student_id": user["id"]}, {"_id": 0}).to_list(500)
        mmap = {m["session_id"]: m for m in my_makeups}
        visible = []
        for s in sessions:
            targets = s.get("class_ids") or []
            mk = mmap.get(s["id"])
            # Jadwal susulan juga memberi hak akses meski siswa di luar kelas target.
            if targets and not (set(targets) & my_class_ids) and not mk:
                continue
            att = await db.attempts.find_one({"session_id": s["id"], "student_id": user["id"]}, {"_id": 0})
            s["attempt_status"] = att["status"] if att else None
            if mk:
                mk_status = window_status(mk["start_time"], mk["end_time"])
                done_here = bool(att and att.get("makeup_id") == mk["id"] and att["status"] != "berlangsung")
                s["makeup"] = {
                    "id": mk["id"], "start_time": mk["start_time"], "end_time": mk["end_time"],
                    "duration_minutes": mk.get("duration_minutes") or s["duration_minutes"],
                    "reason": mk.get("reason", ""),
                    "status": "sudah_dikerjakan" if done_here else mk_status,
                }
                # Sesi tetap bisa dibuka bila salah satu jendela (reguler / susulan) aktif.
                if s["status"] != "berlangsung" and mk_status == "berlangsung":
                    s["status"] = "berlangsung"
                    s["active_window"] = "susulan"
                elif s["status"] == "berlangsung":
                    s["active_window"] = "sesi"
                elif s["status"] == "selesai" and mk_status == "akan_datang":
                    s["status"] = "akan_datang"
                if s["status"] == "berlangsung" and s.get("active_window") == "susulan":
                    s["effective_end_time"] = mk["end_time"]
                    s["effective_duration"] = mk.get("duration_minutes") or s["duration_minutes"]
            visible.append(s)
        return visible
    else:
        for s in sessions:
            classes = await db.classes.find({"id": {"$in": s.get("class_ids", [])}}, {"_id": 0, "name": 1}).to_list(100)
            s["class_names"] = [c["name"] for c in classes]
            s["makeup_count"] = await db.makeups.count_documents({"session_id": s["id"]})
    return sessions


@api_router.post("/sessions")
async def create_session(body: SessionBody, user: dict = Depends(require_roles("admin", "guru"))):
    ses = Session(**body.model_dump())
    await db.sessions.insert_one(ses.model_dump())
    return ses.model_dump()


@api_router.put("/sessions/{sid}")
async def update_session(sid: str, body: SessionBody, user: dict = Depends(require_roles("admin", "guru"))):
    await db.sessions.update_one({"id": sid}, {"$set": body.model_dump()})
    return await db.sessions.find_one({"id": sid}, {"_id": 0})


@api_router.delete("/sessions/{sid}")
async def delete_session(sid: str, user: dict = Depends(require_roles("admin", "guru"))):
    await db.sessions.delete_one({"id": sid})
    # Cascade: tanpa ini, attempt & susulan yatim terus terhitung pada dashboard/stats.
    res = await db.attempts.delete_many({"session_id": sid})
    await db.makeups.delete_many({"session_id": sid})
    return {"ok": True, "attempts_deleted": res.deleted_count}


# ------------------------------------------------------------------ MAKEUP EXAMS (ujian susulan)
def window_status(start_iso: str, end_iso: str, now: Optional[datetime] = None) -> str:
    """akan_datang | berlangsung | selesai untuk sebuah rentang waktu."""
    now = now or datetime.now(timezone.utc)
    try:
        start = datetime.fromisoformat(start_iso)
        end = datetime.fromisoformat(end_iso)
    except Exception:
        return "selesai"
    if now < start:
        return "akan_datang"
    if now > end:
        return "selesai"
    return "berlangsung"


async def eligible_student_ids(session: dict) -> List[str]:
    """Siswa yang menjadi peserta sesi: anggota kelas target, atau semua siswa bila kosong."""
    targets = session.get("class_ids") or []
    if not targets:
        users = await db.users.find({"role": "siswa"}, {"_id": 1}).to_list(5000)
        return [str(u["_id"]) for u in users]
    classes = await db.classes.find({"id": {"$in": targets}}, {"_id": 0, "student_ids": 1}).to_list(500)
    ids, seen = [], set()
    for c in classes:
        for sid in c.get("student_ids") or []:
            if sid not in seen:
                seen.add(sid)
                ids.append(sid)
    return ids


async def find_students_by_ids(ids: List[str]) -> List[dict]:
    """Ambil dokumen siswa dari daftar id string (ObjectId), dikembalikan via clean_user."""
    oids = []
    for i in ids:
        try:
            oids.append(ObjectId(i))
        except Exception:
            continue
    if not oids:
        return []
    docs = await db.users.find({"_id": {"$in": oids}, "role": "siswa"}).to_list(5000)
    return [clean_user(d) for d in docs]


async def get_makeup(session_id: str, student_id: str) -> Optional[dict]:
    return await db.makeups.find_one({"session_id": session_id, "student_id": student_id}, {"_id": 0})


async def enrich_makeup(mk: dict, session: Optional[dict] = None) -> dict:
    """Tambahkan judul sesi, durasi efektif, dan status pengerjaan."""
    if session is None:
        session = await db.sessions.find_one({"id": mk["session_id"]}, {"_id": 0})
    mk["session_title"] = session["title"] if session else "-"
    mk["session_start_time"] = session["start_time"] if session else None
    mk["session_end_time"] = session["end_time"] if session else None
    mk["effective_duration"] = mk.get("duration_minutes") or (session.get("duration_minutes", 60) if session else 60)
    att = await db.attempts.find_one(
        {"session_id": mk["session_id"], "student_id": mk["student_id"]},
        {"_id": 0, "status": 1, "score": 1, "makeup_id": 1, "submitted_at": 1})
    mk["attempt_status"] = att["status"] if att else None
    mk["score"] = att.get("score") if att else None
    done_here = bool(att and att.get("makeup_id") == mk["id"] and att["status"] != "berlangsung")
    if done_here:
        mk["status"] = "sudah_dikerjakan"
    else:
        mk["status"] = window_status(mk["start_time"], mk["end_time"])
    return mk


@api_router.get("/makeups")
async def list_makeups(session_id: Optional[str] = None,
                       user: dict = Depends(require_roles("admin", "guru"))):
    q = {"session_id": session_id} if session_id else {}
    items = await db.makeups.find(q, {"_id": 0}).sort("start_time", -1).to_list(2000)
    cache: dict = {}
    for mk in items:
        sid = mk["session_id"]
        if sid not in cache:
            cache[sid] = await db.sessions.find_one({"id": sid}, {"_id": 0})
        await enrich_makeup(mk, cache[sid])
    return items


@api_router.get("/makeups/summary")
async def makeups_summary(user: dict = Depends(require_roles("admin", "guru"))):
    """Jumlah susulan terjadwal per sesi — untuk badge di daftar sesi."""
    items = await db.makeups.find({}, {"_id": 0, "session_id": 1}).to_list(5000)
    counts: dict = {}
    for mk in items:
        counts[mk["session_id"]] = counts.get(mk["session_id"], 0) + 1
    return counts


@api_router.get("/makeups/me")
async def my_makeups(user: dict = Depends(require_roles("siswa"))):
    items = await db.makeups.find({"student_id": user["id"]}, {"_id": 0}).sort("start_time", -1).to_list(500)
    for mk in items:
        await enrich_makeup(mk)
    return items


@api_router.get("/makeups/absentees/{session_id}")
async def session_absentees(session_id: str, user: dict = Depends(require_roles("admin", "guru"))):
    """Siswa peserta sesi yang belum menyelesaikan ujian — kandidat ujian susulan."""
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Sesi tidak ditemukan")
    await enrich_session(session)
    ids = await eligible_student_ids(session)
    if not ids:
        return {"session": session, "absentees": []}
    students = await find_students_by_ids(ids)
    students.sort(key=lambda s: (s.get("name") or "").lower())
    attempts = await db.attempts.find(
        {"session_id": session_id, "student_id": {"$in": ids}}, {"_id": 0}).to_list(5000)
    amap = {a["student_id"]: a for a in attempts}
    makeups = await db.makeups.find({"session_id": session_id}, {"_id": 0}).to_list(2000)
    mmap = {m["student_id"]: m for m in makeups}
    out = []
    for s in students:
        att = amap.get(s["id"])
        if att and att["status"] != "berlangsung" and att.get("makeup_id") is None and att.get("score") is not None:
            # sudah mengerjakan pada jadwal reguler dan sudah bernilai
            continue
        if att and att["status"] != "berlangsung" and att.get("makeup_id"):
            continue  # sudah menyelesaikan susulan
        mk = mmap.get(s["id"])
        if att is None:
            reason_hint = "Tidak hadir / belum memulai ujian"
        elif att["status"] == "berlangsung":
            reason_hint = "Sudah memulai tetapi belum dikumpulkan"
        else:
            reason_hint = "Dikumpulkan tanpa nilai"
        out.append({
            "id": s["id"], "name": s["name"], "identifier": s.get("identifier", ""),
            "email": s.get("email", ""),
            "has_attempt": att is not None,
            "attempt_status": att["status"] if att else None,
            "reason_hint": reason_hint,
            "makeup": await enrich_makeup(dict(mk), session) if mk else None,
        })
    return {"session": session, "absentees": out}


@api_router.post("/makeups")
async def create_makeups(body: MakeupBody, user: dict = Depends(require_roles("admin", "guru"))):
    session = await db.sessions.find_one({"id": body.session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Sesi tidak ditemukan")
    if not body.student_ids:
        raise HTTPException(status_code=400, detail="Pilih minimal satu siswa")
    try:
        start = datetime.fromisoformat(body.start_time)
        end = datetime.fromisoformat(body.end_time)
    except Exception:
        raise HTTPException(status_code=400, detail="Format waktu tidak valid")
    if end <= start:
        raise HTTPException(status_code=400, detail="Waktu selesai harus setelah waktu mulai")
    if body.duration_minutes is not None and body.duration_minutes < 1:
        raise HTTPException(status_code=400, detail="Durasi minimal 1 menit")

    students = await find_students_by_ids(body.student_ids)
    smap = {s["id"]: s for s in students}
    created, updated, skipped = 0, 0, []
    for sid in body.student_ids:
        st = smap.get(sid)
        if not st:
            skipped.append({"student_id": sid, "reason": "Akun siswa tidak ditemukan"})
            continue
        doc = {
            "session_id": body.session_id,
            "student_id": sid,
            "student_name": st["name"],
            "student_identifier": st.get("identifier", ""),
            "start_time": body.start_time,
            "end_time": body.end_time,
            "duration_minutes": body.duration_minutes,
            "reason": (body.reason or "").strip(),
            "created_by": user["id"],
            "created_by_name": user["name"],
            "created_at": now_iso(),
        }
        existing = await get_makeup(body.session_id, sid)
        if existing:
            await db.makeups.update_one({"id": existing["id"]}, {"$set": {
                k: v for k, v in doc.items() if k not in ("created_at", "created_by", "created_by_name")}})
            updated += 1
        else:
            doc["id"] = new_id()
            await db.makeups.insert_one(dict(doc))
            created += 1
    return {"created": created, "updated": updated, "skipped": skipped}


@api_router.put("/makeups/{mid}")
async def update_makeup(mid: str, body: MakeupUpdateBody,
                        user: dict = Depends(require_roles("admin", "guru"))):
    mk = await db.makeups.find_one({"id": mid}, {"_id": 0})
    if not mk:
        raise HTTPException(status_code=404, detail="Jadwal susulan tidak ditemukan")
    try:
        start = datetime.fromisoformat(body.start_time)
        end = datetime.fromisoformat(body.end_time)
    except Exception:
        raise HTTPException(status_code=400, detail="Format waktu tidak valid")
    if end <= start:
        raise HTTPException(status_code=400, detail="Waktu selesai harus setelah waktu mulai")
    if body.duration_minutes is not None and body.duration_minutes < 1:
        raise HTTPException(status_code=400, detail="Durasi minimal 1 menit")
    await db.makeups.update_one({"id": mid}, {"$set": {
        "start_time": body.start_time, "end_time": body.end_time,
        "duration_minutes": body.duration_minutes, "reason": (body.reason or "").strip()}})
    fresh = await db.makeups.find_one({"id": mid}, {"_id": 0})
    return await enrich_makeup(fresh)


@api_router.delete("/makeups/{mid}")
async def delete_makeup(mid: str, user: dict = Depends(require_roles("admin", "guru"))):
    mk = await db.makeups.find_one({"id": mid}, {"_id": 0})
    if not mk:
        raise HTTPException(status_code=404, detail="Jadwal susulan tidak ditemukan")
    await db.makeups.delete_one({"id": mid})
    return {"ok": True}


# ------------------------------------------------------------------ EXAM (student)
def sanitize_question(q: dict, perm: Optional[List[int]] = None) -> dict:
    opts = q.get("options", [])
    if perm and q["type"] == "pg":
        opts = [opts[i] for i in perm if i < len(opts)]
    return {
        "id": q["id"], "type": q["type"], "text": q["text"],
        "options": opts, "weight": q.get("weight", 1.0), "image_path": q.get("image_path"),
    }


@api_router.post("/exam/start")
async def start_exam(body: StartAttemptBody, user: dict = Depends(require_roles("siswa"))):
    session = await db.sessions.find_one({"id": body.session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Sesi tidak ditemukan")
    await enrich_session(session)

    # Jendela aktif: jadwal reguler sesi, atau jadwal ujian susulan milik siswa ini.
    mk = await get_makeup(body.session_id, user["id"])
    mk_status = window_status(mk["start_time"], mk["end_time"]) if mk else None
    if session["status"] == "berlangsung":
        window = "sesi"
        eff_end = session["end_time"]
        eff_duration = session.get("duration_minutes", 60)
    elif mk and mk_status == "berlangsung":
        window = "susulan"
        eff_end = mk["end_time"]
        eff_duration = mk.get("duration_minutes") or session.get("duration_minutes", 60)
    else:
        if mk and mk_status == "akan_datang":
            raise HTTPException(status_code=400,
                                detail="Ujian susulan Anda belum dimulai. Silakan kembali sesuai jadwal susulan.")
        if mk and mk_status == "selesai":
            raise HTTPException(status_code=400, detail="Jadwal ujian susulan Anda sudah berakhir")
        if session["status"] == "akan_datang":
            raise HTTPException(status_code=400, detail="Sesi belum dimulai")
        raise HTTPException(status_code=400, detail="Sesi sudah berakhir")

    attempt = await db.attempts.find_one({"session_id": body.session_id, "student_id": user["id"]}, {"_id": 0})
    retake = False
    if attempt and attempt["status"] != "berlangsung":
        if window == "susulan" and attempt.get("makeup_id") != mk["id"]:
            retake = True  # pengerjaan lama diarsipkan, siswa mengerjakan ulang lewat susulan
        else:
            raise HTTPException(status_code=400, detail="Anda sudah mengerjakan sesi ini")

    pkg = await db.packages.find_one({"id": session["package_id"]}, {"_id": 0})
    if not pkg:
        raise HTTPException(status_code=400, detail="Paket soal untuk sesi ini sudah dihapus")
    questions = await db.questions.find({"id": {"$in": pkg.get("question_ids", [])}}, {"_id": 0}).to_list(2000)
    qmap = {q["id"]: q for q in questions}

    def build_order():
        order = list(pkg.get("question_ids", []))
        if pkg.get("shuffle_questions"):
            random.shuffle(order)
        perm = {}
        if pkg.get("shuffle_options"):
            for q in questions:
                if q["type"] == "pg" and q.get("options"):
                    idxs = list(range(len(q["options"])))
                    random.shuffle(idxs)
                    perm[q["id"]] = idxs
        return order, perm

    if retake:
        order_ids, option_perm = build_order()
        history = list(attempt.get("previous_attempts") or [])
        history.append({
            "score": attempt.get("score"), "status": attempt["status"],
            "submitted_at": attempt.get("submitted_at"), "started_at": attempt.get("started_at"),
            "answers": attempt.get("answers", {}), "violations": attempt.get("violations", []),
        })
        reset = {
            "answers": {}, "details": [], "status": "berlangsung", "score": None,
            "earned": None, "total_possible": None, "needs_grading": False,
            "started_at": now_iso(), "submitted_at": None, "violations": [],
            "auto_submitted_reason": None, "question_order": order_ids, "option_perm": option_perm,
            "is_makeup": True, "makeup_id": mk["id"], "effective_end": eff_end,
            "effective_duration": eff_duration, "previous_attempts": history,
        }
        await db.attempts.update_one({"id": attempt["id"]}, {"$set": reset})
        attempt = {**attempt, **reset}
    elif not attempt:
        order_ids, option_perm = build_order()
        attempt = {
            "id": new_id(), "session_id": body.session_id, "student_id": user["id"],
            "student_name": user["name"], "student_identifier": user.get("identifier", ""),
            "package_id": session["package_id"], "answers": {}, "status": "berlangsung",
            "score": None, "started_at": now_iso(), "submitted_at": None,
            "needs_grading": False, "question_order": order_ids, "option_perm": option_perm,
            "violations": [],
            "is_makeup": window == "susulan",
            "makeup_id": mk["id"] if window == "susulan" else None,
            "effective_end": eff_end, "effective_duration": eff_duration,
        }
        await db.attempts.insert_one(dict(attempt))
    else:
        # melanjutkan pengerjaan yang sedang berlangsung — segarkan batas waktu efektif
        await db.attempts.update_one({"id": attempt["id"]}, {"$set": {
            "effective_end": eff_end, "effective_duration": eff_duration,
            "is_makeup": window == "susulan" or bool(attempt.get("is_makeup")),
            "makeup_id": mk["id"] if window == "susulan" else attempt.get("makeup_id"),
        }})

    order_ids = attempt.get("question_order") or list(pkg.get("question_ids", []))
    option_perm = attempt.get("option_perm", {})
    display = [sanitize_question(qmap[qid], option_perm.get(qid)) for qid in order_ids if qid in qmap]

    lock = await get_exam_lock()
    return {
        "attempt_id": attempt["id"],
        "session": {"id": session["id"], "title": session["title"],
                    "duration_minutes": eff_duration, "end_time": eff_end},
        "started_at": attempt["started_at"],
        "answers": attempt.get("answers", {}),
        "questions": display,
        "lock": lock,
        "violations": len(attempt.get("violations", [])),
        "is_makeup": window == "susulan",
    }


def attempt_question_ids(attempt: dict, pkg: Optional[dict]) -> List[str]:
    """Question ids for an attempt, falling back to the stored details/order when the
    package was deleted after the exam was taken (keeps results/PDF working)."""
    ids = list((pkg or {}).get("question_ids") or [])
    if not ids:
        ids = [d.get("question_id") for d in (attempt.get("details") or [])]
    if not ids:
        ids = list(attempt.get("question_order") or [])
    return [i for i in ids if i]


def compute_grade(pkg: dict, questions: dict, answers: dict, essay_scores: dict = None):
    """Returns (details, needs_grading, score_or_none)."""
    essay_scores = essay_scores or {}
    weighted = pkg.get("scoring_method") == "weighted"
    details = []
    total_possible = 0.0
    earned = 0.0
    needs_grading = False
    for qid in pkg.get("question_ids", []):
        q = questions.get(qid)
        if not q:
            continue
        w = q.get("weight", 1.0) if weighted else 1.0
        total_possible += w
        ans = answers.get(qid)
        d = {"question_id": qid, "answer": ans, "type": q["type"],
             "points_possible": w, "points_earned": 0.0, "is_correct": None}
        if q["type"] in ("pg", "truefalse"):
            correct = (ans is not None) and (str(ans) == str(q.get("correct_answer")))
            d["is_correct"] = correct
            d["points_earned"] = w if correct else 0.0
            d["correct_answer"] = q.get("correct_answer")
        else:  # essay
            if qid in essay_scores and essay_scores[qid] is not None:
                pts = max(0.0, min(float(essay_scores[qid]), w))
                d["points_earned"] = pts
                d["is_correct"] = None
            else:
                d["needs_grading"] = True
                needs_grading = True
        earned += d["points_earned"]
        details.append(d)
    if needs_grading:
        score = None
    elif not total_possible:
        score = 0.0
    else:
        raw = earned / total_possible * 100
        rounding = pkg.get("rounding", "2desimal")
        if rounding == "bulat":
            raw = round(raw)
        elif rounding == "1desimal":
            raw = round(raw, 1)
        else:
            raw = round(raw, 2)
        min_score = float(pkg.get("min_score", 0) or 0)
        score = max(raw, min_score)
        if rounding == "bulat":
            score = round(score)
        elif rounding == "1desimal":
            score = round(score, 1)
        else:
            score = round(score, 2)
    return details, needs_grading, score, round(earned, 2), round(total_possible, 2)


async def finalize_attempt(attempt: dict, answers: dict) -> dict:
    """Convert shuffled indices, grade, and persist. Shared by submit + auto-submit."""
    pkg = await db.packages.find_one({"id": attempt["package_id"]}, {"_id": 0}) or {}
    if not pkg.get("question_ids"):
        pkg = {**pkg, "question_ids": attempt_question_ids(attempt, pkg)}
    qlist = await db.questions.find({"id": {"$in": pkg["question_ids"]}}, {"_id": 0}).to_list(2000)
    qmap = {q["id"]: q for q in qlist}
    perm_map = attempt.get("option_perm", {})
    canonical = {}
    for qid, ans in (answers or {}).items():
        q = qmap.get(qid)
        if q and q.get("type") == "pg" and qid in perm_map and ans not in (None, ""):
            try:
                canonical[qid] = str(perm_map[qid][int(ans)])
            except (ValueError, IndexError, TypeError):
                canonical[qid] = ans
        else:
            canonical[qid] = ans
    details, needs_grading, score, earned, total = compute_grade(pkg, qmap, canonical)
    update = {
        "answers": canonical, "details": details, "needs_grading": needs_grading,
        "score": score, "earned": earned, "total_possible": total,
        "status": "menunggu_koreksi" if needs_grading else "selesai",
        "submitted_at": now_iso(),
    }
    await db.attempts.update_one({"id": attempt["id"]}, {"$set": update})
    return update


@api_router.post("/exam/submit")
async def submit_exam(body: SubmitBody, user: dict = Depends(require_roles("siswa"))):
    attempt = await db.attempts.find_one({"session_id": body.session_id, "student_id": user["id"]}, {"_id": 0})
    if not attempt:
        raise HTTPException(status_code=404, detail="Percobaan tidak ditemukan")
    if attempt["status"] != "berlangsung":
        raise HTTPException(status_code=400, detail="Sesi sudah dikumpulkan")
    update = await finalize_attempt(attempt, body.answers)
    return {"status": update["status"], "score": update["score"], "needs_grading": update["needs_grading"]}


@api_router.post("/exam/save/{session_id}")
async def save_progress(session_id: str, body: dict, user: dict = Depends(require_roles("siswa"))):
    await db.attempts.update_one(
        {"session_id": session_id, "student_id": user["id"], "status": "berlangsung"},
        {"$set": {"answers": body.get("answers", {})}})
    return {"ok": True}


# ------------------------------------------------------------------ RESULTS
@api_router.get("/results/session/{session_id}")
async def results_by_session(session_id: str, user: dict = Depends(require_roles("admin", "guru"))):
    attempts = await db.attempts.find({"session_id": session_id}, {"_id": 0}).to_list(2000)
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    return {"session": session, "attempts": attempts}


@api_router.get("/results/me")
async def my_results(user: dict = Depends(require_roles("siswa"))):
    attempts = await db.attempts.find(
        {"student_id": user["id"], "status": {"$ne": "berlangsung"}}, {"_id": 0}
    ).sort("submitted_at", -1).to_list(1000)
    for a in attempts:
        s = await db.sessions.find_one({"id": a["session_id"]}, {"_id": 0})
        a["session_title"] = s["title"] if s else "-"
        a["kkm"] = s.get("kkm", 75) if s else 75
    return attempts


@api_router.get("/results/detail/{attempt_id}")
async def result_detail(attempt_id: str, user: dict = Depends(get_current_user)):
    attempt = await db.attempts.find_one({"id": attempt_id}, {"_id": 0})
    if not attempt:
        raise HTTPException(status_code=404, detail="Tidak ditemukan")
    if user["role"] == "siswa" and attempt["student_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Akses ditolak")
    pkg = await db.packages.find_one({"id": attempt["package_id"]}, {"_id": 0}) or {}
    qlist = await db.questions.find({"id": {"$in": attempt_question_ids(attempt, pkg)}},
                                    {"_id": 0}).to_list(2000)
    qmap = {q["id"]: q for q in qlist}
    session = await db.sessions.find_one({"id": attempt["session_id"]}, {"_id": 0})
    enriched = []
    for d in attempt.get("details", []):
        q = qmap.get(d["question_id"], {})
        enriched.append({**d, "text": q.get("text"), "options": q.get("options", []),
                         "correct_answer": q.get("correct_answer"), "image_path": q.get("image_path")})
    attempt["details"] = enriched
    attempt["session_title"] = session["title"] if session else "-"
    attempt["scoring_method"] = pkg.get("scoring_method")
    return attempt


@api_router.post("/results/grade/{attempt_id}")
async def grade_essay(attempt_id: str, body: GradeEssayBody, user: dict = Depends(require_roles("admin", "guru"))):
    attempt = await db.attempts.find_one({"id": attempt_id}, {"_id": 0})
    if not attempt:
        raise HTTPException(status_code=404, detail="Tidak ditemukan")
    pkg = await db.packages.find_one({"id": attempt["package_id"]}, {"_id": 0}) or {}
    if not pkg.get("question_ids"):
        pkg = {**pkg, "question_ids": attempt_question_ids(attempt, pkg)}
    qlist = await db.questions.find({"id": {"$in": pkg["question_ids"]}}, {"_id": 0}).to_list(2000)
    qmap = {q["id"]: q for q in qlist}
    details, needs_grading, score, earned, total = compute_grade(
        pkg, qmap, attempt.get("answers", {}), body.scores)
    await db.attempts.update_one({"id": attempt_id}, {"$set": {
        "details": details, "needs_grading": needs_grading, "score": score,
        "earned": earned, "total_possible": total,
        "status": "menunggu_koreksi" if needs_grading else "selesai",
    }})
    return {"score": score, "needs_grading": needs_grading}


# ------------------------------------------------------------------ CLASSES
async def enrich_class(c: dict) -> dict:
    c["student_count"] = len(c.get("student_ids", []))
    return c


@api_router.get("/classes")
async def list_classes(user: dict = Depends(require_roles("admin", "guru"))):
    items = await db.classes.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for c in items:
        await enrich_class(c)
    return items


@api_router.post("/classes")
async def create_class(body: ClassBody, user: dict = Depends(require_roles("admin", "guru"))):
    cls = SchoolClass(**body.model_dump())
    await db.classes.insert_one(cls.model_dump())
    return await enrich_class(cls.model_dump())


@api_router.put("/classes/{cid}")
async def update_class(cid: str, body: ClassBody, user: dict = Depends(require_roles("admin", "guru"))):
    await db.classes.update_one({"id": cid}, {"$set": body.model_dump()})
    c = await db.classes.find_one({"id": cid}, {"_id": 0})
    return await enrich_class(c)


@api_router.delete("/classes/{cid}")
async def delete_class(cid: str, user: dict = Depends(require_roles("admin", "guru"))):
    await db.classes.delete_one({"id": cid})
    return {"ok": True}


# ------------------------------------------------------- CLASS ROSTER (student accounts)
class ClassStudentCreate(BaseModel):
    name: str
    email: EmailStr
    password: str
    identifier: Optional[str] = None


class AttachStudentsBody(BaseModel):
    student_ids: List[str] = []


async def _get_class_or_404(cid: str) -> dict:
    cls = await db.classes.find_one({"id": cid}, {"_id": 0})
    if not cls:
        raise HTTPException(status_code=404, detail="Kelas tidak ditemukan")
    return cls


async def _class_roster(cls: dict) -> List[dict]:
    ids = cls.get("student_ids", []) or []
    oids = []
    for sid in ids:
        try:
            oids.append(ObjectId(sid))
        except Exception:
            continue
    if not oids:
        return []
    users = await db.users.find({"_id": {"$in": oids}, "role": "siswa"}).to_list(2000)
    roster = []
    for u in users:
        attempts = await db.attempts.count_documents({"student_id": str(u["_id"]), "status": "selesai"})
        roster.append({
            "id": str(u["_id"]),
            "name": u.get("name", ""),
            "email": u.get("email", ""),
            "identifier": u.get("identifier", "") or "",
            "created_at": u.get("created_at"),
            "exams_done": attempts,
        })
    roster.sort(key=lambda x: x["name"].lower())
    return roster


@api_router.get("/classes/{cid}/students")
async def list_class_students(cid: str, user: dict = Depends(require_roles("admin", "guru"))):
    """Student accounts that belong to this class + students not in any class yet."""
    cls = await _get_class_or_404(cid)
    roster = await _class_roster(cls)
    member_ids = {s["id"] for s in roster}

    all_classes = await db.classes.find({}, {"_id": 0, "id": 1, "name": 1, "student_ids": 1}).to_list(1000)
    class_of = {}
    for c in all_classes:
        for sid in c.get("student_ids", []) or []:
            class_of.setdefault(sid, []).append(c["name"])

    others = []
    async for u in db.users.find({"role": "siswa"}):
        sid = str(u["_id"])
        if sid in member_ids:
            continue
        others.append({"id": sid, "name": u.get("name", ""), "email": u.get("email", ""),
                       "identifier": u.get("identifier", "") or "",
                       "class_names": class_of.get(sid, [])})
    others.sort(key=lambda x: x["name"].lower())
    return {"class": {"id": cls["id"], "name": cls["name"], "description": cls.get("description", "")},
            "students": roster, "available": others}


@api_router.post("/classes/{cid}/students")
async def add_class_student(cid: str, body: ClassStudentCreate,
                            user: dict = Depends(require_roles("admin"))):
    """Create a login-ready student account and put it straight into this class."""
    cls = await _get_class_or_404(cid)
    email = body.email.lower().strip()
    if not body.name.strip():
        raise HTTPException(status_code=400, detail="Nama siswa wajib diisi")
    if len(body.password or "") < 5:
        raise HTTPException(status_code=400, detail="Password minimal 5 karakter")

    existing = await db.users.find_one({"email": email})
    if existing is not None:
        if existing.get("role") != "siswa":
            raise HTTPException(status_code=400,
                                detail=f"Email sudah dipakai akun {existing.get('role')}")
        raise HTTPException(status_code=400,
                            detail="Email sudah terdaftar sebagai siswa. Gunakan 'Tambah dari akun yang ada'.")

    doc = {"email": email, "password_hash": hash_password(body.password),
           "name": body.name.strip(), "role": "siswa",
           "identifier": (body.identifier or "").strip(), "created_at": now_iso()}
    res = await db.users.insert_one(doc)
    sid = str(res.inserted_id)
    members = list(dict.fromkeys((cls.get("student_ids") or []) + [sid]))
    await db.classes.update_one({"id": cid}, {"$set": {"student_ids": members}})
    return {"id": sid, "name": doc["name"], "email": email,
            "identifier": doc["identifier"], "exams_done": 0}


@api_router.post("/classes/{cid}/students/attach")
async def attach_class_students(cid: str, body: AttachStudentsBody,
                                user: dict = Depends(require_roles("admin", "guru"))):
    """Move/attach existing student accounts into this class."""
    cls = await _get_class_or_404(cid)
    valid = []
    for sid in body.student_ids:
        try:
            u = await db.users.find_one({"_id": ObjectId(sid), "role": "siswa"})
        except Exception:
            u = None
        if u:
            valid.append(sid)
    members = list(dict.fromkeys((cls.get("student_ids") or []) + valid))
    await db.classes.update_one({"id": cid}, {"$set": {"student_ids": members}})
    return {"added": len(set(valid) - set(cls.get("student_ids") or [])), "total": len(members)}


@api_router.delete("/classes/{cid}/students/{sid}")
async def remove_class_student(cid: str, sid: str, delete_account: bool = False,
                               user: dict = Depends(require_roles("admin"))):
    """Remove a student from the class; optionally delete the login account too."""
    cls = await _get_class_or_404(cid)
    members = [x for x in (cls.get("student_ids") or []) if x != sid]
    await db.classes.update_one({"id": cid}, {"$set": {"student_ids": members}})
    if delete_account:
        try:
            await db.users.delete_one({"_id": ObjectId(sid), "role": "siswa"})
        except Exception:
            pass
        await db.classes.update_many({"student_ids": sid}, {"$pull": {"student_ids": sid}})
    return {"ok": True, "deleted_account": bool(delete_account)}


@api_router.get("/classes/{cid}/students/xlsx")
async def export_class_roster(cid: str, user: dict = Depends(require_roles("admin", "guru"))):
    """Excel list of student accounts of a class (name, NIS, username) to hand out."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    cls = await _get_class_or_404(cid)
    roster = await _class_roster(cls)
    school = await db.settings.find_one({"key": "school"}, {"_id": 0}) or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Akun Siswa"
    ws.sheet_view.showGridLines = False
    widths = [6, 34, 20, 38, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    r = 1
    if school.get("name"):
        ws.merge_cells(f"A{r}:E{r}")
        c = ws.cell(row=r, column=1, value=school["name"].upper())
        c.font = Font(bold=True, size=13, color=XL_GREEN)
        r += 1
    ws.merge_cells(f"A{r}:E{r}")
    c = ws.cell(row=r, column=1, value=f"Daftar Akun Siswa — {cls['name']}")
    c.font = Font(bold=True, size=11)
    r += 2

    headers = ["No", "Nama Siswa", "NIS / NISN", "Username (Email Login)", "Jumlah Ujian"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=XL_GREEN)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _xl_border()
    r += 1
    for idx, s in enumerate(roster, start=1):
        for i, v in enumerate([idx, s["name"], s["identifier"] or "-", s["email"], s["exams_done"]], start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = _xl_border()
            if i in (1, 3, 5):
                cell.alignment = Alignment(horizontal="center")
        r += 1
    if not roster:
        ws.merge_cells(f"A{r}:E{r}")
        ws.cell(row=r, column=1, value="Belum ada siswa di kelas ini.").alignment = Alignment(horizontal="center")
        r += 1
    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    note = ws.cell(row=r, column=1,
                   value="Siswa login memakai Username (Email) dan password yang diberikan admin. "
                         "Password tidak ditampilkan demi keamanan — gunakan tombol 'Reset Password' bila lupa.")
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = re.sub(r"[^A-Za-z0-9._-]+", "-", cls["name"]).strip("-") or "kelas"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="akun-siswa-{safe}.xlsx"'})


# ------------------------------------------------- BULK PASSWORD RESET + LOGIN CARDS
class ResetPasswordsBody(BaseModel):
    mode: str = "random"            # "random" (per student) | "same" (one password for all)
    password: Optional[str] = None  # required when mode == "same"
    student_ids: List[str] = []     # empty -> every student in the class


class CardCredential(BaseModel):
    name: str = ""
    email: str = ""
    identifier: Optional[str] = ""
    password: Optional[str] = ""


class LoginCardsBody(BaseModel):
    login_url: str = ""
    include_password: bool = True
    credentials: List[CardCredential] = []   # empty -> roster with blank password field


def _random_password(n: int = 8) -> str:
    """Readable password: no 0/O/1/l mix-ups, easy to type on a school PC."""
    alphabet = "abcdefghjkmnpqrstuvwxyz"
    digits = "23456789"
    body = "".join(random.choice(alphabet) for _ in range(max(3, n - 3)))
    return body + "".join(random.choice(digits) for _ in range(3))


@api_router.post("/classes/{cid}/students/reset-passwords")
async def reset_class_passwords(cid: str, body: ResetPasswordsBody,
                                user: dict = Depends(require_roles("admin"))):
    """Reset the login password of every (or selected) student of a class in one go."""
    cls = await _get_class_or_404(cid)
    roster = await _class_roster(cls)
    if body.student_ids:
        wanted = set(body.student_ids)
        roster = [s for s in roster if s["id"] in wanted]
    if not roster:
        raise HTTPException(status_code=400, detail="Tidak ada siswa di kelas ini")

    mode = (body.mode or "random").lower()
    if mode not in ("random", "same"):
        raise HTTPException(status_code=400, detail="Mode tidak valid")
    if mode == "same":
        if len(body.password or "") < 5:
            raise HTTPException(status_code=400, detail="Password minimal 5 karakter")

    creds = []
    for s in roster:
        pwd = body.password if mode == "same" else _random_password()
        try:
            await db.users.update_one({"_id": ObjectId(s["id"]), "role": "siswa"},
                                      {"$set": {"password_hash": hash_password(pwd)}})
        except Exception:
            continue
        creds.append({"id": s["id"], "name": s["name"], "email": s["email"],
                      "identifier": s.get("identifier", ""), "password": pwd})

    return {"count": len(creds), "class_name": cls["name"], "credentials": creds}


@api_router.post("/classes/{cid}/students/cards/pdf")
async def class_login_cards(cid: str, body: LoginCardsBody,
                            user: dict = Depends(require_roles("admin", "guru"))):
    """Printable login cards (name / username / password) to hand out to students."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as pdfcanvas

    cls = await _get_class_or_404(cid)
    school = await db.settings.find_one({"key": "school"}, {"_id": 0}) or {}

    items = [c.model_dump() for c in body.credentials]
    if not items:
        items = [{"name": s["name"], "email": s["email"],
                  "identifier": s.get("identifier", ""), "password": ""}
                 for s in await _class_roster(cls)]
    if not items:
        raise HTTPException(status_code=400, detail="Belum ada siswa di kelas ini")

    buf = io.BytesIO()
    c = pdfcanvas.Canvas(buf, pagesize=A4)
    page_w, page_h = A4
    margin = 12 * mm
    cols, rows = 2, 5
    gap = 4 * mm
    card_w = (page_w - 2 * margin - gap) / cols
    card_h = (page_h - 2 * margin - (rows - 1) * gap) / rows
    green = colors.HexColor("#1E3A30")
    grey = colors.HexColor("#6B7280")
    line = colors.HexColor("#D9D9CF")
    login_url = (body.login_url or "").replace("https://", "").replace("http://", "").rstrip("/")
    school_name = (school.get("name") or "CBT Ujian Online").strip()

    def draw_card(idx, item):
        pos = idx % (cols * rows)
        col = pos % cols
        row = pos // cols
        x = margin + col * (card_w + gap)
        y = page_h - margin - (row + 1) * card_h - row * gap

        c.setStrokeColor(line)
        c.setLineWidth(0.8)
        c.roundRect(x, y, card_w, card_h, 3 * mm, stroke=1, fill=0)

        # header band
        c.setFillColor(green)
        c.roundRect(x, y + card_h - 9 * mm, card_w, 9 * mm, 3 * mm, stroke=0, fill=1)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 8.5)
        c.drawString(x + 5 * mm, y + card_h - 6 * mm, school_name.upper()[:38])
        c.setFont("Helvetica", 7.5)
        c.drawRightString(x + card_w - 5 * mm, y + card_h - 6 * mm, cls["name"][:22])

        ty = y + card_h - 15 * mm
        c.setFillColor(grey)
        c.setFont("Helvetica", 6.5)
        c.drawString(x + 5 * mm, ty, "KARTU LOGIN SISWA")
        ty -= 6 * mm
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(x + 5 * mm, ty, (item.get("name") or "-")[:34])
        if item.get("identifier"):
            ty -= 4.5 * mm
            c.setFillColor(grey)
            c.setFont("Helvetica", 7.5)
            c.drawString(x + 5 * mm, ty, f"NIS/NISN: {item['identifier']}")

        ty -= 7 * mm
        for label, value in (("Username", item.get("email") or "-"),
                             ("Password", (item.get("password") or "") if body.include_password else "")):
            c.setFillColor(grey)
            c.setFont("Helvetica", 7)
            c.drawString(x + 5 * mm, ty, label.upper())
            c.setFillColor(colors.black)
            c.setFont("Courier-Bold", 9.5)
            shown = value if value else "________________"
            c.drawString(x + 22 * mm, ty - 0.3 * mm, shown[:26])
            c.setStrokeColor(line)
            c.setLineWidth(0.4)
            c.line(x + 5 * mm, ty - 2.2 * mm, x + card_w - 5 * mm, ty - 2.2 * mm)
            ty -= 7 * mm

        c.setFillColor(grey)
        c.setFont("Helvetica", 6.5)
        if login_url:
            c.drawString(x + 5 * mm, y + 7.5 * mm, f"Buka: {login_url}"[:52])
            c.drawString(x + 5 * mm, y + 4 * mm, "Masuk dengan username & password di atas.")
        else:
            c.drawString(x + 5 * mm, y + 4 * mm, "Masuk ke aplikasi CBT dengan username & password di atas.")

    for i, item in enumerate(items):
        if i > 0 and i % (cols * rows) == 0:
            c.showPage()
        draw_card(i, item)
    c.showPage()
    c.save()
    buf.seek(0)
    safe = re.sub(r"[^A-Za-z0-9._-]+", "-", cls["name"]).strip("-") or "kelas"
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="kartu-login-{safe}.pdf"'})


# ------------------------------------------------------------------ STUDENT IMPORT (Excel)
STUDENT_COLS = ["nama", "kelas", "nis", "username", "password"]
STUDENT_COL_HELP = {
    "nama": "Nama lengkap siswa. Wajib diisi.",
    "kelas": "Nama kelas / rombel, contoh: Kelas X-A. Bila kelas belum ada akan dibuat otomatis. "
             "Boleh dikosongkan bila siswa belum masuk kelas.",
    "nis": "NIS / NISN siswa. Boleh dikosongkan.",
    "username": "Username untuk login. Harus berupa alamat email karena kolom login memakai email. "
                "Contoh: ani.siswa@sekolah.id",
    "password": "Password awal login siswa. Minimal 5 karakter. Wajib diisi untuk siswa baru.",
}
STUDENT_SAMPLE = [
    ["Ani Rahmawati", "Kelas X-A", "0051234561", "ani.rahmawati@sekolah.id", "siswa12345"],
    ["Budi Santoso", "Kelas X-A", "0051234562", "budi.santoso@sekolah.id", "siswa12345"],
    ["Citra Dewi", "Kelas X-B", "0051234563", "citra.dewi@sekolah.id", "siswa12345"],
]
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


@api_router.get("/students/import-template")
async def student_import_template(user: dict = Depends(require_roles("admin", "guru"))):
    """A ready-to-fill, nicely formatted Excel workbook for bulk student import."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    school = await db.settings.find_one({"key": "school"}, {"_id": 0}) or {}
    classes = await db.classes.find({}, {"_id": 0, "name": 1}).to_list(1000)
    class_names = sorted({c["name"] for c in classes})

    thin = _xl_border()
    wb = Workbook()

    # ---------------------------------------------------------- Petunjuk
    guide = wb.active
    guide.title = "Petunjuk"
    guide.sheet_view.showGridLines = False
    guide.column_dimensions["A"].width = 4
    guide.column_dimensions["B"].width = 22
    guide.column_dimensions["C"].width = 92

    r = 1
    if school.get("name"):
        guide.merge_cells(f"B{r}:C{r}")
        c = guide.cell(row=r, column=2, value=school["name"].upper())
        c.font = Font(bold=True, size=13, color=XL_GREEN)
        r += 1
    guide.merge_cells(f"B{r}:C{r}")
    c = guide.cell(row=r, column=2, value="TEMPLATE IMPOR DATA SISWA")
    c.font = Font(bold=True, size=15, color=XL_GREEN)
    guide.row_dimensions[r].height = 22
    r += 1
    guide.merge_cells(f"B{r}:C{r}")
    c = guide.cell(row=r, column=2, value="Isi lembar \"Data Siswa\", simpan sebagai .xlsx, lalu unggah pada menu Manajemen Kelas → Impor Siswa.")
    c.font = Font(size=10, color="7A7A72")
    r += 2

    steps = [
        "Buka lembar \"Data Siswa\" (tab di bawah).",
        "Hapus 3 baris contoh berwarna abu-abu, lalu isi data siswa Anda mulai baris 4.",
        "Satu baris = satu siswa. Jangan mengubah atau menghapus baris judul kolom.",
        "Kolom username WAJIB berupa alamat email — itulah yang diketik siswa saat login.",
        "Bila nama kelas belum terdaftar, kelas baru akan dibuat otomatis saat impor.",
        "Bila username sudah terdaftar, data siswa akan diperbarui (bukan diduplikasi).",
        "Simpan file (.xlsx atau .csv), lalu unggah lewat tombol \"Pilih File & Impor\".",
    ]
    guide.merge_cells(f"B{r}:C{r}")
    c = guide.cell(row=r, column=2, value="LANGKAH PENGISIAN")
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=XL_GREEN)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    r += 1
    for i, step in enumerate(steps, start=1):
        num = guide.cell(row=r, column=2, value=f"Langkah {i}")
        num.font = Font(size=10, bold=True, color=XL_GREEN)
        num.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        num.border = thin
        txt = guide.cell(row=r, column=3, value=step)
        txt.font = Font(size=10)
        txt.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        txt.border = thin
        guide.row_dimensions[r].height = 20
        r += 1
    r += 1

    guide.merge_cells(f"B{r}:C{r}")
    c = guide.cell(row=r, column=2, value="KETERANGAN KOLOM")
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=XL_GREEN)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    r += 1
    for col in STUDENT_COLS:
        head = guide.cell(row=r, column=2, value=col)
        head.font = Font(size=10, bold=True, color=XL_GREEN)
        head.fill = PatternFill("solid", fgColor=XL_GREEN_SOFT)
        head.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        head.border = thin
        desc = guide.cell(row=r, column=3, value=STUDENT_COL_HELP[col])
        desc.font = Font(size=10)
        desc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        desc.border = thin
        guide.row_dimensions[r].height = 30
        r += 1

    if class_names:
        r += 1
        guide.merge_cells(f"B{r}:C{r}")
        c = guide.cell(row=r, column=2, value="Kelas yang sudah terdaftar: " + " · ".join(class_names))
        c.font = Font(size=9, italic=True, color="7A7A72")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ---------------------------------------------------------- Data Siswa
    ws = wb.create_sheet("Data Siswa")
    ws.sheet_view.showGridLines = False
    ncol = len(STUDENT_COLS)
    last = get_column_letter(ncol)

    ws.merge_cells(f"A1:{last}1")
    c = ws.cell(row=1, column=1, value="DATA SISWA — ISI MULAI BARIS 4")
    c.font = Font(bold=True, size=13, color=XL_GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{last}2")
    c = ws.cell(row=2, column=1, value="Baris abu-abu di bawah hanya contoh — silakan hapus sebelum mengunggah. username harus berupa email.")
    c.font = Font(size=9, italic=True, color="9A9A92")
    c.alignment = Alignment(horizontal="center", vertical="center")

    labels = {"nama": "nama", "kelas": "kelas", "nis": "nis",
              "username": "username", "password": "password"}
    for i, col in enumerate(STUDENT_COLS, start=1):
        cell = ws.cell(row=3, column=i, value=labels[col])
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=XL_GREEN)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
        cell.comment = _xl_comment(STUDENT_COL_HELP[col])
    ws.row_dimensions[3].height = 26

    for ri, row in enumerate(STUDENT_SAMPLE, start=4):
        for ci, v in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = Font(size=10, italic=True, color="9A9A92")
            cell.fill = PatternFill("solid", fgColor=XL_STRIPE)
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if ci == 3:
                cell.number_format = "@"  # keep leading zeros of NIS

    # empty, pre-formatted rows so the sheet stays tidy while typing
    for ri in range(4 + len(STUDENT_SAMPLE), 4 + len(STUDENT_SAMPLE) + 40):
        for ci in range(1, ncol + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = thin
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if ci == 3:
                cell.number_format = "@"

    if class_names:
        dv = DataValidation(type="list", formula1='"' + ",".join(class_names)[:250] + '"',
                            allow_blank=True, showDropDown=False)
        dv.prompt = "Pilih kelas yang ada atau tulis nama kelas baru"
        dv.promptTitle = "Kelas"
        dv.error = "Kelas belum terdaftar — tetap boleh diisi, kelas baru akan dibuat otomatis."
        dv.errorStyle = "warning"
        ws.add_data_validation(dv)
        dv.add(f"B4:B{4 + len(STUDENT_SAMPLE) + 39}")

    for col, w in zip("ABCDE", [30, 18, 18, 34, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "3:3"

    wb.active = wb.index(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_data_siswa.xlsx"})


@api_router.post("/students/import")
async def import_students(file: UploadFile = File(...),
                         class_id: Optional[str] = Form(None),
                         user: dict = Depends(require_roles("admin"))):
    """Bulk-create student accounts from Excel/CSV and place them into classes.

    When `class_id` is given (import launched from inside a class), every row that
    leaves the 'kelas' column empty is placed into that class.
    """
    default_class = None
    if class_id:
        default_class = await db.classes.find_one({"id": class_id}, {"_id": 0})
        if not default_class:
            raise HTTPException(status_code=404, detail="Kelas tujuan tidak ditemukan")
    raw = await file.read()
    name = (file.filename or "").lower()
    try:
        if name.endswith((".xlsx", ".xls")):
            df = pd.read_excel(io.BytesIO(raw), dtype=str, sheet_name="Data Siswa", header=2)
        else:
            df = pd.read_csv(io.BytesIO(raw), dtype=str)
    except ValueError:
        # workbook without the expected sheet name -> fall back to the first sheet
        try:
            df = pd.read_excel(io.BytesIO(raw), dtype=str)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Gagal membaca file: {e}")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal membaca file: {e}")

    df.columns = [str(c).strip().lower() for c in df.columns]
    aliases = {"nama siswa": "nama", "name": "nama", "email": "username",
               "user name": "username", "nisn": "nis", "nis/nisn": "nis",
               "nisn/nis": "nis", "kata sandi": "password", "sandi": "password",
               "rombel": "kelas", "kelas/rombel": "kelas"}
    df.rename(columns=aliases, inplace=True)
    if "nama" not in df.columns or "username" not in df.columns:
        raise HTTPException(
            status_code=400,
            detail="Kolom wajib tidak ditemukan. Pastikan ada kolom 'nama' dan 'username'. "
                   "Unduh template untuk format yang benar.")

    def cell(row, key):
        v = row.get(key, "")
        if v is None:
            return ""
        v = str(v).strip()
        return "" if v.lower() in ("nan", "none") else v

    classes = await db.classes.find({}, {"_id": 0}).to_list(1000)
    class_by_name = {c["name"].strip().lower(): c for c in classes}
    class_members = {c["id"]: set(c.get("student_ids", [])) for c in classes}
    created_classes, touched_classes = [], set()
    created = updated = 0
    errors = []
    seen_emails = set()

    for i, row in df.iterrows():
        rownum = i + 4 if name.endswith((".xlsx", ".xls")) else i + 2
        nama = cell(row, "nama")
        username = cell(row, "username").lower()
        password = cell(row, "password")
        nis = cell(row, "nis")
        kelas = cell(row, "kelas")
        if not (nama or username or password or nis or kelas):
            continue  # blank row
        if not kelas and default_class is not None:
            kelas = default_class["name"]
        if not nama:
            errors.append(f"Baris {rownum}: nama wajib diisi")
            continue
        if not username:
            errors.append(f"Baris {rownum}: username (email) wajib diisi")
            continue
        if not EMAIL_RE.match(username):
            errors.append(f"Baris {rownum}: username '{username}' harus berupa email, contoh nama@sekolah.id")
            continue
        if username in seen_emails:
            errors.append(f"Baris {rownum}: username '{username}' dobel di dalam file")
            continue
        seen_emails.add(username)

        existing = await db.users.find_one({"email": username})
        if existing is not None and existing.get("role") != "siswa":
            errors.append(f"Baris {rownum}: '{username}' sudah dipakai akun {existing.get('role')}")
            continue
        if existing is None and len(password) < 5:
            errors.append(f"Baris {rownum}: password minimal 5 karakter")
            continue

        if existing is None:
            doc = {"email": username, "password_hash": hash_password(password),
                   "name": nama, "role": "siswa", "identifier": nis,
                   "created_at": now_iso()}
            res = await db.users.insert_one(doc)
            sid = str(res.inserted_id)
            created += 1
        else:
            sid = str(existing["_id"])
            upd = {"name": nama, "identifier": nis or existing.get("identifier", "")}
            if password:
                if len(password) < 5:
                    errors.append(f"Baris {rownum}: password minimal 5 karakter, password lama dipertahankan")
                else:
                    upd["password_hash"] = hash_password(password)
            await db.users.update_one({"_id": existing["_id"]}, {"$set": upd})
            updated += 1

        if kelas:
            key = kelas.strip().lower()
            cls = class_by_name.get(key)
            if cls is None:
                cls = SchoolClass(name=kelas.strip(), description="Dibuat dari impor siswa").model_dump()
                await db.classes.insert_one(dict(cls))
                class_by_name[key] = cls
                class_members[cls["id"]] = set()
                created_classes.append(cls["name"])
            class_members[cls["id"]].add(sid)
            touched_classes.add(cls["id"])

    added_to_class = 0
    for cid in touched_classes:
        members = sorted(class_members[cid])
        before = len(next((c.get("student_ids", []) for c in classes if c["id"] == cid), []))
        await db.classes.update_one({"id": cid}, {"$set": {"student_ids": members}})
        added_to_class += max(0, len(members) - before)

    return {"created": created, "updated": updated,
            "classes_created": created_classes,
            "added_to_class": added_to_class,
            "errors": errors}



# ------------------------------------------------------------------ QUESTION IMPORT
IMPORT_TEMPLATE = (
    "type,text,option_a,option_b,option_c,option_d,option_e,correct,weight,category,image_url\n"
    "pg,Berapa hasil 5 + 3?,6,7,8,9,10,C,1,Matematika,\n"
    "pg,Ibu kota Provinsi Jawa Barat adalah ...,Bogor,Bandung,Bekasi,Cimahi,Depok,B,1,IPS,\n"
    "truefalse,Matahari terbit dari timur.,,,,,,benar,1,IPA,\n"
    "essay,Jelaskan proses fotosintesis.,,,,,,,2,IPA,\n"
    "pg,Perhatikan gambar berikut.,A,B,C,D,E,A,1,IPA,https://contoh.com/gambar.png\n"
)


@api_router.get("/questions/import-template")
async def import_template(user: dict = Depends(require_roles("admin", "guru"))):
    return StreamingResponse(
        io.BytesIO(IMPORT_TEMPLATE.encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=template_soal.csv"},
    )


@api_router.post("/questions/import")
async def import_questions(file: UploadFile = File(...), user: dict = Depends(require_roles("admin", "guru"))):
    raw = await file.read()
    name = (file.filename or "").lower()
    try:
        if name.endswith(".xlsx") or name.endswith(".xls"):
            df = pd.read_excel(io.BytesIO(raw), dtype=str)
        else:
            df = pd.read_csv(io.BytesIO(raw), dtype=str)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal membaca file: {e}")

    df.columns = [str(c).strip().lower() for c in df.columns]
    cats = await db.categories.find({}, {"_id": 0}).to_list(1000)
    cat_by_name = {c["name"].strip().lower(): c["id"] for c in cats}

    imported = 0
    errors = []
    letter_idx = {"a": "0", "b": "1", "c": "2", "d": "3", "e": "4"}
    for i, row in df.iterrows():
        rownum = i + 2
        try:
            qtype = str(row.get("type", "")).strip().lower()
            text = str(row.get("text", "")).strip()
            if not text or text.lower() == "nan" or qtype not in ("pg", "truefalse", "essay"):
                errors.append(f"Baris {rownum}: tipe/teks tidak valid")
                continue
            # category
            cat_name = str(row.get("category", "")).strip()
            cat_id = None
            if cat_name and cat_name.lower() != "nan":
                key = cat_name.lower()
                if key not in cat_by_name:
                    nc = Category(name=cat_name)
                    await db.categories.insert_one(nc.model_dump())
                    cat_by_name[key] = nc.id
                cat_id = cat_by_name[key]
            weight = 1.0
            try:
                wv = str(row.get("weight", "") or "").strip()
                weight = float(wv) if wv and wv.lower() != "nan" else 1.0
            except (ValueError, TypeError):
                weight = 1.0
            if weight != weight:  # NaN guard
                weight = 1.0

            options, correct = [], None
            if qtype == "pg":
                slots = []
                for col in ("option_a", "option_b", "option_c", "option_d", "option_e"):
                    v = row.get(col, "")
                    v = "" if (v is None or str(v).lower() == "nan") else str(v).strip()
                    slots.append(v)
                # keep A..E positions intact (only drop trailing empty slots) so the
                # answer key letter always points at the right option
                while slots and slots[-1] == "":
                    slots.pop()
                options = slots
                raw_c = str(row.get("correct", "")).strip().lower()
                if raw_c in letter_idx:
                    correct = letter_idx[raw_c]
                elif raw_c.isdigit():
                    correct = raw_c
                else:
                    errors.append(f"Baris {rownum}: kunci PG tidak valid")
                    continue
                if int(correct) >= len(options) or not options[int(correct)]:
                    errors.append(f"Baris {rownum}: kunci PG menunjuk opsi yang kosong")
                    continue
            elif qtype == "truefalse":
                raw_c = str(row.get("correct", "")).strip().lower()
                correct = "true" if raw_c in ("true", "benar", "b", "1") else "false"

            image_path = None
            img_url = str(row.get("image_url", "") or "").strip()
            if img_url and img_url.lower() != "nan" and img_url.startswith("http"):
                image_path = await fetch_image_to_storage(img_url, user["id"])

            ques = Question(category_id=cat_id, type=qtype, text=text,
                            options=options, correct_answer=correct, weight=weight,
                            image_path=image_path)
            await db.questions.insert_one(ques.model_dump())
            imported += 1
        except Exception as e:
            errors.append(f"Baris {rownum}: {e}")

    return {"imported": imported, "errors": errors}


# ------------------------------------------------------------------ RESULT PDF
@api_router.get("/results/detail/{attempt_id}/pdf")
async def result_pdf(attempt_id: str, user: dict = Depends(get_current_user)):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    attempt = await db.attempts.find_one({"id": attempt_id}, {"_id": 0})
    if not attempt:
        raise HTTPException(status_code=404, detail="Tidak ditemukan")
    if user["role"] == "siswa" and attempt["student_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Akses ditolak")

    pkg = await db.packages.find_one({"id": attempt["package_id"]}, {"_id": 0}) or {}
    qlist = await db.questions.find({"id": {"$in": attempt_question_ids(attempt, pkg)}},
                                    {"_id": 0}).to_list(2000)
    qmap = {q["id"]: q for q in qlist}
    session = await db.sessions.find_one({"id": attempt["session_id"]}, {"_id": 0})
    kkm = session.get("kkm", 75) if session else 75

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm,
                            leftMargin=18 * mm, rightMargin=18 * mm)
    styles = getSampleStyleSheet()
    green = colors.HexColor("#1e3a30")
    terra = colors.HexColor("#c0563f")
    h = ParagraphStyle("h", parent=styles["Title"], textColor=green, fontSize=18, spaceAfter=2)
    sub = ParagraphStyle("sub", parent=styles["Normal"], textColor=colors.grey, fontSize=9)
    label = ParagraphStyle("lbl", parent=styles["Normal"], fontSize=9, textColor=colors.grey)
    val = ParagraphStyle("val", parent=styles["Normal"], fontSize=11)

    elems = []
    elems.append(Paragraph("KARTU HASIL UJIAN", h))
    elems.append(Paragraph("Computer Based Test", sub))
    elems.append(Spacer(1, 10 * mm))

    passed = attempt.get("score") is not None and attempt["score"] >= kkm
    score_txt = str(attempt["score"]) if attempt.get("score") is not None else "Menunggu Koreksi"
    info = [
        ["Nama Siswa", attempt.get("student_name", "-"), "Nilai Akhir", score_txt],
        ["NISN / NIP", attempt.get("student_identifier") or "-", "KKM", str(kkm)],
        ["Sesi Ujian", session["title"] if session else "-", "Status",
         "LULUS" if passed else ("BELUM LULUS" if attempt.get("score") is not None else "-")],
        ["Metode Nilai", "Berbobot" if pkg.get("scoring_method") == "weighted" else "Persentase",
         "Poin", f"{attempt.get('earned', 0)}/{attempt.get('total_possible', 0)}"],
    ]
    t = Table(info, colWidths=[32 * mm, 60 * mm, 28 * mm, 46 * mm])
    t.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.grey),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.grey),
        ("TEXTCOLOR", (3, 0), (3, 0), terra if not passed else green),
        ("FONTNAME", (3, 0), (3, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#e0e0d8")),
    ]))
    elems.append(t)
    elems.append(Spacer(1, 8 * mm))

    elems.append(Paragraph("Rincian Jawaban", ParagraphStyle("s2", parent=styles["Heading2"], textColor=green, fontSize=12)))
    elems.append(Spacer(1, 3 * mm))

    rows = [["No", "Soal", "Tipe", "Poin", "Hasil"]]
    tlabel = {"pg": "PG", "truefalse": "B/S", "essay": "Esai"}
    for i, d in enumerate(attempt.get("details", [])):
        q = qmap.get(d["question_id"], {})
        qtext = (q.get("text") or "")[:70] + ("..." if len(q.get("text") or "") > 70 else "")
        if d.get("type") == "essay":
            res = "Menunggu" if d.get("needs_grading") else "Dinilai"
        else:
            res = "Benar" if d.get("is_correct") else "Salah"
        rows.append([str(i + 1), Paragraph(qtext, ParagraphStyle("c", fontSize=8)),
                     tlabel.get(d.get("type"), "-"),
                     f"{d.get('points_earned', 0)}/{d.get('points_possible', 0)}", res])
    dt = Table(rows, colWidths=[10 * mm, 92 * mm, 16 * mm, 22 * mm, 22 * mm], repeatRows=1)
    dt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), green),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f6f0")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e0e0d8")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elems.append(dt)
    elems.append(Spacer(1, 10 * mm))
    elems.append(Paragraph(f"Dicetak pada {datetime.now(timezone.utc).strftime('%d-%m-%Y %H:%M UTC')}", sub))

    doc.build(elems)
    buf.seek(0)
    fname = f"hasil-{attempt.get('student_name', 'siswa')}.pdf".replace(" ", "_")
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})



# ------------------------------------------------------------------ IMAGE UPLOAD
ALLOWED_IMG = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
               "webp": "image/webp", "gif": "image/gif"}


async def fetch_image_to_storage(url: str, user_id: str) -> Optional[str]:
    """Download an image from a URL and store it. Returns storage path or None."""
    try:
        resp = requests.get(url, timeout=20, stream=True,
                            headers={"User-Agent": "Mozilla/5.0 (compatible; CBT-Ujian/1.0)"})
        resp.raise_for_status()
        data = resp.content
        if len(data) > 5 * 1024 * 1024:
            return None
        ct = resp.headers.get("Content-Type", "").split(";")[0].strip().lower()
        ext = {"image/png": "png", "image/jpeg": "jpg", "image/webp": "webp", "image/gif": "gif"}.get(ct)
        if not ext:
            ext = url.rsplit(".", 1)[-1].lower().split("?")[0]
            if ext not in ALLOWED_IMG:
                return None
            ct = ALLOWED_IMG[ext]
        path = f"{APP_NAME}/questions/{user_id}/{new_id()}.{ext}"
        result = put_object(path, data, ct)
        await db.files.insert_one({
            "id": new_id(), "storage_path": result["path"], "original_filename": url,
            "content_type": ct, "size": result.get("size"), "is_deleted": False,
            "created_at": now_iso(),
        })
        return result["path"]
    except Exception:
        return None


@api_router.post("/uploads/image")
async def upload_image(file: UploadFile = File(...), user: dict = Depends(require_roles("admin", "guru"))):
    fname = file.filename or ""
    ext = fname.rsplit(".", 1)[-1].lower() if "." in fname else ""
    if ext not in ALLOWED_IMG:
        raise HTTPException(status_code=400, detail="Format harus png/jpg/webp/gif")
    data = await file.read()
    if len(data) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Ukuran gambar maksimal 5MB")
    ct = ALLOWED_IMG[ext]
    path = f"{APP_NAME}/questions/{user['id']}/{new_id()}.{ext}"
    try:
        result = await store_file(path, data, ct)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Gagal mengunggah gambar: {e}")
    await db.files.insert_one({
        "id": new_id(), "storage_path": result["path"], "original_filename": fname,
        "content_type": ct, "size": result.get("size"), "is_deleted": False,
        "created_at": now_iso(),
    })
    return {"path": result["path"]}


@api_router.get("/files/{path:path}")
async def get_file(path: str, authorization: Optional[str] = Header(None), auth: Optional[str] = Query(None)):
    token = None
    if authorization and authorization.startswith("Bearer "):
        token = authorization[7:]
    elif auth:
        token = auth
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    record = await db.files.find_one({"storage_path": path, "is_deleted": False})
    if not record:
        raise HTTPException(status_code=404, detail="File tidak ditemukan")
    try:
        data, content_type = await load_file(path)
    except Exception:
        raise HTTPException(status_code=404, detail="File tidak ditemukan")
    return Response(content=data, media_type=record.get("content_type", content_type))


# ------------------------------------------------------------------ STUDENT REPORT PDF
@api_router.get("/report/student/{student_id}/pdf")
async def student_report_pdf(student_id: str, user: dict = Depends(get_current_user)):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart

    if student_id == "me":
        student_id = user["id"]
    if user["role"] == "siswa" and student_id != user["id"]:
        raise HTTPException(status_code=403, detail="Akses ditolak")
    stu = await db.users.find_one({"_id": ObjectId(student_id)})
    if not stu:
        raise HTTPException(status_code=404, detail="Siswa tidak ditemukan")

    attempts = await db.attempts.find(
        {"student_id": student_id, "status": {"$ne": "berlangsung"}}, {"_id": 0}
    ).sort("submitted_at", 1).to_list(2000)
    pkgs = await db.packages.find({}, {"_id": 0, "id": 1, "category_id": 1}).to_list(2000)
    pkg_cat = {p["id"]: p.get("category_id") for p in pkgs}
    cats = await db.categories.find({}, {"_id": 0}).to_list(1000)
    cat_name = {c["id"]: c["name"] for c in cats}

    rows_data = []
    scores = []
    labels = []
    for a in attempts:
        s = await db.sessions.find_one({"id": a["session_id"]}, {"_id": 0})
        subj = cat_name.get(pkg_cat.get(a.get("package_id")), "Umum")
        kkm = s.get("kkm", 75) if s else 75
        sc = a.get("score")
        status = "Lulus" if (sc is not None and sc >= kkm) else ("Belum Lulus" if sc is not None else "Menunggu")
        rows_data.append([s["title"] if s else "-", subj, str(sc) if sc is not None else "-", str(kkm), status])
        if sc is not None:
            scores.append(sc)
            labels.append((s["title"] if s else "-")[:10])

    green = colors.HexColor("#1e3a30")
    terra = colors.HexColor("#c0563f")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm, leftMargin=18 * mm, rightMargin=18 * mm)
    styles = getSampleStyleSheet()
    sub = ParagraphStyle("sub", parent=styles["Normal"], textColor=colors.grey, fontSize=9)
    elems = await _school_kop(styles, green, sub)
    elems += [Paragraph("RAPOR HASIL BELAJAR SISWA", ParagraphStyle("h", parent=styles["Title"], textColor=green, fontSize=18, spaceAfter=2)),
             Paragraph("Computer Based Test", sub), Spacer(1, 8 * mm)]

    avg = round(sum(scores) / len(scores), 1) if scores else 0
    info = [["Nama Siswa", stu["name"], "Rata-rata", str(avg)],
            ["NISN / NIP", stu.get("identifier", "") or "-", "Total Ujian", str(len(scores))]]
    t = Table(info, colWidths=[32 * mm, 62 * mm, 30 * mm, 42 * mm])
    t.setStyle(TableStyle([("FONTSIZE", (0, 0), (-1, -1), 9), ("TEXTCOLOR", (0, 0), (0, -1), colors.grey),
                           ("TEXTCOLOR", (2, 0), (2, -1), colors.grey), ("FONTNAME", (3, 0), (3, 0), "Helvetica-Bold"),
                           ("TEXTCOLOR", (3, 0), (3, 0), green), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                           ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#e0e0d8"))]))
    elems += [t, Spacer(1, 8 * mm)]

    if scores:
        elems.append(Paragraph("Grafik Perkembangan Nilai", ParagraphStyle("s2", parent=styles["Heading2"], textColor=green, fontSize=12)))
        d = Drawing(460, 190)
        bc = VerticalBarChart()
        bc.x, bc.y, bc.height, bc.width = 30, 20, 150, 420
        bc.data = [scores]
        bc.categoryAxis.categoryNames = labels
        bc.categoryAxis.labels.fontSize = 6
        bc.categoryAxis.labels.angle = 30
        bc.categoryAxis.labels.dy = -6
        bc.valueAxis.valueMin, bc.valueAxis.valueMax, bc.valueAxis.valueStep = 0, 100, 20
        bc.bars[0].fillColor = green
        d.add(bc)
        elems += [d, Spacer(1, 6 * mm)]

    elems.append(Paragraph("Rincian Nilai", ParagraphStyle("s3", parent=styles["Heading2"], textColor=green, fontSize=12)))
    trows = [["Sesi Ujian", "Mapel", "Nilai", "KKM", "Status"]] + (rows_data or [["Belum ada ujian", "-", "-", "-", "-"]])
    dt = Table(trows, colWidths=[70 * mm, 40 * mm, 20 * mm, 18 * mm, 26 * mm], repeatRows=1)
    dt.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), green), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                            ("FONTSIZE", (0, 0), (-1, -1), 8), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f6f0")]),
                            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e0e0d8")),
                            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
    elems += [dt, Spacer(1, 10 * mm),
              Paragraph(f"Dicetak pada {datetime.now(timezone.utc).strftime('%d-%m-%Y %H:%M UTC')}", sub)]
    doc.build(elems)
    buf.seek(0)
    fname = f"rapor-{stu['name']}.pdf".replace(" ", "_")
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


# ------------------------------------------------------------------ CLASS REPORT PDF (bulk)
@api_router.get("/report/class/{class_id}/pdf")
async def class_report_pdf(class_id: str, user: dict = Depends(require_roles("admin", "guru"))):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart

    cls = await db.classes.find_one({"id": class_id}, {"_id": 0})
    if not cls:
        raise HTTPException(status_code=404, detail="Kelas tidak ditemukan")
    sids = cls.get("student_ids", [])
    students = []
    if sids:
        docs = await db.users.find({"_id": {"$in": [ObjectId(s) for s in sids]}}).to_list(2000)
        students = sorted(docs, key=lambda u: u["name"].lower())
    pkgs = await db.packages.find({}, {"_id": 0, "id": 1, "category_id": 1}).to_list(2000)
    pkg_cat = {p["id"]: p.get("category_id") for p in pkgs}
    cats = await db.categories.find({}, {"_id": 0}).to_list(1000)
    cat_name = {c["id"]: c["name"] for c in cats}

    green = colors.HexColor("#1e3a30")
    styles = getSampleStyleSheet()
    sub = ParagraphStyle("sub", parent=styles["Normal"], textColor=colors.grey, fontSize=9)
    name_style = ParagraphStyle("nm", parent=styles["Heading1"], textColor=green, fontSize=15)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm, leftMargin=18 * mm, rightMargin=18 * mm)
    elems = await _school_kop(styles, green, sub)
    elems += [Paragraph(f"RAPOR KELAS {cls['name'].upper()}", ParagraphStyle("h", parent=styles["Title"], textColor=green, fontSize=18, spaceAfter=2)),
             Paragraph(f"{len(students)} siswa · Computer Based Test", sub), Spacer(1, 8 * mm)]
    if not students:
        elems.append(Paragraph("Belum ada siswa di kelas ini.", styles["Normal"]))

    for i, stu in enumerate(students):
        if i > 0:
            elems.append(PageBreak())
        sid = str(stu["_id"])
        attempts = await db.attempts.find({"student_id": sid, "status": {"$ne": "berlangsung"}}, {"_id": 0}).sort("submitted_at", 1).to_list(2000)
        rows_data, scores, labels = [], [], []
        for a in attempts:
            s = await db.sessions.find_one({"id": a["session_id"]}, {"_id": 0})
            subj = cat_name.get(pkg_cat.get(a.get("package_id")), "Umum")
            kkm = s.get("kkm", 75) if s else 75
            sc = a.get("score")
            status = "Lulus" if (sc is not None and sc >= kkm) else ("Belum Lulus" if sc is not None else "Menunggu")
            rows_data.append([s["title"] if s else "-", subj, str(sc) if sc is not None else "-", str(kkm), status])
            if sc is not None:
                scores.append(sc)
                labels.append((s["title"] if s else "-")[:10])
        elems.append(Paragraph(f"Rapor: {stu['name']}", name_style))
        avg = round(sum(scores) / len(scores), 1) if scores else 0
        info = [["Nama", stu["name"], "Rata-rata", str(avg)],
                ["NISN/NIP", stu.get("identifier", "") or "-", "Total Ujian", str(len(scores))]]
        t = Table(info, colWidths=[28 * mm, 64 * mm, 28 * mm, 42 * mm])
        t.setStyle(TableStyle([("FONTSIZE", (0, 0), (-1, -1), 9), ("TEXTCOLOR", (0, 0), (0, -1), colors.grey),
                               ("TEXTCOLOR", (2, 0), (2, -1), colors.grey), ("FONTNAME", (3, 0), (3, 0), "Helvetica-Bold"),
                               ("TEXTCOLOR", (3, 0), (3, 0), green), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                               ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#e0e0d8"))]))
        elems += [t, Spacer(1, 5 * mm)]
        if scores:
            d = Drawing(460, 170)
            bc = VerticalBarChart()
            bc.x, bc.y, bc.height, bc.width = 30, 15, 135, 420
            bc.data = [scores]
            bc.categoryAxis.categoryNames = labels
            bc.categoryAxis.labels.fontSize = 6
            bc.categoryAxis.labels.angle = 30
            bc.categoryAxis.labels.dy = -6
            bc.valueAxis.valueMin, bc.valueAxis.valueMax, bc.valueAxis.valueStep = 0, 100, 20
            bc.bars[0].fillColor = green
            d.add(bc)
            elems += [d, Spacer(1, 4 * mm)]
        trows = [["Sesi Ujian", "Mapel", "Nilai", "KKM", "Status"]] + (rows_data or [["Belum ada ujian", "-", "-", "-", "-"]])
        dt = Table(trows, colWidths=[70 * mm, 40 * mm, 20 * mm, 18 * mm, 26 * mm], repeatRows=1)
        dt.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), green), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                                ("FONTSIZE", (0, 0), (-1, -1), 8), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f6f0")]),
                                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e0e0d8")),
                                ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
        elems.append(dt)

    doc.build(elems)
    buf.seek(0)
    fname = f"rapor-kelas-{cls['name']}.pdf".replace(" ", "_")
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


# ------------------------------------------------------------------ AUTO-SUBMIT (cron)
async def run_auto_submit():
    now = datetime.now(timezone.utc)
    attempts = await db.attempts.find({"status": "berlangsung"}, {"_id": 0}).to_list(5000)
    count = 0
    for att in attempts:
        session = await db.sessions.find_one({"id": att["session_id"]}, {"_id": 0})
        if not session:
            continue
        # Batas waktu efektif: jendela susulan bila attempt dibuat lewat susulan.
        end_iso = att.get("effective_end") or session["end_time"]
        duration = att.get("effective_duration") or session.get("duration_minutes", 60)
        try:
            end = datetime.fromisoformat(end_iso)
            started = datetime.fromisoformat(att["started_at"])
        except Exception:
            continue
        deadline = min(end, started + timedelta(minutes=duration))
        if now >= deadline:
            await finalize_attempt(att, att.get("answers", {}))
            count += 1
    logging.getLogger(__name__).info(f"Auto-submit finalized {count} attempt(s)")
    return count


@api_router.post("/cron/auto-submit")
async def cron_auto_submit(request: Request, background: BackgroundTasks):
    # Cron endpoints must ack 2xx immediately; enqueue/background the actual work.
    secret = os.environ.get("WEBHOOK_CRON_SECRET", "")
    auth = request.headers.get("Authorization", "")
    token = auth[7:] if auth.startswith("Bearer ") else ""
    if not secret or not hmac.compare_digest(token, secret):
        raise HTTPException(status_code=401, detail="Unauthorized")
    background.add_task(run_auto_submit)
    return {"accepted": True}


# ------------------------------------------------------------------ ITEM ANALYTICS
@api_router.get("/analytics/session/{session_id}")
async def analytics_session(session_id: str, user: dict = Depends(require_roles("admin", "guru"))):
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Sesi tidak ditemukan")
    pkg = await db.packages.find_one({"id": session["package_id"]}, {"_id": 0})
    if not pkg:
        raise HTTPException(status_code=400, detail="Paket soal untuk sesi ini sudah dihapus")
    qlist = await db.questions.find({"id": {"$in": pkg.get("question_ids", [])}}, {"_id": 0}).to_list(2000)
    qmap = {q["id"]: q for q in qlist}
    settings = await db.settings.find_one({"key": "difficulty"}, {"_id": 0})
    pkg_easy = pkg.get("easy_min")
    pkg_med = pkg.get("medium_min")
    if pkg_easy is not None and pkg_med is not None:
        easy_min, medium_min, source = pkg_easy, pkg_med, "paket"
    else:
        easy_min = settings.get("easy_min", 70) if settings else 70
        medium_min = settings.get("medium_min", 40) if settings else 40
        source = "global"
    attempts = await db.attempts.find(
        {"session_id": session_id, "status": {"$in": ["selesai", "menunggu_koreksi"]}}, {"_id": 0}
    ).to_list(5000)

    items = []
    for qid in pkg.get("question_ids", []):
        q = qmap.get(qid)
        if not q:
            continue
        total = answered = correct_count = 0
        pts_earned = pts_possible = 0.0
        for att in attempts:
            det = next((d for d in att.get("details", []) if d["question_id"] == qid), None)
            if not det:
                continue
            total += 1
            if det.get("answer") not in (None, ""):
                answered += 1
            if q["type"] in ("pg", "truefalse"):
                if det.get("is_correct"):
                    correct_count += 1
            else:
                pts_earned += det.get("points_earned", 0) or 0
                pts_possible += det.get("points_possible", 0) or 0
        if q["type"] in ("pg", "truefalse"):
            p = (correct_count / total) if total else 0
        else:
            p = (pts_earned / pts_possible) if pts_possible else 0
        pct = round(p * 100, 1)
        difficulty = "Mudah" if pct >= easy_min else ("Sedang" if pct >= medium_min else "Sulit")
        items.append({
            "question_id": qid, "text": q["text"], "type": q["type"],
            "total": total, "answered": answered,
            "correct": correct_count if q["type"] in ("pg", "truefalse") else None,
            "percent_correct": pct, "difficulty": difficulty,
        })
    return {"session_title": session["title"], "participants": len(attempts),
            "items": items, "thresholds": {"easy_min": easy_min, "medium_min": medium_min, "source": source}}


# ------------------------------------------------------------------ SESSION RESULT EXPORT (Excel)
STATUS_ID = {"selesai": "Selesai", "menunggu_koreksi": "Menunggu Koreksi",
             "berlangsung": "Berlangsung"}
QTYPE_ID = {"pg": "Pilihan Ganda", "truefalse": "Benar/Salah", "essay": "Esai"}
MONTH_ID = ["", "Jan", "Feb", "Mar", "Apr", "Mei", "Jun",
            "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]


def fmt_dt_id(iso: Optional[str]) -> str:
    """'19 Agu 2026, 14.58' — readable Indonesian date/time for spreadsheets."""
    if not iso:
        return "-"
    try:
        dt = datetime.fromisoformat(iso)
    except (ValueError, TypeError):
        return "-"
    return f"{dt.day} {MONTH_ID[dt.month]} {dt.year}, {dt:%H.%M}"


def _xl_comment(text: str):
    from openpyxl.comments import Comment
    c = Comment(text, "CBT Ujian")
    c.width = 340
    c.height = 110
    return c


XL_GREEN = "1E3A30"
XL_GREEN_SOFT = "E8EDEA"
XL_TERRA = "C0563F"
XL_TERRA_SOFT = "FBEAE5"
XL_STRIPE = "F6F6F0"
XL_LINE = "D9D9CF"
XL_GOLD_SOFT = "FDF3D8"


def _xl_border(color=XL_LINE):
    from openpyxl.styles import Border, Side
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _predikat(score: Optional[float]) -> str:
    if score is None:
        return "-"
    if score >= 90:
        return "A"
    if score >= 80:
        return "B"
    if score >= 70:
        return "C"
    if score >= 60:
        return "D"
    return "E"


@api_router.get("/export/session/{session_id}/xlsx")
async def export_session_results(session_id: str, user: dict = Depends(require_roles("admin", "guru"))):
    """A polished, print-ready workbook of a session's results.

    Sheet 1 "Rekap Nilai"    - school letterhead, session info, per-student scores + summary
    Sheet 2 "Rincian Jawaban" - point matrix (students x questions)
    Sheet 3 "Analisis Butir"  - per-question difficulty analysis
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Sesi tidak ditemukan")
    pkg = await db.packages.find_one({"id": session["package_id"]}, {"_id": 0}) or {}
    cat = await db.categories.find_one({"id": pkg.get("category_id")}, {"_id": 0}) or {}
    school = await db.settings.find_one({"key": "school"}, {"_id": 0}) or {}
    attempts = await db.attempts.find(
        {"session_id": session_id, "status": {"$ne": "berlangsung"}}, {"_id": 0}
    ).to_list(5000)
    attempts.sort(key=lambda a: (a.get("student_name") or "").lower())

    q_ids = list(pkg.get("question_ids") or [])
    if not q_ids and attempts:
        q_ids = [d.get("question_id") for d in (attempts[0].get("details") or [])]
    qlist = await db.questions.find({"id": {"$in": q_ids}}, {"_id": 0}).to_list(2000)
    qmap = {q["id"]: q for q in qlist}

    # class name per student
    classes = await db.classes.find({}, {"_id": 0}).to_list(1000)
    cls_of = {}
    for c in classes:
        for sid in c.get("student_ids", []):
            cls_of.setdefault(sid, []).append(c["name"])

    kkm = float(session.get("kkm", 75) or 75)
    weighted = pkg.get("scoring_method") == "weighted"
    thin = _xl_border()

    wb = Workbook()

    # ============================================================ SHEET 1
    ws = wb.active
    ws.title = "Rekap Nilai"
    headers = ["No", "Nama Siswa", "NISN/NIP", "Kelas", "Status", "Benar", "Salah",
               "Kosong", "Poin", "Nilai", "Predikat", "Keterangan", "Pelanggaran",
               "Waktu Kumpul"]
    ncol = len(headers)
    last_col = get_column_letter(ncol)

    def band(row, text, *, size=11, bold=True, color=XL_GREEN, height=None, italic=False):
        ws.merge_cells(f"A{row}:{last_col}{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=bold, size=size, color=color, italic=italic)
        c.alignment = Alignment(horizontal="center", vertical="center")
        if height:
            ws.row_dimensions[row].height = height
        return c

    r = 1
    if school.get("name"):
        band(r, school["name"].upper(), size=14, height=20); r += 1
    if school.get("address"):
        band(r, school["address"], size=9, bold=False, color="7A7A72"); r += 1
    band(r, "REKAP NILAI HASIL UJIAN", size=13, height=22); r += 1
    band(r, session["title"], size=11, bold=False, color="7A7A72"); r += 1
    r += 1

    # session info block (2 label/value pairs per row)
    info = [
        ("Paket Soal", pkg.get("title", "-")),
        ("Mata Pelajaran", cat.get("name", "Umum")),
        ("Jumlah Soal", len(q_ids)),
        ("Metode Penilaian", "Berbobot" if weighted else "Persentase"),
        ("Durasi", f"{session.get('duration_minutes', 0)} menit"),
        ("KKM", kkm),
        ("Mulai", fmt_dt_id(session.get("start_time"))),
        ("Selesai", fmt_dt_id(session.get("end_time"))),
    ]
    info_start = r
    for i in range(0, len(info), 2):
        for j, (label, value) in enumerate(info[i:i + 2]):
            lc = 1 + j * 7
            lab = ws.cell(row=r, column=lc, value=label)
            lab.font = Font(size=9, color="7A7A72", bold=True)
            ws.merge_cells(start_row=r, start_column=lc + 1, end_row=r, end_column=lc + 5)
            val = ws.cell(row=r, column=lc + 1, value=value)
            val.font = Font(size=9)
            val.alignment = Alignment(horizontal="left", vertical="center")
        r += 1
    ws.cell(row=info_start, column=1)
    r += 1

    head_row = r
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=head_row, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=XL_GREEN)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin
    ws.row_dimensions[head_row].height = 30

    scores, rows_written = [], 0
    for idx, a in enumerate(attempts, start=1):
        details = a.get("details") or []
        benar = sum(1 for d in details if d.get("is_correct") is True)
        salah = sum(1 for d in details if d.get("is_correct") is False)
        kosong = sum(1 for d in details if d.get("answer") in (None, ""))
        sc = a.get("score")
        rr = head_row + idx
        vals = [
            idx, a.get("student_name", "-"), a.get("student_identifier") or "-",
            ", ".join(cls_of.get(a.get("student_id"), [])) or "-",
            STATUS_ID.get(a.get("status"), a.get("status", "-")),
            benar, salah, kosong,
            f"{a.get('earned', 0)}/{a.get('total_possible', 0)}",
            sc if sc is not None else "-",
            _predikat(sc),
            ("Lulus" if sc >= kkm else "Belum Lulus") if sc is not None else "Menunggu",
            len(a.get("violations") or []),
            fmt_dt_id(a.get("submitted_at")) + (" (Susulan)" if a.get("is_makeup") else ""),
        ]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=rr, column=ci, value=v)
            c.border = thin
            c.font = Font(size=10, bold=ci in (2, 10))
            c.alignment = Alignment(
                horizontal="left" if ci in (2, 3, 4) else "center", vertical="center")
            if idx % 2 == 0:
                c.fill = PatternFill("solid", fgColor=XL_STRIPE)
        nviol = len(a.get("violations") or [])
        if nviol:
            vc = ws.cell(row=rr, column=13)
            vc.fill = PatternFill("solid", fgColor=XL_TERRA_SOFT)
            vc.font = Font(size=10, bold=True, color=XL_TERRA)
        if sc is not None:
            ws.cell(row=rr, column=10).number_format = "0.00"
            below = sc < kkm
            for ci in (10, 11, 12):
                cc = ws.cell(row=rr, column=ci)
                cc.fill = PatternFill("solid", fgColor=XL_TERRA_SOFT if below else XL_GREEN_SOFT)
                cc.font = Font(size=10, bold=True, color=XL_TERRA if below else XL_GREEN)
            scores.append(sc)
        rows_written += 1

    if rows_written == 0:
        rr = head_row + 1
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=ncol)
        c = ws.cell(row=rr, column=1, value="Belum ada peserta yang mengumpulkan.")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(size=10, italic=True, color="7A7A72")
        c.border = thin
        rows_written = 1

    # summary block
    sr = head_row + rows_written + 2
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=ncol)
    t = ws.cell(row=sr, column=1, value="RINGKASAN")
    t.font = Font(bold=True, size=10, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=XL_GREEN)
    t.alignment = Alignment(horizontal="center", vertical="center")

    lulus = sum(1 for x in scores if x >= kkm)
    summary = [
        ("Jumlah Peserta", len(attempts)),
        ("Sudah Dinilai", len(scores)),
        ("Rata-rata Nilai", round(sum(scores) / len(scores), 2) if scores else "-"),
        ("Nilai Tertinggi", max(scores) if scores else "-"),
        ("Nilai Terendah", min(scores) if scores else "-"),
        ("Jumlah Lulus", lulus),
        ("Belum Lulus", len(scores) - lulus),
        ("Ketuntasan", f"{round(lulus / len(scores) * 100, 1)}%" if scores else "-"),
    ]
    row = sr + 1
    for i in range(0, len(summary), 2):
        for j, (label, value) in enumerate(summary[i:i + 2]):
            lc = 1 + j * 7
            ws.merge_cells(start_row=row, start_column=lc, end_row=row, end_column=lc + 2)
            lab = ws.cell(row=row, column=lc, value=label)
            lab.font = Font(size=10, color="7A7A72", bold=True)
            lab.alignment = Alignment(horizontal="left", vertical="center")
            lab.border = thin
            ws.merge_cells(start_row=row, start_column=lc + 3, end_row=row, end_column=lc + 5)
            val = ws.cell(row=row, column=lc + 3, value=value)
            val.font = Font(size=10, bold=True, color=XL_GREEN)
            val.alignment = Alignment(horizontal="center", vertical="center")
            val.fill = PatternFill("solid", fgColor=XL_GREEN_SOFT)
            val.border = thin
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    note = ws.cell(row=row, column=1, value=(
        f"Predikat: A ≥ 90 · B ≥ 80 · C ≥ 70 · D ≥ 60 · E < 60   |   "
        f"Dicetak {fmt_dt_id(now_iso())}"))
    note.font = Font(size=8, italic=True, color="9A9A92")
    note.alignment = Alignment(horizontal="center")

    widths = [5, 28, 15, 14, 16, 8, 8, 9, 11, 9, 10, 14, 12, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)
    ws.auto_filter.ref = f"A{head_row}:{last_col}{head_row + rows_written}"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{head_row}:{head_row}"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

    # ============================================================ SHEET 2
    ws2 = wb.create_sheet("Rincian Jawaban")
    qcols = [qmap[q] for q in q_ids if q in qmap]
    h2 = ["No", "Nama Siswa", "NISN/NIP"] + [f"S{i + 1}" for i in range(len(qcols))] + ["Poin", "Nilai"]
    n2 = len(h2)
    lc2 = get_column_letter(n2)
    ws2.merge_cells(f"A1:{lc2}1")
    c = ws2.cell(row=1, column=1, value=f"RINCIAN PEROLEHAN POIN PER SOAL — {session['title']}")
    c.font = Font(bold=True, size=12, color=XL_GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 20
    ws2.merge_cells(f"A2:{lc2}2")
    c = ws2.cell(row=2, column=1, value="Angka pada kolom soal = poin yang diperoleh siswa. Arahkan kursor ke judul kolom untuk melihat teks soal.")
    c.font = Font(size=8, italic=True, color="9A9A92")
    c.alignment = Alignment(horizontal="center")

    for i, h in enumerate(h2, start=1):
        c = ws2.cell(row=4, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=XL_GREEN)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin
        if 4 <= i <= 3 + len(qcols):
            q = qcols[i - 4]
            kunci = ""
            if q["type"] == "pg" and q.get("correct_answer") is not None:
                try:
                    kunci = f" | Kunci: {chr(65 + int(q['correct_answer']))}"
                except (ValueError, TypeError):
                    kunci = ""
            elif q["type"] == "truefalse":
                kunci = f" | Kunci: {'Benar' if q.get('correct_answer') == 'true' else 'Salah'}"
            c.comment = _xl_comment(f"[{QTYPE_ID.get(q['type'], q['type'])}] {q['text']}{kunci}")
    ws2.row_dimensions[4].height = 24

    for idx, a in enumerate(attempts, start=1):
        rr = 4 + idx
        dmap = {d["question_id"]: d for d in (a.get("details") or [])}
        vals = [idx, a.get("student_name", "-"), a.get("student_identifier") or "-"]
        for q in qcols:
            d = dmap.get(q["id"])
            vals.append(d.get("points_earned") if d else "-")
        vals += [f"{a.get('earned', 0)}/{a.get('total_possible', 0)}",
                 a.get("score") if a.get("score") is not None else "-"]
        for ci, v in enumerate(vals, start=1):
            c = ws2.cell(row=rr, column=ci, value=v)
            c.border = thin
            c.font = Font(size=10, bold=ci in (2, n2))
            c.alignment = Alignment(horizontal="left" if ci in (2, 3) else "center", vertical="center")
            if idx % 2 == 0:
                c.fill = PatternFill("solid", fgColor=XL_STRIPE)
            if 4 <= ci <= 3 + len(qcols):
                d = dmap.get(qcols[ci - 4]["id"])
                if d and d.get("is_correct") is True:
                    c.fill = PatternFill("solid", fgColor=XL_GREEN_SOFT)
                elif d and d.get("is_correct") is False:
                    c.fill = PatternFill("solid", fgColor=XL_TERRA_SOFT)
                elif d and d.get("type") == "essay":
                    c.fill = PatternFill("solid", fgColor=XL_GOLD_SOFT)

    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 15
    for i in range(4, 4 + len(qcols)):
        ws2.column_dimensions[get_column_letter(i)].width = 6.5
    ws2.column_dimensions[get_column_letter(n2 - 1)].width = 11
    ws2.column_dimensions[get_column_letter(n2)].width = 9
    ws2.freeze_panes = "D5"
    ws2.sheet_view.showGridLines = False
    ws2.page_setup.orientation = "landscape"
    ws2.page_setup.fitToWidth = 1
    ws2.sheet_properties.pageSetUpPr.fitToPage = True

    # ============================================================ SHEET 3
    ws3 = wb.create_sheet("Analisis Butir")
    settings = await db.settings.find_one({"key": "difficulty"}, {"_id": 0})
    if pkg.get("easy_min") is not None and pkg.get("medium_min") is not None:
        easy_min, medium_min, src = pkg["easy_min"], pkg["medium_min"], "khusus paket"
    else:
        easy_min = settings.get("easy_min", 70) if settings else 70
        medium_min = settings.get("medium_min", 40) if settings else 40
        src = "global"

    h3 = ["No", "Tipe", "Soal", "Kunci", "Benar", "Peserta", "% Benar", "Kesukaran"]
    lc3 = get_column_letter(len(h3))
    ws3.merge_cells(f"A1:{lc3}1")
    c = ws3.cell(row=1, column=1, value=f"ANALISIS BUTIR SOAL — {session['title']}")
    c.font = Font(bold=True, size=12, color=XL_GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 20
    ws3.merge_cells(f"A2:{lc3}2")
    c = ws3.cell(row=2, column=1, value=(
        f"Ambang ({src}): Mudah ≥ {easy_min}% · Sedang ≥ {medium_min}% · Sulit < {medium_min}%"))
    c.font = Font(size=8, italic=True, color="9A9A92")
    c.alignment = Alignment(horizontal="center")

    for i, h in enumerate(h3, start=1):
        c = ws3.cell(row=4, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=XL_GREEN)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin
    ws3.row_dimensions[4].height = 24

    diff_fill = {"Mudah": XL_GREEN_SOFT, "Sedang": XL_GOLD_SOFT, "Sulit": XL_TERRA_SOFT}
    diff_font = {"Mudah": XL_GREEN, "Sedang": "8A6D1F", "Sulit": XL_TERRA}
    for i, q in enumerate(qcols, start=1):
        total = benar = 0
        pts_e = pts_p = 0.0
        for a in attempts:
            d = next((x for x in (a.get("details") or []) if x["question_id"] == q["id"]), None)
            if not d:
                continue
            total += 1
            if q["type"] in ("pg", "truefalse"):
                if d.get("is_correct"):
                    benar += 1
            else:
                pts_e += d.get("points_earned") or 0
                pts_p += d.get("points_possible") or 0
        if q["type"] in ("pg", "truefalse"):
            pct = round(benar / total * 100, 1) if total else 0.0
        else:
            pct = round(pts_e / pts_p * 100, 1) if pts_p else 0.0
        label = "Mudah" if pct >= easy_min else ("Sedang" if pct >= medium_min else "Sulit")
        kunci = "-"
        if q["type"] == "pg" and q.get("correct_answer") is not None:
            try:
                kunci = chr(65 + int(q["correct_answer"]))
            except (ValueError, TypeError):
                kunci = "-"
        elif q["type"] == "truefalse":
            kunci = "Benar" if q.get("correct_answer") == "true" else "Salah"
        rr = 4 + i
        vals = [i, QTYPE_ID.get(q["type"], q["type"]), q["text"], kunci,
                benar if q["type"] in ("pg", "truefalse") else "-", total, pct / 100, label]
        for ci, v in enumerate(vals, start=1):
            c = ws3.cell(row=rr, column=ci, value=v)
            c.border = thin
            c.font = Font(size=10)
            c.alignment = Alignment(horizontal="left" if ci == 3 else "center",
                                    vertical="center", wrap_text=ci == 3)
            if i % 2 == 0:
                c.fill = PatternFill("solid", fgColor=XL_STRIPE)
        ws3.cell(row=rr, column=7).number_format = "0.0%"
        lab = ws3.cell(row=rr, column=8)
        lab.fill = PatternFill("solid", fgColor=diff_fill[label])
        lab.font = Font(size=10, bold=True, color=diff_font[label])

    for col, w in zip("ABCDEFGH", [5, 13, 62, 9, 9, 9, 10, 12]):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A5"
    ws3.sheet_view.showGridLines = False
    ws3.page_setup.orientation = "landscape"
    ws3.page_setup.fitToWidth = 1
    ws3.sheet_properties.pageSetUpPr.fitToPage = True

    # ============================================================ SHEET 4 (lockdown)
    # include in-progress attempts here so teachers can spot cheating live
    all_attempts = await db.attempts.find({"session_id": session_id}, {"_id": 0}).to_list(5000)
    all_attempts.sort(key=lambda a: (a.get("student_name") or "").lower())
    viol_rows = []
    for a in all_attempts:
        for k, v in enumerate(a.get("violations") or [], start=1):
            viol_rows.append([a.get("student_name", "-"), a.get("student_identifier") or "-",
                              k, v.get("label") or v.get("type", "-"), fmt_dt_id(v.get("at")),
                              "Ya" if a.get("auto_submitted_reason") == "pelanggaran" else "Tidak",
                              STATUS_ID.get(a.get("status"), a.get("status", "-"))])
    ws4 = wb.create_sheet("Pelanggaran")
    ws4.sheet_view.showGridLines = False
    h4 = ["Nama Siswa", "NISN/NIP", "Pelanggaran ke-", "Jenis Pelanggaran",
          "Waktu", "Dikumpulkan Otomatis", "Status Ujian"]
    lc4 = get_column_letter(len(h4))
    ws4.merge_cells(f"A1:{lc4}1")
    c = ws4.cell(row=1, column=1, value=f"CATATAN PELANGGARAN MODE UJIAN KETAT — {session['title']}")
    c.font = Font(bold=True, size=12, color=XL_GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 20
    ws4.merge_cells(f"A2:{lc4}2")
    c = ws4.cell(row=2, column=1, value=(
        "Tercatat setiap kali siswa keluar dari layar ujian (pindah tab, minimize, keluar layar penuh, "
        f"atau menekan tombol terlarang). Batas pelanggaran: {(await get_exam_lock())['max_violations']}x."))
    c.font = Font(size=8, italic=True, color="9A9A92")
    c.alignment = Alignment(horizontal="center")
    for i, h in enumerate(h4, start=1):
        cell = ws4.cell(row=4, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=XL_GREEN)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws4.row_dimensions[4].height = 24
    if viol_rows:
        for ri, row in enumerate(viol_rows, start=5):
            for ci, v in enumerate(row, start=1):
                cell = ws4.cell(row=ri, column=ci, value=v)
                cell.border = thin
                cell.font = Font(size=10, bold=ci == 1)
                cell.alignment = Alignment(horizontal="left" if ci in (1, 2, 4) else "center",
                                           vertical="center")
                if ri % 2 == 1:
                    cell.fill = PatternFill("solid", fgColor=XL_STRIPE)
            if row[5] == "Ya":
                cell = ws4.cell(row=ri, column=6)
                cell.fill = PatternFill("solid", fgColor=XL_TERRA_SOFT)
                cell.font = Font(size=10, bold=True, color=XL_TERRA)
        ws4.auto_filter.ref = f"A4:{lc4}{4 + len(viol_rows)}"
    else:
        ws4.merge_cells(start_row=5, start_column=1, end_row=5, end_column=len(h4))
        cell = ws4.cell(row=5, column=1, value="Tidak ada pelanggaran tercatat pada sesi ini.")
        cell.font = Font(size=10, italic=True, color="7A7A72")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
    for col, w in zip("ABCDEFG", [28, 15, 15, 30, 20, 20, 18]):
        ws4.column_dimensions[col].width = w
    ws4.freeze_panes = "A5"
    ws4.page_setup.orientation = "landscape"
    ws4.page_setup.fitToWidth = 1
    ws4.sheet_properties.pageSetUpPr.fitToPage = True


    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = "".join(ch if ch.isalnum() or ch in "-_ " else "" for ch in session["title"]).strip()
    fname = f"hasil-{safe or 'sesi'}.xlsx".replace(" ", "_")
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})



# ------------------------------------------------------------------ CLASS GRADE EXPORT (Excel)
@api_router.get("/export/class/{class_id}/xlsx")
async def export_class_grades(class_id: str, user: dict = Depends(require_roles("admin", "guru"))):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    cls = await db.classes.find_one({"id": class_id}, {"_id": 0})
    if not cls:
        raise HTTPException(status_code=404, detail="Kelas tidak ditemukan")
    sids = cls.get("student_ids", [])
    students = []
    if sids:
        docs = await db.users.find({"_id": {"$in": [ObjectId(s) for s in sids]}}).to_list(2000)
        students = sorted([clean_user(u) for u in docs], key=lambda x: x["name"].lower())
    sessions = await db.sessions.find(
        {"$or": [{"class_ids": class_id}, {"class_ids": {"$size": 0}}, {"class_ids": {"$exists": False}}]},
        {"_id": 0}
    ).sort("start_time", 1).to_list(1000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Nilai"
    ws.append([f"REKAP NILAI - {cls['name']}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(4, 4 + len(sessions)))
    ws["A1"].font = Font(bold=True, size=14, color="1E3A30")
    header = ["No", "Nama Siswa", "NISN/NIP"] + [s["title"] for s in sessions] + ["Rata-rata"]
    ws.append(header)
    green = PatternFill("solid", fgColor="1E3A30")
    thin = Side(style="thin", color="D9D9CF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = green
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for i, st in enumerate(students):
        row = [i + 1, st["name"], st.get("identifier", "")]
        scores = []
        for s in sessions:
            att = await db.attempts.find_one({"session_id": s["id"], "student_id": st["id"]}, {"_id": 0})
            val = att.get("score") if att and att.get("score") is not None else None
            row.append(val if val is not None else "-")
            if val is not None:
                scores.append(val)
        row.append(round(sum(scores) / len(scores), 1) if scores else "-")
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = border
            if cell.column > 3:
                cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 16
    from openpyxl.utils import get_column_letter
    for idx in range(4, 4 + len(sessions) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"rekap-nilai-{cls['name']}.xlsx".replace(" ", "_")
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


# ------------------------------------------------------------------ DASHBOARD
@api_router.get("/dashboard/stats")
async def dashboard_stats(user: dict = Depends(require_roles("admin", "guru"))):
    students = await db.users.count_documents({"role": "siswa"})
    teachers = await db.users.count_documents({"role": "guru"})
    questions = await db.questions.count_documents({})
    packages = await db.packages.count_documents({})
    sessions = await db.sessions.count_documents({})
    attempts = await db.attempts.find({"status": "selesai"}, {"_id": 0, "score": 1}).to_list(5000)
    scores = [a["score"] for a in attempts if a.get("score") is not None]
    avg = round(sum(scores) / len(scores), 1) if scores else 0
    pending = await db.attempts.count_documents({"status": "menunggu_koreksi"})
    return {"students": students, "teachers": teachers, "questions": questions,
            "packages": packages, "sessions": sessions, "avg_score": avg,
            "completed_attempts": len(scores), "pending_grading": pending}


# ------------------------------------------------------------------ CLASS ANALYTICS
@api_router.get("/analytics/classes")
async def analytics_classes(user: dict = Depends(require_roles("admin", "guru"))):
    classes = await db.classes.find({}, {"_id": 0}).to_list(1000)
    result = []
    for c in classes:
        sids = c.get("student_ids", [])
        avg = 0
        completed = 0
        if sids:
            attempts = await db.attempts.find(
                {"student_id": {"$in": sids}, "status": "selesai"}, {"_id": 0, "score": 1}).to_list(5000)
            scores = [a["score"] for a in attempts if a.get("score") is not None]
            avg = round(sum(scores) / len(scores), 1) if scores else 0
            completed = len(scores)
        result.append({"class_id": c["id"], "name": c["name"], "avg_score": avg,
                       "completed": completed, "students": len(sids)})
    sessions = await db.sessions.find({}, {"_id": 0}).sort("start_time", 1).to_list(1000)
    trend = []
    for s in sessions:
        atts = await db.attempts.find({"session_id": s["id"], "status": "selesai"}, {"_id": 0, "score": 1}).to_list(5000)
        sc = [a["score"] for a in atts if a.get("score") is not None]
        if sc:
            trend.append({"session": s["title"][:18], "avg": round(sum(sc) / len(sc), 1)})
    return {"classes": result, "trend": trend}


# ------------------------------------------------------------------ SUBJECT STATS
@api_router.get("/analytics/subjects")
async def analytics_subjects(user: dict = Depends(require_roles("admin", "guru"))):
    cats = await db.categories.find({}, {"_id": 0}).to_list(1000)
    pkgs = await db.packages.find({}, {"_id": 0, "id": 1, "category_id": 1}).to_list(2000)
    pkg_cat = {p["id"]: p.get("category_id") for p in pkgs}
    atts = await db.attempts.find({"status": "selesai"}, {"_id": 0, "score": 1, "package_id": 1}).to_list(20000)
    agg = {}
    for a in atts:
        if a.get("score") is None:
            continue
        agg.setdefault(pkg_cat.get(a.get("package_id")), []).append(a["score"])
    result = []
    for c in cats:
        sc = agg.get(c["id"], [])
        result.append({"category_id": c["id"], "name": c["name"],
                       "avg_score": round(sum(sc) / len(sc), 1) if sc else 0, "attempts": len(sc)})
    uncat = agg.get(None, [])
    if uncat:
        result.append({"category_id": None, "name": "Umum",
                       "avg_score": round(sum(uncat) / len(uncat), 1), "attempts": len(uncat)})
    result.sort(key=lambda r: -r["avg_score"])
    return result


# ------------------------------------------------------------------ NOTIFICATIONS
@api_router.get("/notifications")
async def notifications(user: dict = Depends(require_roles("siswa"))):
    my_classes = await db.classes.find({"student_ids": user["id"]}, {"_id": 0, "id": 1}).to_list(1000)
    my_class_ids = {c["id"] for c in my_classes}
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(1000)
    my_makeups = await db.makeups.find({"student_id": user["id"]}, {"_id": 0}).to_list(500)
    mmap = {m["session_id"]: m for m in my_makeups}
    now = datetime.now(timezone.utc)
    notes = []
    for s in sessions:
        targets = s.get("class_ids") or []
        mk = mmap.get(s["id"])
        if targets and not (set(targets) & my_class_ids) and not mk:
            continue
        try:
            start = datetime.fromisoformat(s["start_time"])
            end = datetime.fromisoformat(s["end_time"])
        except Exception:
            continue
        att = await db.attempts.find_one({"session_id": s["id"], "student_id": user["id"]},
                                        {"_id": 0, "status": 1, "makeup_id": 1})
        done = att and att["status"] != "berlangsung"
        if mk and done and att.get("makeup_id") != mk["id"]:
            done = False  # masih punya hak mengerjakan lewat susulan
        if s.get("announcement"):
            notes.append({"id": f"{s['id']}-ann", "type": "info", "title": s["title"],
                          "message": s["announcement"], "time": s["start_time"]})
        if not done:
            if mk:
                mk_status = window_status(mk["start_time"], mk["end_time"], now)
                alasan = f" Alasan: {mk['reason']}." if mk.get("reason") else ""
                if mk_status == "akan_datang":
                    notes.append({"id": f"{mk['id']}-makeup-open", "type": "upcoming", "title": s["title"],
                                  "message": f"Ujian susulan dijadwalkan untuk Anda.{alasan}",
                                  "time": mk["start_time"]})
                elif mk_status == "berlangsung":
                    notes.append({"id": f"{mk['id']}-makeup-live", "type": "live", "title": s["title"],
                                  "message": f"Ujian susulan Anda sedang dibuka — segera kerjakan.{alasan}",
                                  "time": mk["start_time"]})
            if now < start:
                notes.append({"id": f"{s['id']}-open", "type": "upcoming", "title": s["title"],
                              "message": "Ujian akan segera dibuka.", "time": s["start_time"]})
            elif now <= end:
                notes.append({"id": f"{s['id']}-live", "type": "live", "title": s["title"],
                              "message": "Ujian sedang berlangsung — segera kerjakan.", "time": s["start_time"]})
    notes.sort(key=lambda n: n["time"], reverse=True)
    return notes


# ------------------------------------------------------------------ SCHOOL SETTINGS
@api_router.get("/settings/school")
async def get_school():
    """Public: school identity/theme is needed to brand the login screen."""
    doc = await db.settings.find_one({"key": "school"}, {"_id": 0}) or {}
    return {"name": doc.get("name", ""), "address": doc.get("address", ""),
            "logo_path": doc.get("logo_path"),
            "theme_color": sanitize_theme_color(doc.get("theme_color")) or DEFAULT_THEME_COLOR}


@api_router.put("/settings/school")
async def set_school(body: SchoolBody, user: dict = Depends(require_roles("admin"))):
    payload = body.model_dump()
    payload["theme_color"] = sanitize_theme_color(payload.get("theme_color")) or DEFAULT_THEME_COLOR
    await db.settings.update_one({"key": "school"}, {"$set": {"key": "school", **payload}}, upsert=True)
    return payload


async def _logo_flowable(logo_path):
    from reportlab.platypus import Image
    from reportlab.lib.units import mm
    try:
        if not logo_path:
            return None
        data, _ = await load_file(logo_path)
        return Image(io.BytesIO(data), width=18 * mm, height=18 * mm)
    except Exception:
        return None


async def _school_kop(styles, green, sub):
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import ParagraphStyle
    school = await db.settings.find_one({"key": "school"}, {"_id": 0}) or {}
    if not (school.get("name") or school.get("logo_path")):
        return []
    txt = []
    if school.get("name"):
        txt.append(Paragraph(school["name"], ParagraphStyle("sn", parent=styles["Title"], textColor=green, fontSize=15, spaceAfter=0)))
    if school.get("address"):
        txt.append(Paragraph(school["address"], sub))
    logo = await _logo_flowable(school.get("logo_path"))
    if logo:
        t = Table([[logo, txt]], colWidths=[22 * 2.83, None])
        t.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LEFTPADDING", (0, 0), (0, 0), 0)]))
        head = [t]
    else:
        head = txt
    from reportlab.lib import colors
    line = Table([[""]], colWidths=[520 * 0.35])
    return head + [Spacer(1, 6)]

# ------------------------------------------------------------------ EXAM LOCK (mode ujian ketat)
VIOLATION_LABEL = {
    "tab_hidden": "Pindah tab / minimize",
    "window_blur": "Keluar dari jendela ujian",
    "fullscreen_exit": "Keluar dari layar penuh",
    "copy_attempt": "Mencoba menyalin teks",
    "shortcut_blocked": "Menekan tombol terlarang",
    "reload_attempt": "Mencoba memuat ulang halaman",
}
EXAM_LOCK_DEFAULT = {"enabled": True, "max_violations": 3}


class ExamLockBody(BaseModel):
    enabled: bool = True
    max_violations: int = 3


class ViolationBody(BaseModel):
    session_id: str
    type: str = "tab_hidden"


async def get_exam_lock() -> dict:
    doc = await db.settings.find_one({"key": "exam_lock"}, {"_id": 0}) or {}
    return {"enabled": bool(doc.get("enabled", True)),
            "max_violations": int(doc.get("max_violations", 3))}


@api_router.get("/settings/exam-lock")
async def read_exam_lock(user: dict = Depends(get_current_user)):
    """Readable by students too — the exam screen needs the violation limit."""
    return await get_exam_lock()


@api_router.put("/settings/exam-lock")
async def set_exam_lock(body: ExamLockBody, user: dict = Depends(require_roles("admin", "guru"))):
    n = max(1, min(20, int(body.max_violations)))
    doc = {"key": "exam_lock", "enabled": bool(body.enabled), "max_violations": n}
    await db.settings.update_one({"key": "exam_lock"}, {"$set": doc}, upsert=True)
    return {"enabled": doc["enabled"], "max_violations": n}


@api_router.post("/exam/violation")
async def record_violation(body: ViolationBody, user: dict = Depends(require_roles("siswa"))):
    """Log a lockdown violation. Auto-submits the attempt once the limit is reached."""
    attempt = await db.attempts.find_one(
        {"session_id": body.session_id, "student_id": user["id"]}, {"_id": 0})
    if not attempt:
        raise HTTPException(status_code=404, detail="Percobaan tidak ditemukan")
    cfg = await get_exam_lock()
    if attempt["status"] != "berlangsung":
        return {"count": len(attempt.get("violations", [])),
                "max_violations": cfg["max_violations"],
                "auto_submitted": True, "already_submitted": True}

    vtype = body.type if body.type in VIOLATION_LABEL else "tab_hidden"
    entry = {"type": vtype, "label": VIOLATION_LABEL[vtype], "at": now_iso()}
    await db.attempts.update_one({"id": attempt["id"]}, {"$push": {"violations": entry}})
    count = len(attempt.get("violations", [])) + 1

    auto = False
    if cfg["enabled"] and count >= cfg["max_violations"]:
        fresh = await db.attempts.find_one({"id": attempt["id"]}, {"_id": 0})
        await finalize_attempt(fresh, fresh.get("answers", {}))
        await db.attempts.update_one({"id": attempt["id"]},
                                     {"$set": {"auto_submitted_reason": "pelanggaran"}})
        auto = True
        logger.info(f"Exam auto-submitted for {user['email']} after {count} violation(s)")
    return {"count": count, "max_violations": cfg["max_violations"],
            "auto_submitted": auto, "label": entry["label"]}




# ------------------------------------------------------------------ DIFFICULTY SETTINGS
@api_router.get("/settings/difficulty")
async def get_difficulty(user: dict = Depends(require_roles("admin", "guru"))):
    doc = await db.settings.find_one({"key": "difficulty"}, {"_id": 0})
    if not doc:
        return {"easy_min": 70, "medium_min": 40}
    return {"easy_min": doc.get("easy_min", 70), "medium_min": doc.get("medium_min", 40)}


@api_router.put("/settings/difficulty")
async def set_difficulty(body: DifficultyBody, user: dict = Depends(require_roles("admin", "guru"))):
    easy = max(1.0, min(100.0, float(body.easy_min)))
    medium = max(0.0, min(99.0, float(body.medium_min)))
    if medium >= easy:
        raise HTTPException(status_code=400, detail="Ambang 'Sedang' harus lebih kecil dari 'Mudah'")
    await db.settings.update_one(
        {"key": "difficulty"},
        {"$set": {"key": "difficulty", "easy_min": easy, "medium_min": medium}},
        upsert=True)
    return {"easy_min": easy, "medium_min": medium}


# ------------------------------------------------------------------ LEADERBOARD
async def compute_class_leaderboard(cls: dict, category_id: str = None) -> list:
    sids = cls.get("student_ids", [])
    if not sids:
        return []
    pkg_ids = None
    if category_id:
        pkgs = await db.packages.find({"category_id": category_id}, {"_id": 0, "id": 1}).to_list(2000)
        pkg_ids = {p["id"] for p in pkgs}
    docs = await db.users.find({"_id": {"$in": [ObjectId(s) for s in sids]}}).to_list(2000)
    rows = []
    for u in docs:
        uid = str(u["_id"])
        atts = await db.attempts.find({"student_id": uid, "status": "selesai"},
                                      {"_id": 0, "score": 1, "package_id": 1}).to_list(5000)
        scores = [a["score"] for a in atts
                  if a.get("score") is not None and (pkg_ids is None or a.get("package_id") in pkg_ids)]
        rows.append({
            "student_id": uid, "name": u["name"], "identifier": u.get("identifier", ""),
            "avg_score": round(sum(scores) / len(scores), 1) if scores else 0,
            "completed": len(scores),
        })
    rows.sort(key=lambda r: (-r["avg_score"], r["name"].lower()))
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    return rows


@api_router.get("/leaderboard/class/{class_id}")
async def leaderboard_class(class_id: str, user: dict = Depends(require_roles("admin", "guru"))):
    cls = await db.classes.find_one({"id": class_id}, {"_id": 0})
    if not cls:
        raise HTTPException(status_code=404, detail="Kelas tidak ditemukan")
    return {"class_name": cls["name"], "rows": await compute_class_leaderboard(cls)}


@api_router.get("/leaderboard/me")
async def leaderboard_me(category_id: Optional[str] = None, user: dict = Depends(require_roles("siswa"))):
    classes = await db.classes.find({"student_ids": user["id"]}, {"_id": 0}).to_list(1000)
    out = []
    for c in classes:
        out.append({"class_id": c["id"], "class_name": c["name"],
                    "rows": await compute_class_leaderboard(c, category_id)})
    return out


async def compute_global_leaderboard(start=None, end=None, category_id=None):
    students = await db.users.find({"role": "siswa"}).to_list(5000)
    classes = await db.classes.find({}, {"_id": 0}).to_list(1000)
    cls_by_student = {}
    for c in classes:
        for sid in c.get("student_ids", []):
            cls_by_student.setdefault(sid, []).append(c["name"])
    pkg_ids = None
    if category_id:
        pkgs = await db.packages.find({"category_id": category_id}, {"_id": 0, "id": 1}).to_list(2000)
        pkg_ids = {p["id"] for p in pkgs}
    rows = []
    for u in students:
        uid = str(u["_id"])
        atts = await db.attempts.find(
            {"student_id": uid, "status": "selesai"},
            {"_id": 0, "score": 1, "submitted_at": 1, "package_id": 1}).to_list(5000)
        scores = []
        for a in atts:
            if a.get("score") is None:
                continue
            if pkg_ids is not None and a.get("package_id") not in pkg_ids:
                continue
            st = (a.get("submitted_at") or "")[:10]
            if start and (not st or st < start):
                continue
            if end and (not st or st > end):
                continue
            scores.append(a["score"])
        rows.append({
            "student_id": uid, "name": u["name"], "identifier": u.get("identifier", ""),
            "classes": cls_by_student.get(uid, []),
            "avg_score": round(sum(scores) / len(scores), 1) if scores else 0,
            "completed": len(scores),
        })
    rows.sort(key=lambda r: (-r["avg_score"], r["name"].lower()))
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    return rows


@api_router.get("/leaderboard/global")
async def leaderboard_global(start: Optional[str] = None, end: Optional[str] = None,
                             category_id: Optional[str] = None,
                             user: dict = Depends(get_current_user)):
    rows = await compute_global_leaderboard(start, end, category_id)
    return {"rows": rows}


@api_router.get("/export/leaderboard/xlsx")
async def export_leaderboard(start: Optional[str] = None, end: Optional[str] = None,
                             category_id: Optional[str] = None,
                             user: dict = Depends(require_roles("admin", "guru"))):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    rows = await compute_global_leaderboard(start, end, category_id)
    cat_name = "Semua Mapel"
    if category_id:
        cat = await db.categories.find_one({"id": category_id}, {"_id": 0, "name": 1})
        cat_name = cat["name"] if cat else category_id
    period = f"{start or '...'} s/d {end or '...'}" if (start or end) else "Semua waktu"

    wb = Workbook()
    ws = wb.active
    ws.title = "Peringkat Angkatan"
    ws.append(["PERINGKAT ANGKATAN"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws["A1"].font = Font(bold=True, size=14, color="1E3A30")
    ws.append([f"Mapel: {cat_name}", "", f"Periode: {period}"])
    ws.append([])
    header = ["Peringkat", "Nama Siswa", "NISN/NIP", "Kelas", "Ujian Selesai", "Rata-rata"]
    ws.append(header)
    green = PatternFill("solid", fgColor="1E3A30")
    thin = Side(style="thin", color="D9D9CF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hrow = ws.max_row
    for cell in ws[hrow]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = green
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for r in rows:
        ws.append([r["rank"], r["name"], r.get("identifier", ""), ", ".join(r.get("classes", [])),
                   r["completed"], r["avg_score"]])
        for cell in ws[ws.max_row]:
            cell.border = border
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=peringkat-angkatan.xlsx"})


# ------------------------------------------------------------------ BACKUP / RESTORE
BACKUP_COLLECTIONS = ["users", "classes", "categories", "questions", "packages",
                      "sessions", "attempts", "settings", "files", "file_blobs"]
BACKUP_VERSION = 1


def _encode_doc(doc: dict) -> dict:
    """Make a Mongo document JSON-safe (ObjectId -> str, Binary -> base64)."""
    from bson.binary import Binary
    out = {}
    for k, v in doc.items():
        if k == "_id":
            out["_id"] = str(v)
        elif isinstance(v, (Binary, bytes)):
            out[k] = {"__b64__": base64.b64encode(bytes(v)).decode("ascii")}
        elif isinstance(v, datetime):
            out[k] = v.isoformat()
        elif isinstance(v, dict):
            out[k] = _encode_doc(v)
        elif isinstance(v, list):
            out[k] = [_encode_doc(x) if isinstance(x, dict) else x for x in v]
        else:
            out[k] = v
    return out


def _decode_doc(doc: dict) -> dict:
    from bson.binary import Binary
    out = {}
    for k, v in doc.items():
        if k == "_id":
            try:
                out["_id"] = ObjectId(v)
            except Exception:
                out["_id"] = v
        elif isinstance(v, dict) and "__b64__" in v:
            out[k] = Binary(base64.b64decode(v["__b64__"]))
        elif isinstance(v, dict):
            out[k] = _decode_doc(v)
        elif isinstance(v, list):
            out[k] = [_decode_doc(x) if isinstance(x, dict) else x for x in v]
        else:
            out[k] = v
    return out


@api_router.get("/backup/stats")
async def backup_stats(user: dict = Depends(require_roles("admin"))):
    """How much data a backup would contain (shown before downloading)."""
    counts = {}
    for name in BACKUP_COLLECTIONS:
        counts[name] = await db[name].count_documents({})
    blob_size = 0
    async for b in db.file_blobs.find({}, {"size": 1}):
        blob_size += int(b.get("size") or 0)
    return {"counts": counts, "files_bytes": blob_size,
            "students": await db.users.count_documents({"role": "siswa"}),
            "teachers": await db.users.count_documents({"role": "guru"})}


@api_router.get("/backup/export")
async def backup_export(user: dict = Depends(require_roles("admin"))):
    """Download the whole school database (accounts, questions, results, images)."""
    payload = {"version": BACKUP_VERSION, "app": APP_NAME,
               "exported_at": now_iso(), "collections": {}}
    for name in BACKUP_COLLECTIONS:
        docs = await db[name].find({}).to_list(200000)
        payload["collections"][name] = [_encode_doc(d) for d in docs]
    raw = json.dumps(payload, ensure_ascii=False, default=str).encode("utf-8")
    gz = gzip.compress(raw, compresslevel=6)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")
    return StreamingResponse(
        io.BytesIO(gz), media_type="application/gzip",
        headers={"Content-Disposition": f'attachment; filename="backup-cbt-{stamp}.json.gz"'})


@api_router.post("/backup/import")
async def backup_import(file: UploadFile = File(...), mode: str = Form("merge"),
                        user: dict = Depends(require_roles("admin"))):
    """Restore a backup file.

    mode=merge   -> insert/overwrite documents from the file, keep everything else
    mode=replace -> wipe the collections in the file first (full restore)
    """
    mode = (mode or "merge").lower()
    if mode not in ("merge", "replace"):
        raise HTTPException(status_code=400, detail="Mode harus 'merge' atau 'replace'")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="File kosong")
    try:
        if raw[:2] == b"\x1f\x8b":
            raw = gzip.decompress(raw)
        payload = json.loads(raw.decode("utf-8"))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"File backup tidak valid: {e}")

    cols = payload.get("collections")
    if not isinstance(cols, dict) or not cols:
        raise HTTPException(status_code=400, detail="File backup tidak memuat data")
    unknown = [k for k in cols if k not in BACKUP_COLLECTIONS]
    if unknown:
        raise HTTPException(status_code=400,
                            detail=f"File backup memuat data tak dikenal: {', '.join(unknown)}")

    report = {}
    for name, docs in cols.items():
        if mode == "replace":
            await db[name].delete_many({})
        inserted = updated = 0
        for d in docs:
            doc = _decode_doc(d)
            key = None
            if "_id" in doc:
                key = {"_id": doc["_id"]}
            elif "id" in doc:
                key = {"id": doc["id"]}
            elif name == "settings" and "key" in doc:
                key = {"key": doc["key"]}
            elif name == "file_blobs" and "path" in doc:
                key = {"path": doc["path"]}
            if key is None:
                await db[name].insert_one(doc)
                inserted += 1
                continue
            existing = await db[name].find_one(key, {"_id": 1})
            body = {k: v for k, v in doc.items() if k != "_id"}
            if existing is None:
                await db[name].insert_one(doc)
                inserted += 1
            else:
                await db[name].update_one(key, {"$set": body})
                updated += 1
        report[name] = {"inserted": inserted, "updated": updated}

    return {"mode": mode, "version": payload.get("version"),
            "exported_at": payload.get("exported_at"), "result": report}


# ------------------------------------------------------------------ startup
# Self-hosted deployments have no Emergent cron, so an in-process loop can
# finalise expired exam attempts. Enable with INTERNAL_CRON_MINUTES=5 (0 = off).
INTERNAL_CRON_MINUTES = int(os.environ.get("INTERNAL_CRON_MINUTES", "0") or 0)


async def _internal_cron_loop():
    log = logging.getLogger(__name__)
    log.info("Internal auto-submit scheduler every %s minute(s)", INTERNAL_CRON_MINUTES)
    while True:
        try:
            await asyncio.sleep(INTERNAL_CRON_MINUTES * 60)
            await run_auto_submit()
        except asyncio.CancelledError:
            raise
        except Exception as e:
            log.error("Internal auto-submit failed: %s", e)


@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.file_blobs.create_index("path", unique=True)
    if use_emergent_storage():
        try:
            init_storage()
            logger.info("Object storage initialized")
        except Exception as e:
            logging.getLogger(__name__).warning(
                f"Storage init failed ({e}) — uploads will be stored in MongoDB")
    else:
        logger.info("Object storage: MongoDB (self-hosted mode)")
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@example.com").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        await db.users.insert_one({
            "email": admin_email, "password_hash": hash_password(admin_password),
            "name": "Administrator", "role": "admin", "identifier": "",
            "created_at": now_iso(),
        })
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one({"email": admin_email},
                                  {"$set": {"password_hash": hash_password(admin_password)}})
    if INTERNAL_CRON_MINUTES > 0:
        asyncio.create_task(_internal_cron_loop())


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
